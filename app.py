"""
app.py
Flask 웹 대시보드 서버
접속: http://yaglobal.iptime.org:3000/jp_sourcing
"""

import asyncio
import json
import logging
import math
import os
import requests
import threading
from datetime import datetime

from apscheduler.schedulers.background import BackgroundScheduler
from functools import wraps
from flask import Flask, jsonify, render_template, request, Response, session, redirect, url_for, send_from_directory, make_response
import queue

from config import (
    SERVER_HOST, SERVER_PORT, URL_PREFIX,
    AUTO_SCHEDULE_HOUR, AUTO_SCHEDULE_MINUTE,
    LOGIN_USERS, SECRET_KEY, APP_ENV,
    OUTPUT_DIR, DB_DIR,
)
from data_manager import get_status as get_data_status, set_data_root, get_data_root, get_path, ensure_dirs, is_connected
from xebio_search import scrape_nike_sale, load_latest_products, set_app_status, force_close_browser
from cafe_uploader import upload_products, naver_manual_login, has_saved_cookies, delete_cookies, request_upload_stop, is_upload_stop_requested
from exchange import get_jpy_to_krw_rate, get_cached_rate, calc_buying_price, set_margin_rate, get_margin_rate, set_price_config, get_price_config
from user_db import init_db as init_user_db, get_user as get_customer, create_user, check_password as check_customer_pw, username_exists
from werkzeug.security import generate_password_hash, check_password_hash
from post_generator import get_ai_config, set_ai_config, verify_ai_key
from site_config import get_sites_for_ui
from scrape_history import get_history as get_scrape_history
from cafe_schedule import load_schedule, save_schedule, load_check_schedule, save_check_schedule, load_task_schedule, save_task_schedule
from product_db import init_db as init_product_db, get_stats as bigdata_get_stats, search_products as bigdata_search, get_brands as bigdata_get_brands, delete_all as bigdata_delete_all, delete_by_site as bigdata_delete_site, delete_by_ids as bigdata_delete_ids, get_total_count as bigdata_total, export_all as bigdata_export, export_csv as bigdata_export_csv, merge_products as bigdata_merge
from cafe_monitor import start_monitor, stop_monitor, is_monitoring, batch_check_cafe_duplicates
from telegram_bot import start_bot, stop_bot, is_bot_running

# =============================================
# 앱 초기화
# =============================================

APP_VERSION = "자동작업 스케줄"
app = Flask(__name__, template_folder="templates", static_folder="static")
app.secret_key = SECRET_KEY
app.config["TEMPLATES_AUTO_RELOAD"] = True   # 템플릿 변경 즉시 반영
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)


@app.after_request
def add_no_cache(response):
    """브라우저/프록시 캐시 방지 — HTML + JSON 모두 적용"""
    if "text/html" in response.content_type or "application/json" in response.content_type:
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


# =============================================
# 로그인 인증
# =============================================

def login_required(f):
    """로그인 필수 데코레이터 (모든 인증된 사용자)"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            # API 요청이면 JSON 에러 반환 (프론트 SyntaxError 방지)
            if "/api/" in request.path or request.is_json:
                return jsonify({"ok": False, "error": "로그인이 필요합니다"}), 401
            return redirect(f"{URL_PREFIX}/login")
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """관리자 전용 데코레이터"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            if "/api/" in request.path or "/scrape/" in request.path or "/orders" in request.path or request.is_json or request.headers.get("Accept","").startswith("application/json"):
                return jsonify({"ok": False, "error": "로그인이 필요합니다"}), 401
            return redirect(f"{URL_PREFIX}/login")
        # 기존 세션(role 없음)은 admin으로 간주
        if session.get("role", "admin") != "admin":
            return redirect(f"{URL_PREFIX}/shop")
        return f(*args, **kwargs)
    return decorated


# ── 전화번호 정규화 ──────────────────────────
def _normalize_phone(phone: str) -> str:
    """전화번호에서 숫자만 추출 (010-1234-5678 → 01012345678)"""
    if not phone:
        return ""
    import re
    return re.sub(r"[^0-9]", "", phone)


def _format_phone(phone: str) -> str:
    """전화번호 표시용 포맷 (01012345678 → 010-1234-5678)"""
    if not phone:
        return ""
    p = _normalize_phone(phone)
    if len(p) == 11 and p.startswith("010"):
        return f"{p[:3]}-{p[3:7]}-{p[7:]}"
    elif len(p) == 10:
        return f"{p[:3]}-{p[3:6]}-{p[6:]}"
    return p


def _find_user_by_phone(phone: str):
    """전화번호로 기존 회원 찾기 (정규화된 번호 비교)"""
    if not phone:
        return None
    normalized = _normalize_phone(phone)
    if not normalized:
        return None
    from user_db import _conn
    conn = _conn()
    try:
        rows = conn.execute("SELECT * FROM users").fetchall()
        for r in rows:
            db_phone = _normalize_phone(r["phone"] or "")
            if db_phone and db_phone == normalized:
                return {c: r[c] for c in r.keys()}
        return None
    finally:
        conn.close()


# ── 네이버 소셜 로그인 ──────────────────────────
NAVER_CLIENT_ID = "CH3HgXly53mIV7WYrg_c"
NAVER_CLIENT_SECRET = "yPrHZRAHNH"
NAVER_CALLBACK_URL = "https://vintage.theone-biz.com/auth/naver/callback"


@app.route(f"{URL_PREFIX}/auth/naver")
def naver_login():
    """네이버 로그인 시작"""
    import urllib.parse, uuid
    state = uuid.uuid4().hex[:16]
    session["naver_state"] = state
    callback = NAVER_CALLBACK_URL
    params = urllib.parse.urlencode({
        "response_type": "code",
        "client_id": NAVER_CLIENT_ID,
        "redirect_uri": callback,
        "state": state,
    })
    return redirect(f"https://nid.naver.com/oauth2.0/authorize?{params}")


@app.route(f"{URL_PREFIX}/auth/naver/callback")
def naver_callback():
    """네이버 로그인 콜백"""
    code = request.args.get("code", "")
    state = request.args.get("state", "")
    error = request.args.get("error", "")
    if error:
        logger.warning(f"네이버 로그인 거부: {error} - {request.args.get('error_description','')}")
        return redirect(f"{URL_PREFIX}/login")
    if not code:
        return redirect(f"{URL_PREFIX}/login")
    # state 검증 (세션 유실 시에도 진행 허용)
    saved_state = session.pop("naver_state", "")
    if saved_state and state != saved_state:
        logger.warning(f"네이버 state 불일치: {state} != {saved_state}")
        return redirect(f"{URL_PREFIX}/login")

    callback = NAVER_CALLBACK_URL
    logger.info(f"🔵 네이버 콜백 수신: code={code[:10]}... state={state}")
    # 토큰 발급
    try:
        token_res = requests.post("https://nid.naver.com/oauth2.0/token", data={
            "grant_type": "authorization_code",
            "client_id": NAVER_CLIENT_ID,
            "client_secret": NAVER_CLIENT_SECRET,
            "code": code,
            "state": state,
            "redirect_uri": callback,
        }, timeout=10)
        token = token_res.json()
        logger.info(f"🔵 네이버 토큰 응답: {token}")
        access_token = token.get("access_token")
        if not access_token:
            logger.warning(f"네이버 토큰 실패: {token}")
            return redirect(f"{URL_PREFIX}/login")

        # 프로필 조회
        profile_res = requests.get("https://openapi.naver.com/v1/nid/me",
                                   headers={"Authorization": f"Bearer {access_token}"}, timeout=10)
        profile_data = profile_res.json()
        logger.info(f"🔵 네이버 프로필 응답: {profile_data}")
        profile = profile_data.get("response", {})
        naver_id = profile.get("id", "")
        name = profile.get("name", "") or profile.get("nickname", "")
        email = profile.get("email", "")
        phone = _normalize_phone(profile.get("mobile", ""))

        if not naver_id:
            return redirect(f"{URL_PREFIX}/login")

        # 1) 네이버 ID로 기존 소셜 회원 확인
        social_username = f"naver_{naver_id[:12]}"
        customer = get_customer(social_username)

        # 2) 소셜 계정이 없으면 → 전화번호로 기존 일반 회원 찾기 (통합)
        login_username = social_username
        if not customer and phone:
            existing = _find_user_by_phone(phone)
            if existing:
                # 기존 회원 계정으로 통합 로그인
                login_username = existing["username"]
                customer = existing
                # 네이버 ID를 기존 계정에 연결 (naver_id 컬럼 저장)
                try:
                    from user_db import _conn as _uc
                    conn = _uc()
                    conn.execute("UPDATE users SET naver_id=?, email=COALESCE(NULLIF(email,''),?) WHERE username=?",
                                 (naver_id, email, login_username))
                    conn.commit()
                    conn.close()
                except Exception:
                    pass
                logger.info(f"네이버 로그인 → 기존 회원 통합: {login_username} (네이버: {social_username})")

        # 3) 둘 다 없으면 → 신규 소셜 회원가입
        if not customer:
            try:
                from user_db import _conn as _uc
                conn = _uc()
                conn.execute("""INSERT OR IGNORE INTO users (username, password_hash, name, email, phone, status, level, naver_id)
                                VALUES (?,?,?,?,?,?,?,?)""",
                             (social_username, "", name, email, phone, "approved", "b2c", naver_id))
                conn.commit()
                conn.close()
                login_username = social_username
                logger.info(f"네이버 소셜 회원가입: {social_username} ({name})")
                try:
                    from notifier import send_telegram
                    send_telegram(f"👤 <b>네이버 소셜 회원가입</b>\n이름: {name}\n아이디: {social_username}\n📞 {_format_phone(phone)}")
                except Exception:
                    pass
            except Exception as e:
                logger.warning(f"네이버 회원가입 실패: {e}")
                return redirect(f"{URL_PREFIX}/login")
            customer = get_customer(login_username)

        # 로그인 처리
        session["logged_in"] = True
        session["username"] = login_username
        session["role"] = "customer"
        session["level"] = customer["level"] if customer and "level" in customer.keys() else "b2c"
        session["name"] = customer.get("name", "") or name if customer else name
        # 마지막 접속 시간 업데이트
        try:
            from user_db import _conn as _uc
            uc = _uc()
            uc.execute("UPDATE users SET last_login=datetime('now','localtime') WHERE username=?", (login_username,))
            uc.commit()
            uc.close()
        except Exception:
            pass
        logger.info(f"네이버 소셜 로그인: {login_username} ({name})")
        return redirect(f"{URL_PREFIX}/shop")
    except Exception as e:
        logger.error(f"네이버 로그인 오류: {e}")
        return redirect(f"{URL_PREFIX}/login")


@app.route(f"{URL_PREFIX}/login", methods=["GET", "POST"])
def login():
    """로그인 페이지"""
    if session.get("logged_in"):
        if session.get("role", "admin") == "admin":
            return redirect(f"{URL_PREFIX}/dashboard")
        return redirect(f"{URL_PREFIX}/shop")

    error = None
    username = ""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        # 1) 관리자 확인
        if username in LOGIN_USERS and LOGIN_USERS[username] == password:
            session["logged_in"] = True
            session["username"] = username
            session["role"] = "admin"
            session["level"] = "b2b"
            logger.info(f"관리자 로그인: {username}")
            return redirect(f"{URL_PREFIX}/dashboard")
        # 2) 고객 확인
        customer = get_customer(username)
        if customer and check_customer_pw(customer, password):
            cust_status = customer["status"] if "status" in customer.keys() else "approved"
            expires_at = customer["expires_at"] if "expires_at" in customer.keys() else ""
            if cust_status == "pending":
                error = "가입 승인 대기 중입니다. 관리자 승인 후 이용 가능합니다."
                logger.info(f"승인 대기 로그인 시도: {username}")
            elif cust_status == "suspended":
                error = "계정이 사용 불가 상태입니다. 관리자에게 문의해주세요."
                logger.info(f"사용불가 로그인 시도: {username}")
            elif cust_status == "rejected":
                error = "가입이 거절되었습니다. 관리자에게 문의해주세요."
            elif expires_at and expires_at < datetime.now().strftime("%Y-%m-%d"):
                error = f"사용 기간이 만료되었습니다. (만료일: {expires_at}) 관리자에게 문의해주세요."
                logger.info(f"기간 만료 로그인 시도: {username} (만료: {expires_at})")
            else:
                session["logged_in"] = True
                session["username"] = username
                session["role"] = "customer"
                session["level"] = customer["level"] if "level" in customer.keys() else "b2c"
                session["name"] = customer["name"] if "name" in customer.keys() else ""
                # 마지막 접속 시간 업데이트
                try:
                    from user_db import _conn as _uc
                    uc = _uc()
                    uc.execute("UPDATE users SET last_login=datetime('now','localtime') WHERE username=?", (username,))
                    uc.commit()
                    uc.close()
                except Exception:
                    pass
                logger.info(f"고객 로그인: {username} (level={session['level']})")
                return redirect(f"{URL_PREFIX}/shop")
        elif customer:
            error = "비밀번호가 올바르지 않습니다"
        else:
            error = "아이디 또는 비밀번호가 올바르지 않습니다"
        logger.warning(f"로그인 실패: {username}")

    return render_template("login.html",
                           error=error, username=username,
                           url_prefix=URL_PREFIX, env=APP_ENV)


@app.route(f"{URL_PREFIX}/signup", methods=["GET", "POST"])
def signup():
    """회원가입 페이지"""
    if session.get("logged_in"):
        if session.get("role", "admin") == "admin":
            return redirect(f"{URL_PREFIX}/dashboard")
        return redirect(f"{URL_PREFIX}/shop")

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        email = request.form.get("email", "").strip()
        name = request.form.get("name", "").strip()
        phone = _normalize_phone(request.form.get("phone", ""))
        if not username or not password:
            error = "아이디와 비밀번호는 필수입니다"
        elif len(password) < 4:
            error = "비밀번호는 4자 이상이어야 합니다"
        elif username in LOGIN_USERS:
            error = "사용할 수 없는 아이디입니다"
        elif username_exists(username):
            error = "이미 존재하는 아이디입니다"
        elif phone and _find_user_by_phone(phone):
            existing = _find_user_by_phone(phone)
            error = f"이미 가입된 전화번호입니다 (아이디: {existing['username'][:3]}***)"
        else:
            if create_user(username, password, name, phone):
                # 추가 정보 저장 (배송/통관)
                try:
                    from user_db import _conn as user_conn
                    conn = user_conn()
                    conn.execute("""UPDATE users SET email=?, postal_code=?, address=?, address_detail=?, customs_id=?
                                   WHERE username=?""", (
                        email,
                        request.form.get("postal_code", "").strip(),
                        request.form.get("address", "").strip(),
                        request.form.get("address_detail", "").strip(),
                        request.form.get("customs_id", "").strip(),
                        username,
                    ))
                    conn.commit()
                    # 사업자등록증 첨부
                    cert_file = request.files.get("biz_cert")
                    if cert_file and cert_file.filename:
                        ext = os.path.splitext(cert_file.filename)[1].lower()
                        if ext in (".jpg", ".jpeg", ".png", ".pdf"):
                            upload_dir = os.path.join(get_path("db"), "certs")
                            os.makedirs(upload_dir, exist_ok=True)
                            cert_filename = f"{username}_cert{ext}"
                            cert_file.save(os.path.join(upload_dir, cert_filename))
                            conn.execute("UPDATE users SET business_cert_file=? WHERE username=?", (cert_filename, username))
                            conn.commit()
                    conn.close()
                except Exception:
                    pass
                logger.info(f"회원가입 (B2C 자동승인): {username}")
                # 텔레그램 알림
                try:
                    from notifier import send_telegram
                    send_telegram(
                        f"👤 <b>새 회원가입 알림</b>\n"
                        f"🆔 아이디: {username}\n"
                        f"📛 이름: {name or '-'}\n"
                        f"📞 연락처: {phone or '-'}\n"
                        f"✅ B2C 자동 승인 완료"
                    )
                except Exception:
                    pass
                # 가입 환영 문자 발송
                if phone:
                    try:
                        from aligo_sms import send_sms, load_config
                        load_config()
                        msg = (
                            f"[TheOne Vintage] {name or username}님, 회원가입을 환영합니다!\n"
                            f"B2C 회원으로 승인되었습니다.\n"
                            f"B2B 승인은 요청/문의 게시판에서 신청해주세요.\n"
                            f"https://vintage.theone-biz.com"
                        )
                        send_sms(phone, msg, title="TheOne Vintage")
                    except Exception:
                        pass
                return render_template("signup.html", error=None, success=True,
                                       url_prefix=URL_PREFIX, env=APP_ENV)
            else:
                error = "회원가입 실패. 다시 시도해주세요."

    return render_template("signup.html", error=error, success=False,
                           url_prefix=URL_PREFIX, env=APP_ENV)


# ── 사업자 정보 설정 ──────────────────────────
_biz_info_path = os.path.join(get_path("db"), "biz_info.json")
# NAS 경로 접근 불가 시 로컬 폴백
if not os.path.exists(os.path.dirname(_biz_info_path)):
    _biz_info_path = os.path.join(os.path.dirname(__file__), "biz_info.json")

def _load_biz_info():
    import json as _json
    if os.path.exists(_biz_info_path):
        try:
            with open(_biz_info_path, "r", encoding="utf-8") as f:
                return _json.load(f)
        except Exception:
            pass
    return {"active": 1, "biz1": {}, "biz2": {}}

def _save_biz_info(data):
    import json as _json
    try:
        d = os.path.dirname(_biz_info_path)
        if d:
            os.makedirs(d, exist_ok=True)
    except Exception:
        pass
    with open(_biz_info_path, "w", encoding="utf-8") as f:
        _json.dump(data, f, ensure_ascii=False, indent=2)


@app.route(f"{URL_PREFIX}/settings/biz-info", methods=["GET"])
@admin_required
def get_biz_info():
    return jsonify({"ok": True, **_load_biz_info()})


@app.route(f"{URL_PREFIX}/settings/biz-info", methods=["POST"])
@admin_required
def update_biz_info():
    data = request.json or {}
    info = _load_biz_info()
    info["active"] = int(data.get("active", info.get("active", 1)))
    for key in ["biz1", "biz2"]:
        if key in data:
            info[key] = data[key]
    _save_biz_info(info)
    return jsonify({"ok": True, "message": "사업자 정보 저장 완료"})


@app.route(f"{URL_PREFIX}/settings/biz-info/active")
def get_active_biz_info():
    """활성 사업자 정보 (푸터용, 비로그인 접근 가능)"""
    info = _load_biz_info()
    active = info.get("active", 1)
    biz = info.get(f"biz{active}", {})
    return jsonify({"ok": True, **biz})


@app.route(f"{URL_PREFIX}/api/vintage-cafe-products")
@admin_required
def vintage_cafe_products():
    """빈티지 카페 업로드용 상품 목록 (B2C 가격, 전체)"""
    from product_db import _conn
    conn = _conn()
    try:
        rows = conn.execute("""
            SELECT id, site_id, product_code, internal_code, name, name_ko, brand, brand_ko,
                   price_jpy, img_url, link, condition_grade, cafe_status, cafe_uploaded_at,
                   description, description_ko, color, material
            FROM products WHERE source_type='vintage' AND in_stock=1
            ORDER BY created_at DESC
        """).fetchall()
        products = []
        brands_set = set()
        for r in rows:
            b2c = _calc_vintage_price(r["price_jpy"], "b2c")
            name = r["name_ko"] if r["name_ko"] and r["name_ko"] != r["name"] else r["name"]
            brand = r["brand"]
            brands_set.add(brand)
            products.append({
                "id": r["id"],
                "product_code": r["internal_code"] or r["product_code"],
                "name": name,
                "brand": brand,
                "price_jpy": r["price_jpy"],
                "price_krw": b2c,
                "img_url": r["img_url"],
                "link": r["link"],
                "condition_grade": r["condition_grade"] or "",
                "cafe_status": r["cafe_status"] or "대기",
                "cafe_uploaded_at": r["cafe_uploaded_at"] or "",
                "description": r["description_ko"] or r["description"] or "",
                "color": r["color"] or "",
                "material": r["material"] or "",
            })
        return jsonify({"ok": True, "products": products, "brands": sorted(brands_set), "total": len(products)})
    finally:
        conn.close()


@app.route("/favicon.ico")
def favicon():
    return send_from_directory("static", "favicon.ico", mimetype="image/x-icon")


@app.route("/googleaccc97bb8d10ca5d.html")
def google_verification():
    return send_from_directory("static", "googleaccc97bb8d10ca5d.html")


@app.route("/robots.txt")
def robots_txt():
    return Response(
        "User-agent: *\nAllow: /\nAllow: /shop\nAllow: /shop/api/notices\nAllow: /shop/api/reviews\nDisallow: /dashboard\nDisallow: /orders\nDisallow: /members\nDisallow: /scrape\nDisallow: /settings\nDisallow: /api/\nSitemap: https://vintage.theone-biz.com/sitemap.xml\n",
        mimetype="text/plain"
    )


@app.route("/sitemap.xml")
def sitemap_xml():
    xml = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml += '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    xml += '  <url><loc>https://vintage.theone-biz.com/shop</loc><changefreq>daily</changefreq><priority>1.0</priority></url>\n'
    xml += '  <url><loc>https://vintage.theone-biz.com/</loc><changefreq>daily</changefreq><priority>0.8</priority></url>\n'
    xml += '</urlset>'
    return Response(xml, mimetype="application/xml")


@app.route(f"{URL_PREFIX}/shop")
def shop():
    """고객용 빈티지 상품 카탈로그 (비회원도 접근 가능)"""
    logged_in = session.get("logged_in", False)
    user_level = session.get("level", "b2c") if logged_in else "guest"
    display_name = session.get("name", "") or session.get("username", "")
    return render_template("shop.html",
                           url_prefix=URL_PREFIX, env=APP_ENV,
                           username=display_name,
                           is_admin=session.get("role", "") == "admin",
                           user_level=user_level,
                           logged_in=logged_in)


@app.route(f"{URL_PREFIX}/shop/mypage")
@login_required
def shop_mypage():
    """고객용 마이페이지"""
    return render_template("mypage.html",
                           url_prefix=URL_PREFIX, env=APP_ENV,
                           username=session.get("username"),
                           is_admin=session.get("role", "admin") == "admin")


@app.route(f"{URL_PREFIX}/shop/my-orders")
@login_required
def shop_my_orders():
    return redirect(f"{URL_PREFIX}/shop/mypage#orders")


@app.route(f"{URL_PREFIX}/shop/api/my-orders")
@login_required
def shop_my_orders_api():
    """고객 본인 주문/문의 리스트"""
    _init_orders_db()
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        rows = conn.execute(
            "SELECT id, type, username, customer_name, brand, product_name, product_code, price, status, order_number, created_at FROM orders WHERE username = ? ORDER BY created_at DESC LIMIT 100",
            (username,)
        ).fetchall()
        orders = [{c: r[c] for c in r.keys()} for r in rows]
        # 상품 이미지 추가
        try:
            from product_db import _conn as prod_conn
            pconn = prod_conn()
            for o in orders:
                code = o.get("product_code", "")
                if code and not code.startswith("ORD"):
                    pr = pconn.execute("SELECT img_url FROM products WHERE internal_code=? OR product_code=? LIMIT 1", (code, code)).fetchone()
                    o["product_img"] = pr["img_url"] if pr else ""
                else:
                    o["product_img"] = ""
            pconn.close()
        except Exception:
            pass
        return jsonify({"ok": True, "orders": orders})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/order-detail/<order_number>")
@login_required
def order_detail_api(order_number):
    """주문번호로 주문 상세 조회"""
    _init_orders_db()
    username = session.get("username", "")
    role = session.get("role", "customer")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        if role == "admin":
            row = conn.execute("SELECT * FROM orders WHERE order_number=?", (order_number,)).fetchone()
        else:
            row = conn.execute("SELECT * FROM orders WHERE order_number=? AND username=?", (order_number, username)).fetchone()
        if not row:
            return jsonify({"ok": False, "message": "주문을 찾을 수 없습니다"})
        order = {c: row[c] for c in row.keys()}
        # 상품 상세 정보 조회
        product = None
        code = order.get("product_code", "")
        if code:
            try:
                from product_db import _conn as prod_conn
                pconn = prod_conn()
                pr = pconn.execute("SELECT * FROM products WHERE internal_code=? AND source_type='vintage' LIMIT 1", (code,)).fetchone()
                if pr:
                    import json as _json
                    product = {c: pr[c] for c in pr.keys()}
                    product["detail_images"] = _json.loads(product.get("detail_images") or "[]") if isinstance(product.get("detail_images"), str) else product.get("detail_images", [])
                    product["price_krw"] = _calc_vintage_price(product.get("price_jpy", 0))
                pconn.close()
            except Exception:
                pass
        return jsonify({"ok": True, "order": order, "product": product})
    finally:
        conn.close()


# ── 배송지 관리 API ──────────────────────────
def _init_address_db():
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS addresses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            label TEXT DEFAULT '',
            name TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            postal_code TEXT DEFAULT '',
            address TEXT DEFAULT '',
            address_detail TEXT DEFAULT '',
            customs_id TEXT DEFAULT '',
            business_number TEXT DEFAULT '',
            is_default INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )""")
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/addresses")
@login_required
def get_addresses():
    """저장된 배송지 목록 (최대 2개)"""
    _init_address_db()
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        rows = conn.execute("SELECT * FROM addresses WHERE username=? ORDER BY is_default DESC, id DESC LIMIT 2", (username,)).fetchall()
        return jsonify({"ok": True, "addresses": [{c: r[c] for c in r.keys()} for r in rows]})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/addresses", methods=["POST"])
@login_required
def save_address():
    """배송지 저장 (최대 2개)"""
    _init_address_db()
    data = request.json or {}
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        count = conn.execute("SELECT count(*) FROM addresses WHERE username=?", (username,)).fetchone()[0]
        addr_id = data.get("id")
        if addr_id:
            # 기존 배송지 수정
            conn.execute("""UPDATE addresses SET label=?, name=?, phone=?, postal_code=?, address=?,
                            address_detail=?, customs_id=?, business_number=?, is_default=? WHERE id=? AND username=?""",
                         (data.get("label",""), data.get("name",""), data.get("phone",""),
                          data.get("postal_code",""), data.get("address",""), data.get("address_detail",""),
                          data.get("customs_id",""), data.get("business_number",""),
                          1 if data.get("is_default") else 0, addr_id, username))
        elif count >= 2:
            return jsonify({"ok": False, "message": "배송지는 최대 2개까지 저장 가능합니다"})
        else:
            is_default = 1 if (count == 0 or data.get("is_default")) else 0
            conn.execute("""INSERT INTO addresses (username, label, name, phone, postal_code, address,
                            address_detail, customs_id, business_number, is_default)
                            VALUES (?,?,?,?,?,?,?,?,?,?)""",
                         (username, data.get("label",""), data.get("name",""), data.get("phone",""),
                          data.get("postal_code",""), data.get("address",""), data.get("address_detail",""),
                          data.get("customs_id",""), data.get("business_number",""), is_default))
        # is_default 설정 시 나머지 해제
        if data.get("is_default"):
            new_id = addr_id or conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.execute("UPDATE addresses SET is_default=0 WHERE username=? AND id!=?", (username, new_id))
        conn.commit()
        return jsonify({"ok": True, "message": "배송지 저장 완료"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/addresses/<int:addr_id>", methods=["DELETE"])
@login_required
def delete_address(addr_id):
    """배송지 삭제"""
    _init_address_db()
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("DELETE FROM addresses WHERE id=? AND username=?", (addr_id, username))
        conn.commit()
        return jsonify({"ok": True, "message": "삭제 완료"})
    finally:
        conn.close()


# ── 장바구니 API ──────────────────────────
def _init_cart_db():
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS cart (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            product_code TEXT NOT NULL,
            brand TEXT DEFAULT '',
            product_name TEXT DEFAULT '',
            price TEXT DEFAULT '',
            price_jpy INTEGER DEFAULT 0,
            img_url TEXT DEFAULT '',
            is_sold_out INTEGER DEFAULT 0,
            checked_at TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(username, product_code)
        )""")
        conn.commit()
        # 기존 테이블에 컬럼 추가
        for col, default in [("is_sold_out", "0"), ("checked_at", "''"), ("exchange_rate", "''")]:
            try:
                conn.execute(f"ALTER TABLE cart ADD COLUMN {col} DEFAULT {default}")
                conn.commit()
            except Exception:
                pass
    except Exception:
        pass
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/cart", methods=["GET"])
@login_required
def get_cart():
    _init_cart_db()
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        rows = conn.execute("SELECT * FROM cart WHERE username=? ORDER BY created_at DESC", (username,)).fetchall()
        items = [{c: r[c] for c in r.keys()} for r in rows]
        # cart 테이블의 is_sold_out 값 사용
        for item in items:
            is_sold = item.get("is_sold_out", 0)
            checked = item.get("checked_at", "")
            item["sold_out"] = bool(is_sold)
            item["order_status"] = "품절" if is_sold else ("주문가능" if checked else "확인중")
        return jsonify({"ok": True, "items": items, "count": len(items)})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/cart", methods=["POST"])
@login_required
def add_to_cart():
    _init_cart_db()
    data = request.json or {}
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        product_code = data.get("code", "")
        # 중복 체크
        existing = conn.execute("SELECT id FROM cart WHERE username=? AND product_code=?", (username, product_code)).fetchone()
        if existing:
            count = conn.execute("SELECT count(*) FROM cart WHERE username=?", (username,)).fetchone()[0]
            return jsonify({"ok": True, "count": count, "duplicate": True, "message": "이미 장바구니에 담긴 상품입니다."})

        # 가격을 고객 레벨에 맞게 재계산
        cart_price = data.get("price", "")
        cart_jpy = data.get("price_jpy", 0)
        if cart_jpy and cart_jpy > 0:
            try:
                _ur = conn.execute("SELECT level FROM users WHERE username=?", (username,)).fetchone()
                _lvl = _ur["level"] if _ur else "b2c"
                _rp = _calc_vintage_price(cart_jpy, _lvl)
                if _rp > 0:
                    cart_price = f"{_rp:,}원"
            except Exception:
                pass

        cart_rate = str(get_cached_rate() or 9.23)
        conn.execute("""INSERT INTO cart (username, product_code, brand, product_name, price, price_jpy, img_url, exchange_rate)
                        VALUES (?,?,?,?,?,?,?,?)""",
                     (username, product_code, data.get("brand",""), data.get("name",""),
                      cart_price, cart_jpy, data.get("img_url",""), cart_rate))
        conn.commit()
        count = conn.execute("SELECT count(*) FROM cart WHERE username=?", (username,)).fetchone()[0]

        # 백그라운드 품절 체크 → cart 테이블에 저장
        t = threading.Thread(target=_bg_check_single_cart_item, args=(product_code,), daemon=True)
        t.start()

        return jsonify({"ok": True, "count": count})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/cart/<int:item_id>", methods=["DELETE"])
@login_required
def remove_from_cart(item_id):
    _init_cart_db()
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("DELETE FROM cart WHERE id=? AND username=?", (item_id, username))
        conn.commit()
        count = conn.execute("SELECT count(*) FROM cart WHERE username=?", (username,)).fetchone()[0]
        return jsonify({"ok": True, "count": count})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/myinfo", methods=["GET"])
@login_required
def get_myinfo():
    """본인 회원정보 조회"""
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        row = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        if not row:
            return jsonify({"ok": False})
        user = {c: (row[c] or "") for c in row.keys() if c != "password_hash"}
        user["phone"] = _format_phone(user.get("phone", ""))
        return jsonify({"ok": True, "user": user})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/myinfo", methods=["POST"])
@login_required
def update_myinfo():
    """본인 회원정보 수정"""
    data = request.json or {}
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        fields = {k: data[k].strip() for k in
                  ["name","email","phone","postal_code","address","address_detail","customs_id","business_number"]
                  if k in data and data[k] is not None}
        if not fields:
            return jsonify({"ok": False, "message": "변경할 정보 없음"})
        # 전화번호 정규화
        if "phone" in fields:
            fields["phone"] = _normalize_phone(fields["phone"])
        sets = ", ".join(f"{k}=?" for k in fields)
        conn.execute(f"UPDATE users SET {sets} WHERE username=?", list(fields.values()) + [username])
        conn.commit()
        return jsonify({"ok": True, "message": "저장 완료"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/change-password", methods=["POST"])
@login_required
def change_password():
    """비밀번호 변경"""
    data = request.json or {}
    current = data.get("current", "")
    new_pw = data.get("new_password", "")
    if not current or not new_pw:
        return jsonify({"ok": False, "message": "현재 비밀번호와 새 비밀번호를 입력하세요"})
    if len(new_pw) < 4:
        return jsonify({"ok": False, "message": "새 비밀번호는 4자 이상이어야 합니다"})
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        row = conn.execute("SELECT password_hash FROM users WHERE username=?", (username,)).fetchone()
        if not row or not check_password_hash(row["password_hash"], current):
            return jsonify({"ok": False, "message": "현재 비밀번호가 틀립니다"})
        conn.execute("UPDATE users SET password_hash=? WHERE username=?",
                     (generate_password_hash(new_pw), username))
        conn.commit()
        return jsonify({"ok": True, "message": "비밀번호가 변경되었습니다"})
    finally:
        conn.close()


# ── 고객 요청/문의 게시판 ──────────────────────────
def _init_board_db():
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS board (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            name TEXT DEFAULT '',
            category TEXT DEFAULT 'general',
            title TEXT NOT NULL,
            content TEXT DEFAULT '',
            status TEXT DEFAULT 'open',
            admin_reply TEXT DEFAULT '',
            replied_at TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )""")
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/board", methods=["GET"])
@login_required
def get_board():
    _init_board_db()
    username = session.get("username", "")
    role = session.get("role", "customer")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        if role == "admin":
            rows = conn.execute("SELECT * FROM board ORDER BY created_at DESC LIMIT 200").fetchall()
        else:
            rows = conn.execute("SELECT * FROM board WHERE username=? ORDER BY created_at DESC LIMIT 100", (username,)).fetchall()
        posts = [{c: r[c] for c in r.keys()} for r in rows]
        return jsonify({"ok": True, "posts": posts})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/board", methods=["POST"])
@login_required
def create_post():
    _init_board_db()
    data = request.json or {}
    username = session.get("username", "")
    name = session.get("name", "")
    if not name:
        try:
            user_row = get_customer(username)
            if user_row and "name" in user_row.keys():
                name = user_row["name"] or ""
        except Exception:
            pass
    title = (data.get("title") or "").strip()
    content = (data.get("content") or "").strip()
    category = data.get("category", "general")
    if not title:
        return jsonify({"ok": False, "message": "제목을 입력해주세요"})
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("INSERT INTO board (username, name, category, title, content) VALUES (?,?,?,?,?)",
                     (username, name, category, title, content))
        conn.commit()
        # 텔레그램 알림 (관리자 글은 제외)
        if session.get("role") != "admin":
            cat_labels = {"brand":"브랜드 추가","feature":"기능 요청","inquiry":"문의","general":"기타"}
            try:
                from notifier import send_telegram
                send_telegram(
                    f"📋 <b>고객 요청 게시판</b>\n"
                    f"👤 {username}" + (f" ({name})" if name else "") + f"\n"
                    f"📂 {cat_labels.get(category, category)}\n"
                    f"📝 {title}\n"
                    f"💬 {content[:100]}"
                )
            except Exception:
                pass
        return jsonify({"ok": True, "message": "등록 완료"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/board/<int:post_id>/reply", methods=["POST"])
@admin_required
def reply_post(post_id):
    _init_board_db()
    data = request.json or {}
    reply = (data.get("reply") or "").strip()
    status = data.get("status", "answered")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("UPDATE board SET admin_reply=?, status=?, replied_at=datetime('now','localtime') WHERE id=?",
                     (reply, status, post_id))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/board/<int:post_id>", methods=["DELETE"])
@admin_required
def delete_post(post_id):
    _init_board_db()
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("DELETE FROM board WHERE id=?", (post_id,))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


# ── 공지사항 / 후기 게시판 ──────────────────────────
def _init_community_db():
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS notices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT DEFAULT '',
            pinned INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS reviews (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            product_code TEXT DEFAULT '',
            product_name TEXT DEFAULT '',
            brand TEXT DEFAULT '',
            rating INTEGER DEFAULT 5,
            title TEXT NOT NULL,
            content TEXT DEFAULT '',
            img_url TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )""")
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/notices")
def get_notices():
    """공지사항 목록 (비로그인도 가능)"""
    _init_community_db()
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        rows = conn.execute("SELECT * FROM notices ORDER BY pinned DESC, created_at DESC LIMIT 50").fetchall()
        return jsonify({"ok": True, "notices": [{c: r[c] for c in r.keys()} for r in rows]})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/notices", methods=["POST"])
@admin_required
def create_notice():
    """공지 작성 (관리자)"""
    _init_community_db()
    data = request.json or {}
    title = (data.get("title") or "").strip()
    content = (data.get("content") or "").strip()
    pinned = 1 if data.get("pinned") else 0
    if not title:
        return jsonify({"ok": False, "message": "제목을 입력해주세요"})
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("INSERT INTO notices (title, content, pinned) VALUES (?,?,?)", (title, content, pinned))
        conn.commit()
        return jsonify({"ok": True, "message": "공지 등록 완료"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/notices/<int:nid>", methods=["PUT"])
@admin_required
def update_notice(nid):
    """공지 수정 (관리자)"""
    _init_community_db()
    data = request.json or {}
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("UPDATE notices SET title=?, content=?, pinned=?, updated_at=datetime('now','localtime') WHERE id=?",
                     (data.get("title",""), data.get("content",""), 1 if data.get("pinned") else 0, nid))
        conn.commit()
        return jsonify({"ok": True, "message": "수정 완료"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/notices/<int:nid>", methods=["DELETE"])
@admin_required
def delete_notice(nid):
    """공지 삭제 (관리자)"""
    _init_community_db()
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("DELETE FROM notices WHERE id=?", (nid,))
        conn.commit()
        return jsonify({"ok": True, "message": "삭제 완료"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/reviews")
def get_reviews():
    """후기 목록 (비로그인도 가능)"""
    _init_community_db()
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        rows = conn.execute("SELECT * FROM reviews ORDER BY created_at DESC LIMIT 100").fetchall()
        return jsonify({"ok": True, "reviews": [{c: r[c] for c in r.keys()} for r in rows]})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/review-upload", methods=["POST"])
@login_required
def upload_review_image():
    """후기 이미지 업로드 (최대 3장)"""
    from data_manager import get_path
    upload_dir = os.path.join(get_path("db"), "review_images")
    os.makedirs(upload_dir, exist_ok=True)
    files = request.files.getlist("images")
    if not files:
        return jsonify({"ok": False, "message": "이미지를 선택해주세요"})
    urls = []
    for f in files[:3]:
        if f and f.filename:
            import uuid
            ext = os.path.splitext(f.filename)[1].lower() or ".jpg"
            if ext not in (".jpg", ".jpeg", ".png", ".webp"):
                continue
            fname = f"{uuid.uuid4().hex[:12]}{ext}"
            f.save(os.path.join(upload_dir, fname))
            urls.append(f"{URL_PREFIX}/shop/review-img/{fname}")
    return jsonify({"ok": True, "urls": urls})


@app.route(f"{URL_PREFIX}/shop/review-img/<filename>")
def serve_review_image(filename):
    """후기 이미지 서빙"""
    from data_manager import get_path
    upload_dir = os.path.join(get_path("db"), "review_images")
    return send_from_directory(upload_dir, filename)


@app.route(f"{URL_PREFIX}/shop/api/reviews", methods=["POST"])
@login_required
def create_review():
    """후기 작성 (로그인 필요)"""
    _init_community_db()
    data = request.json or {}
    username = session.get("username", "")
    title = (data.get("title") or "").strip()
    content = (data.get("content") or "").strip()
    rating = int(data.get("rating") or 5)
    img_url = data.get("img_url", "")  # 쉼표 구분 다중 이미지
    if not title:
        return jsonify({"ok": False, "message": "제목을 입력해주세요"})
    if rating < 1 or rating > 5:
        rating = 5
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("""INSERT INTO reviews (username, product_code, product_name, brand, rating, title, content, img_url)
                        VALUES (?,?,?,?,?,?,?,?)""",
                     (username, data.get("product_code",""), data.get("product_name",""),
                      data.get("brand",""), rating, title, content, img_url))
        conn.commit()
        return jsonify({"ok": True, "message": "후기 등록 완료"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/reviews/<int:rid>", methods=["DELETE"])
@login_required
def delete_review(rid):
    """후기 삭제 (본인 또는 관리자)"""
    _init_community_db()
    username = session.get("username", "")
    role = session.get("role", "customer")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        if role == "admin":
            conn.execute("DELETE FROM reviews WHERE id=?", (rid,))
        else:
            conn.execute("DELETE FROM reviews WHERE id=? AND username=?", (rid, username))
        conn.commit()
        return jsonify({"ok": True, "message": "삭제 완료"})
    finally:
        conn.close()


def _calc_vintage_cost(jpy: int) -> int:
    """원가 계산 (일본 상품가 + 수수료 + 택배비) × 환율"""
    if not jpy or jpy <= 0:
        return 0
    cfg = _vintage_price_config
    fee = cfg["jp_fee_pct"] / 100
    jp_ship = cfg.get("jp_domestic_shipping", 800)
    rate = get_cached_rate() or 9.23
    cost = (jpy + jp_ship) * (1 + fee) * rate
    return int(math.ceil(cost / 100) * 100)


def _calc_vintage_price(jpy: int, margin_type="b2c") -> int:
    """빈티지 상품 한국 판매가 계산
    B2C: 일본가 기반 정상 계산
    B2B: B2C 가격에서 5% 할인
    """
    if not jpy or jpy <= 0:
        return 0
    cfg = _vintage_price_config
    fee = cfg["jp_fee_pct"] / 100
    markup = cfg["buy_markup_pct"] / 100
    margin = cfg.get("margin_b2c_pct", 15.0) / 100  # 항상 B2C 마진 기준
    jp_ship = cfg.get("jp_domestic_shipping", 800)
    intl_ship = cfg["intl_shipping_krw"]
    rate = get_cached_rate() or 9.23
    jpy_total = (jpy + jp_ship) * (1 + fee)
    raw = jpy_total * rate * (1 + markup) * (1 + margin) + intl_ship
    b2c_price = int(math.ceil(raw / 100) * 100)
    if margin_type == "b2b":
        return int(math.ceil(b2c_price * 0.95 / 100) * 100)  # B2C의 5% 할인
    return b2c_price


# ── 토스페이먼츠 결제 ──────────────────────────
TOSS_CLIENT_KEY = "test_ck_Poxy1XQL8R96wqlAw7GNr7nO5Wml"
TOSS_SECRET_KEY = "test_sk_DpexMgkW36wOYALjW94JVGbR5ozO"


@app.route(f"{URL_PREFIX}/shop/payment")
@login_required
def payment_page():
    """결제 페이지"""
    # 세션 정리 (쿠키 크기 초과 방지)
    session.pop("_pay_order_ids", None)
    return render_template("payment.html",
                           url_prefix=URL_PREFIX,
                           toss_client_key=TOSS_CLIENT_KEY,
                           username=session.get("username", ""))


@app.route(f"{URL_PREFIX}/shop/payment/success")
@login_required
def payment_success():
    """결제 성공 콜백 → 결제 확인 + 주문 저장"""
    payment_key = request.args.get("paymentKey", "")
    order_id = request.args.get("orderId", "")
    amount = request.args.get("amount", 0, type=int)

    if not payment_key or not order_id:
        return redirect(f"{URL_PREFIX}/shop")

    # 토스 결제 확인 API
    import base64
    auth = base64.b64encode(f"{TOSS_SECRET_KEY}:".encode()).decode()
    try:
        resp = requests.post(
            "https://api.tosspayments.com/v1/payments/confirm",
            json={"paymentKey": payment_key, "orderId": order_id, "amount": amount},
            headers={"Authorization": f"Basic {auth}", "Content-Type": "application/json"},
            timeout=10
        )
        result = resp.json()

        if resp.status_code == 200 and result.get("status") == "DONE":
            # 결제 성공 → 주문 저장
            _init_orders_db()
            username = session.get("username", "")
            customer_name = session.get("name", "")
            if not customer_name:
                try:
                    user_row = get_customer(username)
                    if user_row and "name" in user_row.keys():
                        customer_name = user_row["name"] or ""
                except Exception:
                    pass

            meta = result.get("orderName", "")
            payment_memo = f"토스결제 {result.get('method','')} {payment_key[:20]}"

            # URL 파라미터에서 원본 주문 ID들 추출
            pay_order_ids = request.args.get("order_ids", "") or ""

            from user_db import _conn as user_conn
            conn = user_conn()
            try:
                if pay_order_ids:
                    # 개별 주문 ID들의 상태를 confirmed로 업데이트 + 결제 메모 추가
                    ids = [i.strip() for i in pay_order_ids.split(",") if i.strip()]
                    updated_orders = []
                    for oid in ids:
                        try:
                            row = conn.execute("SELECT * FROM orders WHERE id=? AND username=?", (int(oid), username)).fetchone()
                            if row:
                                conn.execute("UPDATE orders SET status='confirmed', memo=? WHERE id=?",
                                             (payment_memo, int(oid)))
                                updated_orders.append(f"{row['brand']} {row['product_name'] or ''}")
                        except Exception:
                            pass
                    conn.commit()
                    detail_text = "\n".join(f"  · {n}" for n in updated_orders) if updated_orders else meta
                else:
                    # 개별 ID 없는 경우 (단건 결제 등) 기존 방식
                    order_number = _generate_order_number(conn)
                    conn.execute("""INSERT INTO orders (type, username, customer_name, brand, product_name, product_code, price, price_jpy, status, memo, order_number)
                                    VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                                 ("order", username, customer_name, "", meta, "", f"{amount:,}원", 0, "confirmed",
                                  payment_memo, order_number))
                    conn.commit()
                    detail_text = meta
            finally:
                conn.close()

            # 텔레그램 알림
            try:
                from notifier import send_telegram
                send_telegram(
                    f"💳 <b>결제 완료!</b>\n"
                    f"👤 {username}" + (f" ({customer_name})" if customer_name else "") + f"\n"
                    f"📦 {detail_text}\n"
                    f"💰 {amount:,}원 ({result.get('method','')})\n"
                    f"🔖 {order_id}"
                )
            except Exception:
                pass

            return redirect(f"{URL_PREFIX}/shop/mypage#orders")
        else:
            error_msg = result.get("message", "결제 확인 실패")
            logger.warning(f"결제 확인 실패: {error_msg}")
            return redirect(f"{URL_PREFIX}/shop/payment/fail?message={error_msg}")
    except Exception as e:
        logger.error(f"결제 확인 오류: {e}")
        return redirect(f"{URL_PREFIX}/shop/payment/fail?message={str(e)[:100]}")


@app.route(f"{URL_PREFIX}/shop/payment/fail")
def payment_fail():
    """결제 실패"""
    message = request.args.get("message", "결제가 취소되었습니다")
    return render_template("payment_fail.html",
                           url_prefix=URL_PREFIX, message=message)


@app.route(f"{URL_PREFIX}/shop/api/update-order-status", methods=["POST"])
@login_required
def update_my_order_status():
    """고객 주문 상태/메모 업데이트 (계좌이체 등)"""
    data = request.json or {}
    order_id = data.get("id")
    status = data.get("status", "")
    memo = data.get("memo", "")
    username = session.get("username", "")
    if not order_id:
        return jsonify({"ok": False})
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        if status:
            conn.execute("UPDATE orders SET status=? WHERE id=? AND username=?", (status, order_id, username))
        if memo:
            conn.execute("UPDATE orders SET memo=? WHERE id=? AND username=?", (memo, order_id, username))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/cancel-order/<int:order_id>", methods=["POST"])
@login_required
def cancel_my_order(order_id):
    """고객이 직접 주문 취소
    - new(신규): 즉시 취소
    - confirmed(주문확인) 이후: 취소요청 → 관리자 승인 필요
    """
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        row = conn.execute("SELECT * FROM orders WHERE id=? AND username=?", (order_id, username)).fetchone()
        if not row:
            return jsonify({"ok": False, "message": "주문을 찾을 수 없습니다"})

        st = row["status"]
        if st in ("cancelled", "cancel_request"):
            return jsonify({"ok": False, "message": "이미 취소되었거나 취소 요청 중입니다"})
        if st not in ("new", "confirmed", "processing"):
            return jsonify({"ok": False, "message": f"'{st}' 상태에서는 취소할 수 없습니다"})

        if st == "new":
            # 신규 주문은 즉시 취소
            conn.execute("UPDATE orders SET status='cancelled' WHERE id=?", (order_id,))
            conn.commit()
            try:
                from notifier import send_telegram
                send_telegram(f"🚫 <b>주문 취소</b>\n👤 {username}\n📦 {row['product_name'] or ''}\n💰 {row['price'] or ''}")
            except Exception:
                pass
            return jsonify({"ok": True, "message": "주문이 취소되었습니다"})
        else:
            # 주문확인 이후는 취소요청 → 관리자 승인 필요
            conn.execute("UPDATE orders SET status='cancel_request' WHERE id=?", (order_id,))
            conn.commit()
            try:
                from notifier import send_telegram
                send_telegram(f"⚠️ <b>취소 요청</b>\n👤 {username}\n📦 {row['product_name'] or ''}\n💰 {row['price'] or ''}\n\n관리자 승인이 필요합니다.")
            except Exception:
                pass
            return jsonify({"ok": True, "message": "취소 요청이 접수되었습니다. 관리자 승인 후 취소됩니다."})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/orders/<int:order_id>/approve-cancel", methods=["POST"])
@admin_required
def approve_cancel(order_id):
    """관리자: 취소 요청 승인/거절"""
    data = request.get_json() or {}
    action = data.get("action", "approve")  # approve / reject
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        row = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not row:
            return jsonify({"ok": False, "message": "주문을 찾을 수 없습니다"})
        if row["status"] != "cancel_request":
            return jsonify({"ok": False, "message": "취소 요청 상태가 아닙니다"})

        if action == "approve":
            conn.execute("UPDATE orders SET status='cancelled' WHERE id=?", (order_id,))
            conn.commit()
            # 고객에게 문자 알림
            try:
                user = conn.execute("SELECT phone FROM users WHERE username=?", (row["username"],)).fetchone()
                if user and user["phone"]:
                    from aligo_sms import send_sms, load_config
                    load_config()
                    send_sms(user["phone"], f"[TheOne Vintage] 주문 취소가 승인되었습니다.\n주문번호: {row.get('order_number','')}\n상품: {row.get('brand','')} {(row.get('product_name',''))[:20]}")
            except Exception:
                pass
            try:
                from notifier import send_telegram
                send_telegram(f"✅ <b>취소 승인</b>\n주문번호: {row.get('order_number','')}\n👤 {row['username']}\n📦 {row['product_name'] or ''}")
            except Exception:
                pass
            return jsonify({"ok": True, "message": "취소가 승인되었습니다"})
        else:
            # 거절 → 이전 상태(confirmed)로 복원
            conn.execute("UPDATE orders SET status='confirmed' WHERE id=?", (order_id,))
            conn.commit()
            try:
                user = conn.execute("SELECT phone FROM users WHERE username=?", (row["username"],)).fetchone()
                if user and user["phone"]:
                    from aligo_sms import send_sms, load_config
                    load_config()
                    send_sms(user["phone"], f"[TheOne Vintage] 주문 취소 요청이 거절되었습니다.\n주문번호: {row.get('order_number','')}\n문의사항은 카카오톡으로 연락주세요.")
            except Exception:
                pass
            return jsonify({"ok": True, "message": "취소 요청이 거절되었습니다"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/coupon-check", methods=["POST"])
@login_required
def coupon_check():
    """쿠폰 코드 확인 (향후 DB 연동 가능)"""
    data = request.get_json() or {}
    code = (data.get("code") or "").strip().upper()
    subtotal = int(data.get("subtotal") or 0)
    if not code:
        return jsonify({"ok": False, "message": "쿠폰 코드를 입력해주세요"})

    # 쿠폰 정의 (향후 DB 테이블로 이동 가능)
    coupons = {
        "WELCOME10": {"type": "percent", "value": 10, "max_discount": 50000, "desc": "신규회원 10% 할인"},
        "THEONE5000": {"type": "fixed", "value": 5000, "desc": "5,000원 할인"},
    }

    coupon = coupons.get(code)
    if not coupon:
        return jsonify({"ok": False, "message": "유효하지 않은 쿠폰 코드입니다"})

    if coupon["type"] == "percent":
        discount = int(subtotal * coupon["value"] / 100)
        max_d = coupon.get("max_discount", 999999999)
        discount = min(discount, max_d)
    else:
        discount = coupon["value"]

    discount = min(discount, subtotal)  # 상품금액 초과 방지
    return jsonify({"ok": True, "discount": discount, "message": coupon["desc"]})


@app.route(f"{URL_PREFIX}/shop/api/notify", methods=["POST"])
@login_required
def shop_notify():
    """주문/문의 시 텔레그램 알림"""
    data = request.json or {}
    ntype = data.get("type", "inquiry")
    brand = data.get("brand", "")
    name = data.get("name", "")
    code = data.get("code", "")
    price = data.get("price", "")
    price_jpy = data.get("price_jpy", 0)
    username = session.get("username", "비회원")
    customer_name = session.get("name", "")
    # 세션에 이름 없으면 DB에서 조회
    if not customer_name and username != "비회원":
        try:
            user_row = get_customer(username)
            if user_row and "name" in user_row.keys():
                customer_name = user_row["name"] or ""
        except Exception:
            pass

    icon = "🛒" if ntype == "order" else "💬"
    label = "주문" if ntype == "order" else "문의"
    user_info = f"{username}" + (f" ({customer_name})" if customer_name else "")

    msg = (
        f"{icon} <b>고객 {label} 알림</b>\n"
        f"👤 {user_info}\n"
        f"🏷 {brand} {name}\n"
        f"🔖 {code}\n"
        f"💰 {price} (¥{price_jpy:,})"
    )
    try:
        from notifier import send_telegram
        send_telegram(msg)
    except Exception as e:
        logger.warning(f"주문 알림 전송 실패: {e}")

    # 주문 중복 체크 (같은 상품코드 + 미처리 상태)
    if ntype == "order" and code:
        from user_db import _conn as _uc2
        _c2 = _uc2()
        try:
            dup = _c2.execute(
                "SELECT id, order_number FROM orders WHERE username=? AND product_code=? AND status IN ('new','confirmed','processing')",
                (username, code)
            ).fetchone()
            if dup:
                return jsonify({"ok": True, "duplicate": True, "message": f"이미 주문된 상품입니다. (주문번호: {dup['order_number']})"})
        finally:
            _c2.close()

    # 주문 DB에 저장
    try:
        _save_order(ntype, username, customer_name, brand, name, code, price, price_jpy)
    except Exception as e:
        logger.warning(f"주문 저장 실패: {e}")
    return jsonify({"ok": True})


def _init_orders_db():
    """주문 테이블 초기화"""
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT DEFAULT 'inquiry',
            username TEXT DEFAULT '',
            customer_name TEXT DEFAULT '',
            brand TEXT DEFAULT '',
            product_name TEXT DEFAULT '',
            product_code TEXT DEFAULT '',
            price TEXT DEFAULT '',
            price_jpy INTEGER DEFAULT 0,
            status TEXT DEFAULT 'new',
            memo TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )""")
        conn.commit()
        # 마이그레이션: order_number 컬럼 추가
        try:
            conn.execute("ALTER TABLE orders ADD COLUMN order_number TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass
        # 마이그레이션: 택배사/송장번호/환율 컬럼 추가
        for col in ["courier", "tracking_no", "exchange_rate"]:
            try:
                conn.execute(f"ALTER TABLE orders ADD COLUMN {col} TEXT DEFAULT ''")
                conn.commit()
            except Exception:
                pass
        # 기존 주문에 order_number 부여
        rows = conn.execute("SELECT id, created_at FROM orders WHERE order_number='' OR order_number IS NULL ORDER BY id").fetchall()
        for r in rows:
            on = _generate_order_number(conn, r["created_at"])
            conn.execute("UPDATE orders SET order_number=? WHERE id=?", (on, r["id"]))
        if rows:
            conn.commit()
    except Exception:
        pass
    finally:
        conn.close()


_order_seq = 0
def _generate_order_number(conn=None, created_at=None):
    """주문번호 생성: ORD-YYMMDD-HHMMSS + 시퀀스 (중복 방지)"""
    global _order_seq
    _order_seq += 1
    dt = datetime.now()
    return f"ORD-{dt.strftime('%y%m%d')}-{dt.strftime('%H%M%S')}{_order_seq:02d}"


@app.route(f"{URL_PREFIX}/shop/api/bulk-order", methods=["POST"])
@login_required
def bulk_order():
    """장바구니 다중 주문 — 1건의 다중주문으로 저장"""
    data = request.get_json() or {}
    items = data.get("items", [])
    if not items:
        return jsonify({"ok": False, "message": "주문할 상품이 없습니다"})

    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        customer = conn.execute("SELECT name FROM users WHERE username=?", (username,)).fetchone()
        customer_name = customer["name"] if customer else username
    finally:
        conn.close()

    if len(items) == 1:
        # 1건이면 일반 주문
        it = items[0]
        order_number = _save_order("order", username, customer_name, it["brand"], it["name"], it["code"], it["price"], it.get("price_jpy", 0))
        # 장바구니에서 삭제
        conn2 = user_conn()
        conn2.execute("DELETE FROM cart WHERE id=?", (it.get("cart_id", 0),))
        conn2.commit()
        conn2.close()
        return jsonify({"ok": True, "order_number": order_number})
    else:
        # 다중 주문 — 개별 주문 + 묶음 정보 저장
        total_jpy = sum(it.get("price_jpy", 0) for it in items)
        brands = list(set(it["brand"] for it in items))
        brand_text = brands[0] if len(brands) == 1 else f"{brands[0]} 외 {len(brands)-1}"

        # 개별 주문 저장
        order_numbers = []
        for it in items:
            on = _save_order("order", username, customer_name, it["brand"], it["name"], it["code"], it["price"], it.get("price_jpy", 0))
            order_numbers.append(on)

        # 다중주문 묶음 저장
        _init_orders_db()
        conn3 = user_conn()
        try:
            batch_number = _generate_order_number(conn3)
            # 대표 상품명: "첫 상품명 외 N건"
            first_name = (items[0].get("name") or "")[:30]
            batch_name = f"{first_name} 외 {len(items)-1}건" if len(items) > 1 else first_name
            detail_json = json.dumps([{"order_number": on, "brand": it["brand"], "name": it["name"], "code": it["code"], "price": it["price"]} for on, it in zip(order_numbers, items)], ensure_ascii=False)
            conn3.execute("""INSERT INTO orders (type, username, customer_name, brand, product_name, product_code, price, price_jpy, order_number, memo)
                            VALUES (?,?,?,?,?,?,?,?,?,?)""",
                         ("order", username, customer_name, brand_text, batch_name, ",".join(order_numbers), f"총 {len(items)}건", total_jpy, batch_number, detail_json))
            conn3.commit()
        finally:
            conn3.close()

        # 장바구니에서 삭제
        conn4 = user_conn()
        for it in items:
            conn4.execute("DELETE FROM cart WHERE id=?", (it.get("cart_id", 0),))
        conn4.commit()
        conn4.close()

        return jsonify({"ok": True, "order_number": batch_number, "count": len(items)})


def _save_order(ntype, username, customer_name, brand, product_name, product_code, price, price_jpy):
    _init_orders_db()
    # 가격을 고객 레벨에 맞게 계산 (주문 시점 환율로 고정)
    current_rate = get_cached_rate() or 9.23
    if price_jpy and price_jpy > 0:
        try:
            from user_db import _conn as _ulc
            _uc = _ulc()
            _ur = _uc.execute("SELECT level FROM users WHERE username=?", (username,)).fetchone()
            _uc.close()
            user_lvl = _ur["level"] if _ur else "b2c"
            recalc = _calc_vintage_price(price_jpy, user_lvl)
            if recalc > 0:
                price = f"{recalc:,}원"
        except Exception:
            pass

    # product_code가 고유번호(No.S-)가 아니면 자동 변환
    if product_code and not product_code.startswith("No."):
        try:
            from product_db import _conn as p_conn, _generate_internal_code
            pc = p_conn()
            row = pc.execute("SELECT id, internal_code FROM products WHERE product_code=? LIMIT 1", (product_code,)).fetchone()
            if row:
                if row["internal_code"]:
                    product_code = row["internal_code"]
                else:
                    new_code = _generate_internal_code(pc, "2ndstreet")
                    pc.execute("UPDATE products SET internal_code=? WHERE id=?", (new_code, row["id"]))
                    pc.commit()
                    product_code = new_code
            pc.close()
        except Exception:
            pass

    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        order_number = _generate_order_number(conn)
        conn.execute("""INSERT INTO orders (type, username, customer_name, brand, product_name, product_code, price, price_jpy, order_number, exchange_rate)
                        VALUES (?,?,?,?,?,?,?,?,?,?)""",
                     (ntype, username, customer_name, brand, product_name, product_code, price, price_jpy, order_number, str(current_rate)))
        conn.commit()

        # 주문 접수 시 고객에게 자동 문자 발송
        try:
            user = conn.execute("SELECT phone FROM users WHERE username=?", (username,)).fetchone()
            if user and user["phone"]:
                from aligo_sms import send_sms, load_config
                load_config()
                msg = f"[TheOne Vintage] 주문이 접수되었습니다.\n주문번호: {order_number}\n상품: {brand} {product_name[:20]}\n확인 후 안내드리겠습니다."
                send_sms(user["phone"], msg, title="TheOne Vintage")
                logger.info(f"[SMS] 주문 접수 알림: {username} ({user['phone']})")
        except Exception as e:
            logger.warning(f"[SMS] 주문 접수 알림 실패: {e}")

        # 백그라운드 품절 체크 → 품절이면 자동 취소 + 문자
        order_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        t = threading.Thread(target=_bg_check_and_update_order, args=(order_id, product_code, username), daemon=True)
        t.start()

        return order_number
    finally:
        conn.close()


def _check_product_soldout(product_code):
    """상품 품절 여부 정밀 체크 (Playwright 브라우저) — Windows에서만 DB 업데이트"""
    from product_db import _conn
    conn = _conn()
    try:
        row = conn.execute(
            "SELECT id, link, product_status FROM products WHERE internal_code=? OR product_code=? LIMIT 1",
            (product_code, product_code)
        ).fetchone()
        if not row or not row["link"]:
            return None

        import platform
        is_windows = platform.system() == "Windows"

        # DB에 이미 품절 마킹된 경우
        if row["product_status"] == "sold_out":
            return True

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            import asyncio
            result = asyncio.run(_check_soldout_playwright(row["link"]))
            # Windows에서만 DB 업데이트
            if is_windows:
                if result is True:
                    conn.execute("UPDATE products SET product_status='sold_out', checked_at=? WHERE id=?", (now, row["id"]))
                    conn.commit()
                    logger.info(f"[품절체크] 품절 확인: {product_code}")
                elif result is False:
                    conn.execute("UPDATE products SET product_status='available', checked_at=? WHERE id=?", (now, row["id"]))
                    conn.commit()
            return result
        except Exception as e:
            logger.warning(f"[품절체크] Playwright 오류: {e}")
            return None
    finally:
        conn.close()


def _bg_check_single_cart_item(product_code):
    """백그라운드: 장바구니 단건 품절 체크 → cart 테이블에 저장 (상품DB 미수정)

    Playwright 대신 requests + 정규식으로 빠르게 체크 (0.5~1초)
    403 차단 시 → 확인불가로 처리 (품절 아님)
    """
    from product_db import _conn as p_conn
    from user_db import _conn as u_conn

    pc = p_conn()
    r = pc.execute("SELECT link FROM products WHERE internal_code=? OR product_code=? LIMIT 1", (product_code, product_code)).fetchone()
    pc.close()
    if not r or not r["link"]:
        return

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    is_sold = -1  # -1: 확인불가

    try:
        import requests as _req
        resp = _req.get(r["link"], timeout=10, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            "Accept-Language": "ja,ja-JP;q=0.9",
            "Accept": "text/html,application/xhtml+xml",
        })
        if resp.status_code == 404:
            is_sold = 1
        elif resp.status_code == 200:
            body = resp.text
            if "SOLD OUT" in body or "この商品は売切れ" in body or "この商品は売り切れ" in body:
                is_sold = 1
            elif "Access Denied" in body:
                is_sold = -1  # 차단 → Playwright 재시도
            else:
                is_sold = 0
        else:
            is_sold = -1  # 403 등 → Playwright 재시도
    except Exception:
        is_sold = -1

    # HTTP 차단 시 Playwright로 정확한 체크
    if is_sold == -1:
        try:
            import asyncio
            result = asyncio.run(_check_soldout_playwright(r["link"]))
            if result is True:
                is_sold = 1
            elif result is False:
                is_sold = 0
            # None이면 여전히 확인불가
        except Exception as e:
            logger.warning(f"[장바구니 단건체크] Playwright 폴백 오류: {e}")

    if is_sold >= 0:
        try:
            uc = u_conn()
            uc.execute("UPDATE cart SET is_sold_out=?, checked_at=? WHERE product_code=?", (is_sold, now, product_code))
            uc.commit()
            uc.close()
            logger.info(f"[장바구니 단건체크] {product_code} → {'품절' if is_sold else '주문가능'} (HTTP)")
        except Exception:
            pass


_cart_check_running = False

def _bg_check_cart_soldout_all():
    """백그라운드: 장바구니 전체 품절 체크 → cart 테이블에 저장 (상품DB 미수정)"""
    global _cart_check_running
    if _cart_check_running:
        return
    _cart_check_running = True

    import asyncio
    from user_db import _conn as u_conn
    from product_db import _conn as p_conn

    async def _run():
        from playwright.async_api import async_playwright
        uc = u_conn()
        pc = p_conn()
        rows = uc.execute("SELECT DISTINCT product_code FROM cart WHERE product_code IS NOT NULL AND product_code != ''").fetchall()
        codes = [r[0] for r in rows]

        if not codes:
            uc.close(); pc.close()
            return

        logger.info(f"[장바구니 체크] {len(codes)}개 상품 품절 체크 시작")

        pw = await async_playwright().start()
        browser = await pw.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled", "--lang=ja", "--disable-translate"],
        )
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            locale="ja-JP", timezone_id="Asia/Tokyo",
            viewport={"width": 1280, "height": 900},
        )
        await context.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>false});")
        page = await context.new_page()

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sold = 0; avail = 0

        for i, code in enumerate(codes):
            r = pc.execute("SELECT link FROM products WHERE internal_code=? OR product_code=? LIMIT 1", (code, code)).fetchone()
            if not r or not r["link"]:
                continue
            try:
                resp = await page.goto(r["link"], wait_until="domcontentloaded", timeout=20000)
                await asyncio.sleep(2)
                is_sold = 0
                if resp and resp.status == 404:
                    is_sold = 1
                elif resp and resp.status in (403, 429, 503):
                    continue  # 차단 시 스킵
                else:
                    result = await page.evaluate("""() => {
                        const body = document.body.innerText || '';
                        // 확실한 품절 문구
                        if (body.includes('※申し訳ございません。この商品は売切れ') ||
                            body.includes('※申し訳ございません。この商品は売り切れ') ||
                            body.includes('この商品は現在販売しておりません')) return 'sold';
                        // 가격 요소 주변의 SOLD OUT 체크 (추천상품 영역 오탐 방지)
                        const price = document.querySelector('[itemprop="price"], .priceMain, .priceNum');
                        if (!price) return 'unknown';
                        let parent = price.parentElement;
                        for (let i = 0; i < 4 && parent; i++) {
                            if ((parent.innerText || '').includes('SOLD OUT')) return 'sold';
                            parent = parent.parentElement;
                        }
                        return 'ok';
                    }""")
                    is_sold = 1 if result == "sold" else 0

                uc.execute("UPDATE cart SET is_sold_out=?, checked_at=? WHERE product_code=?", (is_sold, now, code))
                uc.commit()
                if is_sold:
                    sold += 1
                else:
                    avail += 1
            except Exception:
                pass
            await asyncio.sleep(1)

        await browser.close()
        await pw.stop()
        pc.close()
        uc.close()
        logger.info(f"[장바구니 체크] 완료: 주문가능 {avail}개 / 품절 {sold}개")

    try:
        asyncio.run(_run())
    except Exception as e:
        logger.warning(f"[장바구니 체크] 오류: {e}")
    finally:
        _cart_check_running = False


async def _check_soldout_playwright(url):
    """Playwright로 단일 상품 품절 체크"""
    from playwright.async_api import async_playwright
    playwright = None
    browser = None
    try:
        playwright = await async_playwright().start()
        browser = await playwright.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled", "--lang=ja", "--disable-translate"],
        )
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            locale="ja-JP",
            timezone_id="Asia/Tokyo",
            viewport={"width": 1280, "height": 900},
            extra_http_headers={"Accept-Language": "ja,ja-JP;q=0.9"},
        )
        await context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => false});
            Object.defineProperty(navigator, 'language', {get: () => 'ja'});
            Object.defineProperty(navigator, 'languages', {get: () => ['ja', 'ja-JP']});
        """)
        page = await context.new_page()
        resp = await page.goto(url, wait_until="domcontentloaded", timeout=15000)

        # 403/503 등 차단 응답 → 확인불가 (품절 아님)
        if resp is None:
            return None
        if resp.status in (403, 429, 503, 520, 521, 522, 523, 524):
            logger.warning(f"[품절체크] 차단됨 {resp.status}: {url}")
            return None
        if resp.status == 404:
            return True

        import asyncio as _aio
        await _aio.sleep(2)

        # 페이지 내 품절 문구 체크
        result = await page.evaluate("""() => {
            const body = document.body.innerText || '';
            // 차단된 경우 → 확인불가
            if (body.includes('Access Denied') || body.includes("don't have permission")) return 'blocked';
            // 확실한 품절 문구
            if (body.includes('※申し訳ございません。この商品は売切れ') ||
                body.includes('※申し訳ございません。この商品は売り切れ') ||
                body.includes('この商品は現在販売しておりません')) return 'sold_out';
            // 가격 요소 주변의 SOLD OUT 체크 (추천상품 영역 오탐 방지)
            const price = document.querySelector('[itemprop="price"], .priceMain, .priceNum');
            if (!price) return 'no_price';
            let parent = price.parentElement;
            for (let i = 0; i < 4 && parent; i++) {
                if ((parent.innerText || '').includes('SOLD OUT')) return 'sold_out';
                parent = parent.parentElement;
            }
            return 'available';
        }""")

        if result == "blocked":
            logger.warning(f"[품절체크] Access Denied: {url}")
            return None  # 차단 = 확인불가 (품절 아님)
        if result == "sold_out":
            return True
        if result == "no_price":
            return None  # 가격 못 찾음 = 확인불가
        return False  # available
    except Exception as e:
        logger.warning(f"[품절체크] Playwright 예외: {e}")
        return None
    finally:
        if browser:
            await browser.close()
        if playwright:
            await playwright.stop()


def _bg_check_and_update_order(order_id, product_code, username):
    """백그라운드: 주문 상품 품절 체크 → 품절이면 자동 취소 + 문자"""
    try:
        is_sold = _check_product_soldout(product_code)
        if is_sold:
            from user_db import _conn as user_conn
            conn = user_conn()
            try:
                order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
                if order and order["status"] == "new":
                    conn.execute("UPDATE orders SET status='sold_out' WHERE id=?", (order_id,))
                    conn.commit()
                    logger.info(f"[품절체크] 주문 #{order_id} 자동 품절취소 ({product_code})")
                    # SMS 발송
                    user = conn.execute("SELECT phone FROM users WHERE username=?", (username,)).fetchone()
                    if user and user["phone"]:
                        from aligo_sms import send_order_notification, load_config
                        load_config()
                        product_name = (order["product_name"] or "") if "product_name" in order.keys() else ""
                        order_number = (order["order_number"] or "") if "order_number" in order.keys() else ""
                        send_order_notification(user["phone"], order_number, "sold_out", product_name)
                        logger.info(f"[SMS] 품절 자동 알림: {username} ({user['phone']})")
            finally:
                conn.close()
    except Exception as e:
        logger.warning(f"[품절체크] 오류: {e}")


@app.route(f"{URL_PREFIX}/orders")
@admin_required
def get_orders():
    """주문/문의 리스트 조회"""
    _init_orders_db()
    type_filter = request.args.get("type", "")
    status_filter = request.args.get("status", "")
    search_query = request.args.get("q", "").strip()
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        sql = "SELECT * FROM orders WHERE 1=1"
        params = []
        if type_filter:
            sql += " AND type = ?"
            params.append(type_filter)
        if status_filter:
            sql += " AND status = ?"
            params.append(status_filter)
        if search_query:
            sql += " AND (username LIKE ? OR customer_name LIKE ? OR order_number LIKE ? OR product_code LIKE ? OR brand LIKE ? OR product_name LIKE ?)"
            q = f"%{search_query}%"
            params.extend([q, q, q, q, q, q])
        sql += " ORDER BY created_at DESC LIMIT 200"
        rows = conn.execute(sql, params).fetchall()
        orders = []
        # 상품 링크 조회를 위해 product_db 연결
        product_links = {}
        product_extras = {}
        try:
            from product_db import _conn as prod_conn
            pconn = prod_conn()
            for r in rows:
                code = r["product_code"] or ""
                if code and code not in product_extras:
                    pr = pconn.execute("SELECT link, img_url, product_status FROM products WHERE internal_code=? OR product_code=? LIMIT 1", (code, code)).fetchone()
                    product_extras[code] = {
                        "link": pr["link"] if pr else "",
                        "img": pr["img_url"] if pr else "",
                        "sold_out": (pr["product_status"] == "sold_out") if pr and pr["product_status"] else False,
                    }
            pconn.close()
        except Exception:
            pass
        for r in rows:
            o = {c: r[c] for c in r.keys()}
            extras = product_extras.get(r["product_code"], {})
            o["product_link"] = extras.get("link", "")
            o["product_img"] = extras.get("img", "")
            o["product_sold_out"] = extras.get("sold_out", False)
            # 원가/마진 계산
            pname = o.get("product_name", "") or ""
            pjpy = o.get("price_jpy", 0) or 0
            if "일괄결제" in pname or "다중주문" in pname:
                # 다중주문/일괄결제: 개별 주문들의 합계
                batch_cost = 0
                batch_sell = 0
                batch_ids_str = o.get("product_code", "")
                if batch_ids_str:
                    for bid in batch_ids_str.split(","):
                        bid = bid.strip()
                        if bid.isdigit():
                            br = conn.execute("SELECT price_jpy, price FROM orders WHERE id=?", (int(bid),)).fetchone()
                        elif bid.startswith("ORD"):
                            br = conn.execute("SELECT price_jpy, price FROM orders WHERE order_number=?", (bid,)).fetchone()
                        else:
                            continue
                        if br and br["price_jpy"]:
                            batch_cost += _calc_vintage_cost(br["price_jpy"])
                        if br:
                            try:
                                batch_sell += int("".join(ch for ch in str(br["price"] or "0") if ch.isdigit()) or 0)
                            except Exception:
                                pass
                o["cost_krw"] = batch_cost
                o["margin_krw"] = batch_sell - batch_cost if batch_sell > 0 and batch_cost > 0 else 0
            elif pjpy > 0:
                o["cost_krw"] = _calc_vintage_cost(pjpy)
                try:
                    sell_price = int("".join(c for c in str(o.get("price", "0")) if c.isdigit()) or 0)
                except Exception:
                    sell_price = 0
                o["margin_krw"] = sell_price - o["cost_krw"] if sell_price > 0 else 0
            else:
                o["cost_krw"] = 0
                o["margin_krw"] = 0
            orders.append(o)
        # 현재 적용 환율 + 회원 레벨
        current_rate = get_cached_rate() or 0
        user_levels = {}
        try:
            level_rows = conn.execute("SELECT username, level FROM users").fetchall()
            for lr in level_rows:
                user_levels[lr["username"]] = lr["level"] if "level" in lr.keys() else "b2c"
        except Exception:
            pass
        for o in orders:
            lvl = user_levels.get(o.get("username",""), "b2c")
            o["user_level"] = lvl
            # 주문 시점 환율이 있으면 저장된 가격 사용 (재계산 안 함)
            # 환율 기록이 없는 기존 주문만 현재 환율로 재계산
            saved_rate = o.get("exchange_rate", "") or ""
            pjpy = o.get("price_jpy", 0) or 0
            pname = o.get("product_name", "") or ""
            if not saved_rate and pjpy > 0 and "다중주문" not in pname and "일괄결제" not in pname:
                recalc = _calc_vintage_price(pjpy, lvl)
                if recalc > 0:
                    o["price"] = f"{recalc:,}원"
                    cost = o.get("cost_krw", 0) or 0
                    o["margin_krw"] = recalc - cost if recalc > 0 and cost > 0 else 0
        return jsonify({"ok": True, "orders": orders, "rate": current_rate})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/orders/<int:order_id>/related")
@admin_required
def get_related_orders(order_id):
    """일괄결제 관련 개별 주문 조회"""
    _init_orders_db()
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            return jsonify({"ok": False})
        username = order["username"]
        memo = order["memo"] or ""
        product_code = order["product_code"] or ""
        related = []
        # 0) 다중주문 memo에 JSON 저장된 경우
        if memo and memo.startswith("["):
            try:
                detail = json.loads(memo)
                # product_code에 개별 주문번호가 쉼표로 저장됨
                order_nums = [d.strip() for d in product_code.split(",") if d.strip().startswith("ORD")]
                if order_nums:
                    placeholders = ",".join(["?"] * len(order_nums))
                    rows = conn.execute(f"SELECT * FROM orders WHERE order_number IN ({placeholders}) ORDER BY created_at", order_nums).fetchall()
                    related = [{c: r[c] for c in r.keys()} for r in rows]
            except Exception:
                pass

        # 1) product_code에 개별 주문 ID들이 저장된 경우 (예: "18,19,20,21")
        if not related and product_code and all(p.strip().isdigit() for p in product_code.split(",") if p.strip()):
            ids = [int(p.strip()) for p in product_code.split(",") if p.strip()]
            if ids:
                placeholders = ",".join(["?"] * len(ids))
                rows = conn.execute(f"SELECT * FROM orders WHERE id IN ({placeholders}) ORDER BY created_at", ids).fetchall()
                related = [{c: r[c] for c in r.keys()} for r in rows]
        # 2) 같은 결제 메모로 조회
        if not related and memo and "토스결제" in memo:
            rows = conn.execute(
                "SELECT * FROM orders WHERE username=? AND memo=? AND id!=? ORDER BY created_at",
                (username, memo, order_id)
            ).fetchall()
            related = [{c: r[c] for c in r.keys()} for r in rows]
        # 3) 비슷한 시간대(±5분) 주문
        if not related:
            rows = conn.execute("""
                SELECT * FROM orders WHERE username=? AND id!=? AND type='order'
                AND abs(strftime('%s', created_at) - strftime('%s', ?)) < 300
                ORDER BY created_at
            """, (username, order_id, order["created_at"])).fetchall()
            related = [{c: r[c] for c in r.keys()} for r in rows]
        # 상품 이미지 조회
        # 상품 이미지/링크 + 원가 계산
        try:
            from product_db import _conn as prod_conn
            pconn = prod_conn()
            for o in related:
                code = o.get("product_code", "")
                if code:
                    pr = pconn.execute("SELECT img_url, link FROM products WHERE internal_code=? OR product_code=? LIMIT 1", (code, code)).fetchone()
                    o["product_img"] = pr["img_url"] if pr else ""
                    o["product_link"] = pr["link"] if pr else ""
                else:
                    o["product_img"] = ""
                    o["product_link"] = ""
                # 원가 계산
                pjpy = o.get("price_jpy", 0) or 0
                if pjpy > 0:
                    try:
                        o["cost_krw"] = _calc_vintage_cost(pjpy)
                    except Exception:
                        o["cost_krw"] = 0
                else:
                    o["cost_krw"] = 0
            pconn.close()
        except Exception:
            pass
        return jsonify({"ok": True, "related": related, "count": len(related), "rate": get_cached_rate() or 0})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/order-batch/<int:order_id>")
@login_required
def get_order_batch(order_id):
    """다중주문 세부 내역 (고객용)"""
    username = session.get("username", "")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        order = conn.execute("SELECT * FROM orders WHERE id=? AND username=?", (order_id, username)).fetchone()
        if not order:
            return jsonify({"ok": False})
        product_code = order["product_code"] or ""
        order_nums = [d.strip() for d in product_code.split(",") if d.strip().startswith("ORD")]
        if not order_nums:
            return jsonify({"ok": False, "items": []})
        placeholders = ",".join(["?"] * len(order_nums))
        rows = conn.execute(f"SELECT id, type, brand, product_name, product_code, price, status, order_number, created_at FROM orders WHERE order_number IN ({placeholders}) AND username=? ORDER BY created_at",
                           order_nums + [username]).fetchall()
        items = [{c: r[c] for c in r.keys()} for r in rows]
        try:
            from product_db import _conn as prod_conn
            pconn = prod_conn()
            for it in items:
                code = it.get("product_code", "")
                if code:
                    pr = pconn.execute("SELECT img_url FROM products WHERE internal_code=? OR product_code=? LIMIT 1", (code, code)).fetchone()
                    it["product_img"] = pr["img_url"] if pr else ""
            pconn.close()
        except Exception:
            pass
        return jsonify({"ok": True, "items": items})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/orders/<int:order_id>", methods=["DELETE"])
@admin_required
def delete_order(order_id):
    """주문 삭제"""
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
        conn.commit()
        logger.info(f"주문 삭제: #{order_id}")
        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/orders/<int:order_id>", methods=["PATCH"])
@admin_required
def update_order(order_id):
    """주문 상태/메모 업데이트"""
    data = request.json or {}
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        updates = []
        params = []
        if "status" in data:
            updates.append("status = ?")
            params.append(data["status"])
        if "memo" in data:
            updates.append("memo = ?")
            params.append(data["memo"])
        if "courier" in data:
            updates.append("courier = ?")
            params.append(data["courier"])
        if "tracking_no" in data:
            updates.append("tracking_no = ?")
            params.append(data["tracking_no"])
        if not updates:
            return jsonify({"ok": False})
        params.append(order_id)
        conn.execute(f"UPDATE orders SET {', '.join(updates)} WHERE id = ?", params)
        conn.commit()

        # 주문 상태 변경 시 자동 문자 발송
        if "status" in data:
            new_status = data["status"]
            if new_status in ("confirmed", "processing", "shipped", "completed", "sold_out"):
                try:
                    order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
                    if order:
                        username = order["username"]
                        user = conn.execute("SELECT phone FROM users WHERE username=?", (username,)).fetchone()
                        if user and user["phone"]:
                            from aligo_sms import send_order_notification, load_config
                            load_config()
                            product_name = (order["product_name"] or "") if "product_name" in order.keys() else ""
                            if not product_name:
                                product_name = (order["brand"] or "") if "brand" in order.keys() else ""
                            order_number = (order["order_number"] or "") if "order_number" in order.keys() else ""
                            send_order_notification(user["phone"], order_number, new_status, product_name)
                            logger.info(f"[SMS] 주문 알림 발송: {username} ({user['phone']}) → {new_status}")
                except Exception as e:
                    logger.warning(f"[SMS] 자동 발송 실패: {e}")

        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/ai-analyze", methods=["POST"])
def shop_ai_analyze():
    """AI 상품 분석 (관리자/B2B만)"""
    if not session.get("logged_in"):
        return jsonify({"ok": False, "message": "로그인이 필요합니다"})
    if session.get("role") != "admin" and session.get("level") != "b2b":
        return jsonify({"ok": False, "message": "B2B 회원 전용 기능입니다"})
    data = request.json or {}
    brand = data.get("brand", "")
    name = data.get("name", "")
    desc = data.get("description", "")
    condition = data.get("condition", "")
    price_jpy = data.get("price_jpy", 0)
    try:
        from post_generator import get_ai_config, _call_gemini, _call_claude, _call_openai
        config = get_ai_config()
        provider = config.get("provider", "none")
        if provider == "none":
            return jsonify({"ok": False, "message": "AI가 설정되지 않았습니다"})

        grade_labels = {"NS":"신품/미사용","S":"최상급","A":"양호","B":"사용감 있음","C":"사용감 많음","D":"난있음"}
        prompt = f"""당신은 국내 명품 가격 비교 전문 분석가입니다.
아래 상품의 국내 판매 시세를 분석해주세요.

[상품 정보]
- 브랜드: {brand}
- 상품명: {name}
- 상태: {grade_labels.get(condition, condition)}
- 상품 설명: {desc[:500] if desc else '없음'}

[중요 규칙]
- 일본 엔화 가격은 절대 언급하지 마세요
- 국내 판매 플랫폼 기준 가격만 분석하세요
- 국내 플랫폼 가격이 높게 형성되어 있다는 점을 강조하세요

[분석 요청 항목]

1. 📋 상품 식별
   - 정확한 모델명, 품번, 시즌 추정

2. 📊 국내 판매 시세 비교
   - 다음 플랫폼 기준 판매가를 조사하여 비교표로 작성:
     트렌비 / 구구스 / 필웨이 / 머스트잇 / 스마트스토어 / 리본즈 / 번개장터
   - 각 플랫폼별 예상 판매가 범위 (동일 모델, 유사 상태 기준)
   - 가장 저렴한 곳과 가장 비싼 곳 명시

3. 📈 트렌드 & 인기도
   - 이 모델의 국내 인기도 (높음/보통/낮음)
   - 검색 트렌드 및 수요
   - 시즌성/한정성 여부

4. ⚠️ 구매 시 참고사항
   - 정품 확인 포인트
   - 상태 체크 포인트

각 항목을 간결하게 작성하세요."""

        if provider == "gemini" and config.get("gemini_key"):
            result = _call_gemini(prompt)
        elif provider == "claude" and config.get("claude_key"):
            result = _call_claude(prompt)
        elif provider == "openai" and config.get("openai_key"):
            result = _call_openai(prompt)
        else:
            return jsonify({"ok": False, "message": "AI API 키가 설정되지 않았습니다"})

        return jsonify({"ok": True, "analysis": result})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)})


@app.route(f"{URL_PREFIX}/shop/api/image-search", methods=["POST"])
def shop_image_search():
    """📷 이미지 검색 — AI Vision으로 상품 분석 후 키워드 추출"""
    data = request.json or {}
    image_b64 = data.get("image", "")
    mime_type = data.get("mime_type", "image/jpeg")
    if not image_b64:
        return jsonify({"ok": False, "message": "이미지가 없습니다"})

    try:
        from post_generator import get_ai_config, _get_gemini, _get_openai
        config = get_ai_config()

        prompt = """이 이미지의 상품을 분석해주세요.

다음 정보를 추출하세요:
1. 브랜드명 (영문)
2. 상품 카테고리 (가방/지갑/의류/신발/악세서리 등)
3. 세부 종류 (숄더백/토트백/핸드백/클러치/지갑/자켓/코트 등)
4. 색상
5. 소재 (가죽/캔버스/나일론 등)

반드시 아래 JSON 형식으로만 응답하세요:
{"brand":"브랜드명","category":"카테고리","type":"세부종류","color":"색상","material":"소재","keywords":"검색용 키워드 (공백 구분, 2~4개)"}"""

        import base64
        result_text = ""

        # Gemini Vision 우선 시도
        if config.get("gemini_key"):
            try:
                client = _get_gemini()
                from google.genai import types
                img_part = types.Part.from_bytes(data=base64.b64decode(image_b64), mime_type=mime_type)
                resp = client.models.generate_content(
                    model="gemini-2.0-flash",
                    contents=[prompt, img_part],
                )
                result_text = resp.text or ""
            except Exception as e:
                logger.warning(f"Gemini Vision 실패: {e}")

        # Gemini 실패 시 OpenAI Vision
        if not result_text and config.get("openai_key"):
            try:
                client = _get_openai()
                resp = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{image_b64}", "detail": "low"}},
                        ],
                    }],
                    max_tokens=300,
                )
                result_text = resp.choices[0].message.content or ""
            except Exception as e:
                logger.warning(f"OpenAI Vision 실패: {e}")

        if not result_text:
            return jsonify({"ok": False, "message": "AI 이미지 분석에 실패했습니다. API 키를 확인해주세요."})

        # JSON 파싱
        import json as _json
        # ```json ... ``` 블록 제거
        cleaned = result_text.strip()
        if "```" in cleaned:
            cleaned = cleaned.split("```json")[-1].split("```")[0].strip() if "```json" in cleaned else cleaned.split("```")[1].split("```")[0].strip()
        try:
            parsed = _json.loads(cleaned)
        except Exception:
            # JSON 파싱 실패 시 텍스트에서 키워드 추출
            parsed = {"keywords": cleaned[:100]}

        brand = parsed.get("brand", "")
        category = parsed.get("category", "")
        item_type = parsed.get("type", "")
        color = parsed.get("color", "")
        material = parsed.get("material", "")
        keywords = parsed.get("keywords", "")

        # 검색 키워드 조합
        search_parts = []
        if brand and brand.upper() not in ("UNKNOWN", "불명", "없음", "N/A", ""):
            search_parts.append(brand)
        if item_type:
            search_parts.append(item_type)
        if color and color not in ("없음", "N/A", ""):
            search_parts.append(color)
        if not search_parts and keywords:
            search_parts = keywords.split()[:3]

        search_keyword = " ".join(search_parts)

        analysis_parts = []
        if brand: analysis_parts.append(f"브랜드: {brand}")
        if item_type: analysis_parts.append(f"종류: {item_type}")
        if color: analysis_parts.append(f"색상: {color}")
        if material: analysis_parts.append(f"소재: {material}")
        analysis_text = " | ".join(analysis_parts)

        return jsonify({
            "ok": True,
            "keywords": search_keyword,
            "analysis": analysis_text,
            "raw": parsed,
        })

    except Exception as e:
        logger.error(f"이미지 검색 오류: {e}")
        return jsonify({"ok": False, "message": f"처리 오류: {str(e)}"})


@app.route(f"{URL_PREFIX}/api/product-images")
@admin_required
def api_product_images():
    """상품 이미지 목록 반환"""
    code = request.args.get("code", "").strip()
    if not code:
        return jsonify({"ok": False, "images": []})
    from product_db import _conn
    conn = _conn()
    try:
        row = conn.execute("SELECT img_url, detail_images FROM products WHERE product_code=? OR internal_code=? LIMIT 1",
                           (code, code)).fetchone()
        if not row:
            return jsonify({"ok": False, "images": []})
        images = []
        thumb = row["img_url"] or ""
        if thumb:
            images.append(thumb)
        try:
            import json as _j
            detail = _j.loads(row["detail_images"] or "[]")
            for u in detail:
                if u and u not in images:
                    images.append(u)
        except Exception:
            pass
        return jsonify({"ok": True, "images": images})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/api/product-images/order", methods=["POST"])
@admin_required
def api_save_product_image_order():
    """상품 이미지 순서 저장"""
    data = request.json or {}
    code = data.get("code", "").strip()
    images = data.get("images", [])
    if not code or not images:
        return jsonify({"ok": False, "message": "코드/이미지 필요"})
    from product_db import _conn
    import json as _j
    conn = _conn()
    try:
        # 첫 번째 이미지를 썸네일로, 나머지를 detail_images로 저장
        thumb = images[0] if images else ""
        detail = _j.dumps(images)
        conn.execute("UPDATE products SET img_url=?, detail_images=? WHERE product_code=? OR internal_code=?",
                     (thumb, detail, code, code))
        conn.commit()
        return jsonify({"ok": True, "message": f"이미지 순서 저장 ({len(images)}개)"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/api/blog-fill-images", methods=["POST"])
@admin_required
def blog_fill_images():
    """블로그 이미지 부족분 AI 생성 (Gemini → Pexels 폴백)"""
    data = request.json or {}
    keyword = data.get("keyword", "luxury brand product")
    count = min(data.get("count", 3), 5)
    code = data.get("code", "")

    images = []
    try:
        # 1순위: Gemini AI 생성
        from data_manager import get_path
        gemini_key = ""
        try:
            env_path = os.path.join(os.path.dirname(__file__), ".env")
            with open(env_path, encoding="utf-8") as ef:
                for line in ef:
                    if line.strip().startswith("GEMINI_API_KEY="):
                        gemini_key = line.strip().split("=", 1)[1].strip()
                        break
        except Exception:
            pass

        if gemini_key and len(gemini_key) > 20:
            from google import genai
            from google.genai import types as _gt
            gclient = genai.Client(api_key=gemini_key)
            img_dir = os.path.join(get_path("db"), "blog_images")
            os.makedirs(img_dir, exist_ok=True)

            for i in range(count):
                try:
                    resp = gclient.models.generate_content(
                        model="gemini-2.5-flash-image",
                        contents=f"Professional product photo: {keyword}. Elegant, warm lighting, luxury boutique. IMPORTANT: Do NOT include ANY text in the image.",
                        config=_gt.GenerateContentConfig(response_modalities=["TEXT", "IMAGE"]),
                    )
                    for part in resp.candidates[0].content.parts:
                        if part.inline_data:
                            from datetime import datetime as _dt
                            fn = f"blog_{_dt.now().strftime('%Y%m%d_%H%M%S')}_{i+1}.png"
                            fp = os.path.join(img_dir, fn)
                            with open(fp, "wb") as f:
                                f.write(part.inline_data.data)
                            # 파일 URL로 접근 가능하게
                            images.append(f"/api/blog-image/{fn}")
                            break
                    import time; time.sleep(2)
                except Exception as e:
                    logger.warning(f"Gemini 이미지 {i+1} 실패: {e}")

        # 2순위: Pexels 폴백
        if len(images) < count:
            PEXELS_KEY = "ZMFMszrhmZ9oy5UTEC0XKa7h8JGytGpnLWkoFDcE4bdqxLv7r507JHEe"
            remain = count - len(images)
            try:
                import requests as _req
                r = _req.get("https://api.pexels.com/v1/search",
                    params={"query": keyword, "per_page": remain, "orientation": "landscape"},
                    headers={"Authorization": PEXELS_KEY}, timeout=10)
                if r.status_code == 200:
                    for p in r.json().get("photos", []):
                        images.append(p["src"]["large"])
            except Exception:
                pass

        return jsonify({"ok": True, "images": images})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)})


@app.route(f"{URL_PREFIX}/api/blog-image/<path:filename>")
@admin_required
def serve_blog_image(filename):
    """블로그 AI 이미지 서빙"""
    from flask import send_from_directory
    img_dir = os.path.join(get_path("db"), "blog_images")
    return send_from_directory(img_dir, filename)


@app.route(f"{URL_PREFIX}/shop/api/product-by-code")
def shop_api_product_by_code():
    """internal_code(고유번호) 또는 product_code(원본코드)로 상품 1건 조회"""
    code = request.args.get("code", "").strip()
    if not code:
        return jsonify({"ok": False})
    from product_db import _conn
    conn = _conn()
    try:
        # 1) internal_code 우선 매칭
        row = conn.execute(
            "SELECT * FROM products WHERE internal_code=? AND source_type='vintage' LIMIT 1",
            (code,)
        ).fetchone()
        # 2) 없으면 product_code(원본)로 재조회
        if not row:
            row = conn.execute(
                "SELECT * FROM products WHERE product_code=? AND source_type='vintage' LIMIT 1",
                (code,)
            ).fetchone()
        if not row:
            return jsonify({"ok": False})
        import json as _json
        p = {c: row[c] for c in row.keys()}
        p["detail_images"] = _json.loads(p.get("detail_images") or "[]") if isinstance(p.get("detail_images"), str) else p.get("detail_images", [])
        user_level = session.get("level", "b2c")
        p["price_krw"] = _calc_vintage_price(p.get("price_jpy", 0), "b2b" if user_level == "b2b" else "b2c")
        p["price_b2c"] = _calc_vintage_price(p.get("price_jpy", 0), "b2c")
        p["product_code"] = p.get("internal_code") or p.get("product_code", "")
        # shop API와 동일한 필드 매핑
        p["size_info"] = p.get("color", "")
        p["color_raw"] = p.get("color", "")
        # 한국어 번역 우선 사용
        name_ko = p.get("name_ko", "")
        if name_ko and name_ko.strip() and name_ko != p.get("name", ""):
            p["name"] = name_ko
        desc_ko = p.get("description_ko", "")
        desc_ja = p.get("description", "")
        if desc_ko and desc_ko.strip():
            p["description"] = desc_ko
        elif desc_ja and desc_ja.strip() and desc_ja == "商品のお問い合わせ":
            p["description"] = ""
        return jsonify({"ok": True, "product": p})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/products")
def shop_api_products():
    """고객용 빈티지 상품 API"""
    brand = request.args.get("brand", "").strip()
    condition = request.args.get("condition", "").strip()
    bag_type = request.args.get("bag_type", "").strip()
    color_filter = request.args.get("color", "").strip()
    keyword = request.args.get("keyword", "").strip()
    site = request.args.get("site", "").strip()
    price_min = request.args.get("price_min", 0, type=int)
    price_max = request.args.get("price_max", 0, type=int)
    user_level = session.get("level", "b2c")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 24, type=int)
    sort = request.args.get("sort", "newest")

    # 비회원/미승인 회원은 1페이지만 허용
    is_approved = False
    if session.get("logged_in"):
        if session.get("role") == "admin":
            is_approved = True
        else:
            from user_db import _conn as _uc
            _uconn = _uc()
            try:
                _u = _uconn.execute("SELECT status FROM users WHERE username=?", (session.get("username",""),)).fetchone()
                is_approved = _u and _u["status"] == "approved"
            except Exception:
                pass
            finally:
                _uconn.close()
    if not is_approved and page > 1:
        page = 1

    from product_db import _conn
    conn = _conn()
    try:
        base_where = "source_type='vintage' AND brand NOT LIKE '%OFF%'"
        base_params = []
        if site:
            base_where += " AND site_id = ?"
            base_params.append(site)

        # 사이트 목록
        site_rows = conn.execute(
            f"SELECT site_id, COUNT(*) c FROM products WHERE {base_where} GROUP BY site_id ORDER BY c DESC", base_params
        ).fetchall()
        site_names = {"2ndstreet": "세컨드스트리트", "kindal": "킨달", "brandoff": "브랜드오프", "komehyo": "코메효"}
        sites = [{"id": r["site_id"], "name": site_names.get(r["site_id"], r["site_id"]), "count": r["c"]} for r in site_rows]

        # 브랜드 목록
        brand_rows = conn.execute(
            f"SELECT brand, COUNT(*) c FROM products WHERE {base_where} GROUP BY brand ORDER BY c DESC", base_params
        ).fetchall()
        brands = [{"name": r["brand"], "count": r["c"]} for r in brand_rows]

        # 가방 종류 목록 (상품명 첫 번째 / 앞부분)
        bag_type_map = {
            # 가방
            "ショルダーバッグ": "숄더백", "トートバッグ": "토트백", "リュック": "백팩",
            "ハンドバッグ": "핸드백", "ポーチ": "파우치", "ボストンバッグ": "보스턴백",
            "クラッチバッグ": "클러치", "ウエストバッグ": "웨이스트백",
            "セカンドバッグ": "세컨드백", "バッグ": "가방",
            # 의류
            "スカート": "스커트", "ジャケット": "자켓", "コート": "코트",
            "Tシャツ": "T셔츠", "シャツ": "셔츠", "ブラウス": "블라우스",
            "ワンピース": "원피스", "パンツ": "팬츠", "スラックス": "슬랙스",
            "ニット": "니트", "セーター": "스웨터", "カーディガン": "가디건",
            "パーカー": "후드", "スウェット": "스웨트", "ベスト": "베스트",
            "ドレス": "드레스",
            # 신발
            "パンプス": "펌프스", "スニーカー": "스니커즈", "ブーツ": "부츠",
            "サンダル": "샌들", "ローファー": "로퍼", "ミュール": "뮬",
            "スリッポン": "슬립온", "シューズ": "슈즈",
            # 소품/악세서리
            "財布": "지갑", "ベルト": "벨트", "マフラー": "머플러",
            "帽子": "모자", "キャップ": "캡", "サングラス": "선글라스",
            "ネックレス": "목걸이", "ブレスレット": "팔찌", "リング": "반지",
            "ピアス": "피어싱", "イヤリング": "이어링",
            "スカーフ": "스카프", "ストール": "스톨",
            "キーケース": "키케이스", "キーリング": "키링",
            "コインケース": "코인케이스", "カードケース": "카드케이스",
            "手袋": "장갑", "腕時計": "시계", "ウォッチ": "시계",
            # 기타 가방류
            "ブリーフケース": "브리프케이스", "ビジネスバッグ": "비즈니스백",
            "トラベル": "트래블백", "キャリー": "캐리어",
            "メッセンジャー": "메신저백", "キーポル": "키폴(여행백)",
            # 넥타이/하의
            "ネクタイ": "넥타이", "ボトム": "하의",
            # LV 모델명 (가방으로 분류)
            "ポシェット": "포셰트(미니백)", "アルマ": "가방", "ダヌーブ": "가방",
            "リポーター": "가방", "アマゾン": "가방", "ジェロニモス": "가방",
            "ルーピング": "가방", "ブルームズベリ": "가방", "ナイル": "가방",
            "カバ": "가방", "サック": "가방", "ミュゼット": "가방",
            "ヴィバ": "가방", "アクセサリー": "악세서리",
        }
        bag_rows = conn.execute(f"SELECT name, brand FROM products WHERE {base_where}", base_params).fetchall()
        bag_counts = {}
        for r in bag_rows:
            n = r["name"] or ""
            brand_name = r["brand"] or ""
            matched = False
            for ja, ko in bag_type_map.items():
                if ja in n:
                    bag_counts[ko] = bag_counts.get(ko, 0) + 1
                    matched = True
                    break
            if not matched:
                # LV 모델명 등 브랜드 고유 모델 → 가방으로 자동 분류
                if "_モノグラム" in n or "_ダミエ" in n or "_エピ" in n or "_タイガ" in n or "_ヴェルニ" in n:
                    bag_counts["가방"] = bag_counts.get("가방", 0) + 1
                else:
                    bag_counts["기타"] = bag_counts.get("기타", 0) + 1
        bag_types = [{"name": k, "count": v} for k, v in sorted(bag_counts.items(), key=lambda x: -x[1])]

        # 컬러 목록 (상품명에서 추출)
        _color_map = {
            "BLK": "블랙", "WHT": "화이트", "RED": "레드", "BLU": "블루",
            "GRN": "그린", "NVY": "네이비", "BRW": "브라운", "GRY": "그레이",
            "PNK": "핑크", "YLW": "옐로우", "ORG": "오렌지", "PPL": "퍼플",
            "GLD": "골드", "SLV": "실버", "BGE": "베이지", "CRM": "크림",
            "KHK": "카키", "ボルドー": "보르도",
            "ブラック": "블랙", "ホワイト": "화이트", "レッド": "레드", "ブルー": "블루",
            "グリーン": "그린", "ネイビー": "네이비", "ブラウン": "브라운", "グレー": "그레이",
            "ピンク": "핑크", "イエロー": "옐로우", "オレンジ": "오렌지", "パープル": "퍼플",
            "ゴールド": "골드", "シルバー": "실버", "ベージュ": "베이지", "クリーム": "크림",
            "カーキ": "카키", "マルチカラー": "멀티컬러",
        }
        _color_keys = set(_color_map.keys())
        color_counts = {}
        for r in bag_rows:  # bag_rows 재사용
            n = r["name"] or ""
            for part in n.split("/"):
                p = part.strip()
                if p in _color_keys:
                    ko = _color_map[p]
                    color_counts[ko] = color_counts.get(ko, 0) + 1
        colors = [{"name": k, "count": v} for k, v in sorted(color_counts.items(), key=lambda x: -x[1])]

        # 상품 조회
        sql = f"SELECT * FROM products WHERE {base_where}"
        params = list(base_params)
        if brand:
            brands_list = [b.strip() for b in brand.split(",") if b.strip()]
            if len(brands_list) == 1:
                sql += " AND brand = ?"
                params.append(brands_list[0])
            elif len(brands_list) > 1:
                placeholders = ",".join(["?"] * len(brands_list))
                sql += f" AND brand IN ({placeholders})"
                params.extend(brands_list)
        if condition:
            cond_list = [c.strip() for c in condition.split(",") if c.strip()]
            if len(cond_list) == 1:
                sql += " AND condition_grade = ?"
                params.append(cond_list[0])
            elif len(cond_list) > 1:
                placeholders = ",".join(["?"] * len(cond_list))
                sql += f" AND condition_grade IN ({placeholders})"
                params.extend(cond_list)
        if bag_type:
            bag_list = [b.strip() for b in bag_type.split(",") if b.strip()]
            has_etc = "기타" in bag_list
            ja_keys = []
            for bt in bag_list:
                if bt == "기타":
                    continue
                for ja, ko in bag_type_map.items():
                    if ko == bt:
                        ja_keys.append(ja)
                        break
            if has_etc and not ja_keys:
                # "기타"만 선택 → 모든 매핑 키워드에 안 걸리는 상품
                not_clauses = " AND ".join(["name NOT LIKE ?" for _ in bag_type_map])
                sql += f" AND ({not_clauses})"
                params.extend([f"%{k}%" for k in bag_type_map.keys()])
            elif has_etc and ja_keys:
                # "기타" + 다른 종류 함께 선택
                like_clauses = " OR ".join(["name LIKE ?"] * len(ja_keys))
                not_clauses = " AND ".join(["name NOT LIKE ?" for _ in bag_type_map])
                sql += f" AND (({like_clauses}) OR ({not_clauses}))"
                params.extend([f"%{k}%" for k in ja_keys])
                params.extend([f"%{k}%" for k in bag_type_map.keys()])
            elif len(ja_keys) == 1:
                sql += " AND name LIKE ?"
                params.append(f"%{ja_keys[0]}%")
            elif len(ja_keys) > 1:
                sql += " AND (" + " OR ".join(["name LIKE ?"] * len(ja_keys)) + ")"
                params.extend([f"%{k}%" for k in ja_keys])
        if color_filter:
            color_list = [c.strip() for c in color_filter.split(",") if c.strip()]
            # 한국어 → 일본어/영문 코드 역매핑
            _color_reverse = {}
            for k, v in _color_map.items():
                _color_reverse.setdefault(v, []).append(k)
            color_ja_keys = []
            for cf in color_list:
                color_ja_keys.extend(_color_reverse.get(cf, []))
            if len(color_ja_keys) == 1:
                sql += " AND name LIKE ?"
                params.append(f"%{color_ja_keys[0]}%")
            elif len(color_ja_keys) > 1:
                sql += " AND (" + " OR ".join(["name LIKE ?"] * len(color_ja_keys)) + ")"
                params.extend([f"%{k}%" for k in color_ja_keys])

        if keyword:
            # 여러 키워드 지원 (공백 구분 → AND 조건)
            words = keyword.split()
            for word in words:
                sql += " AND (name LIKE ? OR name_ko LIKE ? OR brand LIKE ? OR brand_ko LIKE ? OR description LIKE ? OR description_ko LIKE ? OR internal_code LIKE ? OR product_code LIKE ? OR color LIKE ? OR material LIKE ? OR condition_grade LIKE ?)"
                params.extend([f"%{word}%"] * 11)

        # 가격대 필터 (한국 원화 → 엔화 역산, 레벨별 마진 적용)
        if price_min > 0 or price_max > 0:
            cfg = _vintage_price_config
            rate = get_cached_rate() or 9.23
            fee = cfg["jp_fee_pct"] / 100
            markup = cfg["buy_markup_pct"] / 100
            margin_key = "margin_b2b_pct" if user_level == "b2b" else "margin_b2c_pct"
            margin = cfg.get(margin_key, 15.0) / 100
            jp_ship = cfg.get("jp_domestic_shipping", 800)
            intl_ship = cfg["intl_shipping_krw"]
            # 역산: krw = (jpy + jp_ship) * (1+fee) * rate * (1+markup) * (1+margin) + intl_ship
            # jpy = (krw - intl_ship) / ((1+fee) * rate * (1+markup) * (1+margin)) - jp_ship
            multiplier = (1 + fee) * rate * (1 + markup) * (1 + margin)
            if price_min > 0 and multiplier > 0:
                jpy_min = max(0, (price_min - intl_ship) / multiplier - jp_ship)
                sql += " AND price_jpy >= ?"
                params.append(int(jpy_min))
            if price_max > 0 and multiplier > 0:
                jpy_max = (price_max - intl_ship) / multiplier - jp_ship
                sql += " AND price_jpy <= ?"
                params.append(int(jpy_max))

        if sort == "price_asc":
            sql += " ORDER BY price_jpy ASC"
        elif sort == "price_desc":
            sql += " ORDER BY price_jpy DESC"
        else:
            sql += " ORDER BY created_at DESC"

        # 총 개수
        count_sql = sql.replace("SELECT *", "SELECT COUNT(*) c", 1)
        total = conn.execute(count_sql, params).fetchone()["c"]

        sql += " LIMIT ? OFFSET ?"
        params.extend([per_page, (page - 1) * per_page])
        rows = conn.execute(sql, params).fetchall()

        products = []
        for r in rows:
            # 상세 이미지에서 상품 이미지만 필터링
            detail_imgs = []
            try:
                import json as _json
                imgs = _json.loads(r["detail_images"]) if r["detail_images"] else []
                detail_imgs = [img for img in imgs if any(ext in img.lower() for ext in ['.jpg', '.jpeg', '.png', '.webp'])]
            except Exception:
                pass

            # 설명: description_ko 우선, 없으면 description
            desc = ""
            desc_ko = r["description_ko"] if "description_ko" in r.keys() else ""
            desc_ja = r["description"] if "description" in r.keys() else ""
            if desc_ko and desc_ko.strip():
                desc = desc_ko
            elif desc_ja and desc_ja.strip() and desc_ja != "商品のお問い合わせ":
                desc = desc_ja

            products.append({
                "id": r["id"],
                "site_id": r["site_id"],
                "name": r["name_ko"] if r["name_ko"] and r["name_ko"] != r["name"] else r["name"],
                "name_ja": r["name"],
                "brand": r["brand"],
                "price_jpy": r["price_jpy"],
                "price_krw": _calc_vintage_price(r["price_jpy"], user_level),
                "price_b2c": _calc_vintage_price(r["price_jpy"], "b2c") if user_level == "b2b" else 0,
                "price_b2b": _calc_vintage_price(r["price_jpy"], "b2b") if user_level == "b2b" else 0,
                "user_level": user_level,
                "img_url": r["img_url"],
                "link": r["link"],
                "condition_grade": r["condition_grade"] if "condition_grade" in r.keys() else "",
                "product_code": r["internal_code"] if "internal_code" in r.keys() and r["internal_code"] else r["product_code"] or "",
                "size_info": r["color"] if "color" in r.keys() else "",
                "material": r["material"] if "material" in r.keys() else "",
                "color_raw": r["color"] if "color" in r.keys() else "",
                "description": desc,
                "detail_images": detail_imgs[:12],
                "sold_out": (r["product_status"] == "sold_out") if "product_status" in r.keys() and r["product_status"] else False,
            })

        return jsonify({
            "products": products,
            "sites": sites,
            "brands": brands,
            "bag_types": bag_types,
            "colors": colors,
            "conditions": ["NS", "S", "A", "B", "C", "D"],
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": (total + per_page - 1) // per_page,
            "exchange_rate": get_cached_rate() or 9.23,
        })
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/logout")
def logout():
    """로그아웃"""
    session.clear()
    return redirect(f"{URL_PREFIX}/login")


# ── 회원관리 API ─────────────────────────────
@app.route(f"{URL_PREFIX}/members")
@admin_required
def get_members():
    """회원 목록 조회"""
    from user_db import _conn
    q = request.args.get("q", "").strip()
    conn = _conn()
    try:
        if q:
            rows = conn.execute(
                "SELECT * FROM users WHERE username LIKE ? OR name LIKE ? OR phone LIKE ? ORDER BY created_at DESC",
                (f"%{q}%", f"%{q}%", f"%{q}%")
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall()
        members = [{"username": r["username"], "name": r["name"], "phone": r["phone"],
                     "role": r["role"], "level": r["level"] if "level" in r.keys() else "b2c",
                     "status": r["status"] if "status" in r.keys() else "approved",
                     "expires_at": r["expires_at"] if "expires_at" in r.keys() else "",
                     "last_login": r["last_login"] if "last_login" in r.keys() else "",
                     "created_at": r["created_at"]} for r in rows]
        return jsonify({"members": members, "total": len(members)})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/members/<path:username>", methods=["DELETE"])
@admin_required
def delete_member(username):
    """회원 삭제"""
    from user_db import _conn
    conn = _conn()
    try:
        result = conn.execute("DELETE FROM users WHERE username = ?", (username,))
        conn.commit()
        if result.rowcount > 0:
            logger.info(f"회원 삭제: {username}")
            return jsonify({"ok": True, "message": f"{username} 삭제 완료"})
        return jsonify({"ok": False, "message": "해당 회원을 찾을 수 없습니다"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/members/<path:username>/status", methods=["POST"])
@admin_required
def change_member_status(username):
    """회원 승인/거절 + 기간 설정"""
    from datetime import timedelta
    data = request.json or {}
    new_status = data.get("status", "")
    period = data.get("period", "")  # free, 1m, 3m, 6m
    if new_status not in ("approved", "rejected", "pending", "suspended"):
        return jsonify({"ok": False, "message": "잘못된 상태"})

    expires_at = ""
    if new_status == "approved" and period:
        if period == "free":
            expires_at = ""  # 무제한
        elif period == "1m":
            expires_at = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
        elif period == "3m":
            expires_at = (datetime.now() + timedelta(days=90)).strftime("%Y-%m-%d")
        elif period == "6m":
            expires_at = (datetime.now() + timedelta(days=180)).strftime("%Y-%m-%d")
        elif period == "12m":
            expires_at = (datetime.now() + timedelta(days=365)).strftime("%Y-%m-%d")

    from user_db import _conn
    conn = _conn()
    try:
        conn.execute("UPDATE users SET status = ?, expires_at = ? WHERE username = ?",
                     (new_status, expires_at, username))
        conn.commit()
        label = {"approved": "승인", "rejected": "거절", "pending": "대기", "suspended": "사용불가"}.get(new_status, new_status)
        exp_msg = f" (만료: {expires_at})" if expires_at else " (무제한)"
        logger.info(f"회원 {label}: {username}{exp_msg}")

        # 상태 변경 시 자동 문자 발송
        try:
            user = conn.execute("SELECT phone, level FROM users WHERE username=?", (username,)).fetchone()
            if user and user["phone"]:
                from aligo_sms import send_sms, load_config
                load_config()
                sms_msgs = {
                    "approved": (
                        f"[TheOne Vintage] 회원가입 승인이 완료되었습니다.\n"
                        f"현재 B2C 등급으로 승인되었습니다.\n"
                        f"B2B 승인을 위해서는 요청/문의 게시판을 통해 문의 부탁드리겠습니다.\n"
                        f"감사합니다.\nhttps://vintage.theone-biz.com"
                    ),
                    "pending": "[TheOne Vintage] 회원 상태가 승인대기로 변경되었습니다.\n관리자 승인 후 이용 가능합니다.\n문의사항은 카카오톡으로 연락해주세요.",
                    "suspended": "[TheOne Vintage] 계정이 사용불가 상태로 변경되었습니다.\n사유 확인은 관리자에게 문의해주세요.",
                    "rejected": "[TheOne Vintage] 회원가입이 거절되었습니다.\n사유 확인은 관리자에게 문의해주세요.",
                }
                msg = sms_msgs.get(new_status)
                if msg:
                    send_sms(user["phone"], msg, title="TheOne Vintage")
                    logger.info(f"[SMS] 회원 상태 알림: {username} → {label} ({user['phone']})")
        except Exception as e:
            logger.warning(f"[SMS] 회원 상태 알림 실패: {e}")

        return jsonify({"ok": True, "message": f"{username} → {label}{exp_msg}"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/members/<path:username>/expiry", methods=["POST"])
@admin_required
def change_member_expiry(username):
    """회원 만료일 직접 변경"""
    data = request.json or {}
    expires_at = data.get("expires_at", "")
    from user_db import _conn
    conn = _conn()
    try:
        conn.execute("UPDATE users SET expires_at=? WHERE username=?", (expires_at, username))
        conn.commit()
        msg = f"{username} → {'무제한' if not expires_at else expires_at + '까지'}"
        logger.info(f"만료일 변경: {msg}")
        return jsonify({"ok": True, "message": msg})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/members/<path:username>/extend", methods=["POST"])
@admin_required
def extend_member(username):
    """회원 기간 연장"""
    from datetime import timedelta
    data = request.json or {}
    period = data.get("period", "3m")
    days_map = {"1m": 30, "3m": 90, "6m": 180, "12m": 365}
    days = days_map.get(period, 90)

    from user_db import _conn
    conn = _conn()
    try:
        row = conn.execute("SELECT expires_at FROM users WHERE username = ?", (username,)).fetchone()
        if not row:
            return jsonify({"ok": False, "message": "해당 회원을 찾을 수 없습니다"})
        # 기존 만료일 기준으로 연장 (이미 만료된 경우 오늘부터)
        current = row["expires_at"] if row["expires_at"] else ""
        if current and current >= datetime.now().strftime("%Y-%m-%d"):
            base = datetime.strptime(current, "%Y-%m-%d")
        else:
            base = datetime.now()
        new_expires = (base + timedelta(days=days)).strftime("%Y-%m-%d")
        conn.execute("UPDATE users SET expires_at = ?, status = 'approved' WHERE username = ?", (new_expires, username))
        conn.commit()
        logger.info(f"회원 기간 연장: {username} → {new_expires}")
        return jsonify({"ok": True, "message": f"{username} → {new_expires}까지 연장"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/members/<path:username>/level", methods=["POST"])
@admin_required
def change_member_level(username):
    """회원 레벨 변경 (b2c/b2b)"""
    data = request.json or {}
    level = data.get("level", "b2c")
    if level not in ("b2c", "b2b"):
        return jsonify({"ok": False, "message": "잘못된 레벨"})
    from user_db import _conn
    conn = _conn()
    try:
        result = conn.execute("UPDATE users SET level = ? WHERE username = ?", (level, username))
        conn.commit()
        if result.rowcount > 0:
            logger.info(f"회원 레벨 변경: {username} → {level}")
            return jsonify({"ok": True, "message": f"{username} → {level.upper()}"})
        return jsonify({"ok": False, "message": "해당 회원을 찾을 수 없습니다"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/members/<path:username>/orders")
@admin_required
def get_member_orders(username):
    """특정 회원의 주문 내역 조회"""
    _init_orders_db()
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        rows = conn.execute("SELECT * FROM orders WHERE username=? ORDER BY created_at DESC LIMIT 100", (username,)).fetchall()
        orders = []
        for r in rows:
            o = {c: r[c] for c in r.keys()}
            # 상품 이미지 조회
            code = o.get("product_code", "")
            if code:
                try:
                    from product_db import _conn as prod_conn
                    pconn = prod_conn()
                    pr = pconn.execute("SELECT img_url FROM products WHERE internal_code=? LIMIT 1", (code,)).fetchone()
                    o["product_img"] = pr["img_url"] if pr else ""
                    pconn.close()
                except Exception:
                    o["product_img"] = ""
            else:
                o["product_img"] = ""
            orders.append(o)
        return jsonify({"ok": True, "orders": orders})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/members/<path:username>/info", methods=["GET"])
@admin_required
def get_member_info(username):
    """회원 상세 정보 조회"""
    from user_db import _conn
    conn = _conn()
    try:
        row = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        if not row:
            return jsonify({"ok": False})
        cols = row.keys()
        return jsonify({"ok": True, "user": {c: (row[c] or "") for c in cols if c != "password_hash"}})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/members/<path:username>/info", methods=["POST"])
@admin_required
def update_member_info(username):
    """회원 배송/통관/사업자 정보 업데이트"""
    data = request.json or {}
    from user_db import _conn
    conn = _conn()
    try:
        fields = {
            "name": data.get("name"),
            "phone": data.get("phone"),
            "postal_code": data.get("postal_code"),
            "address": data.get("address"),
            "address_detail": data.get("address_detail"),
            "customs_id": data.get("customs_id"),
            "business_number": data.get("business_number"),
        }
        updates = []
        params = []
        for k, v in fields.items():
            if v is not None:
                updates.append(f"{k} = ?")
                params.append(v.strip())
        if not updates:
            return jsonify({"ok": False, "message": "변경할 정보 없음"})
        params.append(username)
        conn.execute(f"UPDATE users SET {', '.join(updates)} WHERE username = ?", params)
        conn.commit()
        logger.info(f"회원 정보 수정: {username}")
        return jsonify({"ok": True, "message": "저장 완료"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/shop/api/upload-cert", methods=["POST"])
@login_required
def upload_my_cert():
    """고객이 직접 사업자등록증 업로드"""
    username = session.get("username", "")
    if not username:
        return jsonify({"ok": False, "message": "로그인이 필요합니다"})
    if "file" not in request.files:
        return jsonify({"ok": False, "message": "파일이 없습니다"})
    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "message": "파일이 없습니다"})
    allowed = {".jpg", ".jpeg", ".png", ".pdf"}
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in allowed:
        return jsonify({"ok": False, "message": f"허용 파일: {', '.join(allowed)}"})
    upload_dir = os.path.join(get_path("db"), "certs")
    os.makedirs(upload_dir, exist_ok=True)
    filename = f"{username}_cert{ext}"
    filepath = os.path.join(upload_dir, filename)
    f.save(filepath)
    from user_db import _conn
    conn = _conn()
    try:
        conn.execute("UPDATE users SET business_cert_file = ? WHERE username = ?", (filename, username))
        conn.commit()
        logger.info(f"사업자등록증 업로드 (고객): {username} → {filename}")
        # 텔레그램 알림
        try:
            from notifier import send_telegram
            send_telegram(f"📎 <b>사업자등록증 첨부</b>\n👤 {username}\n📄 {filename}")
        except Exception:
            pass
        return jsonify({"ok": True, "message": "사업자등록증이 업로드되었습니다", "filename": filename})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/members/<path:username>/upload-cert", methods=["POST"])
@admin_required
def upload_member_cert(username):
    """사업자등록증 파일 업로드 (관리자)"""
    if "file" not in request.files:
        return jsonify({"ok": False, "message": "파일이 없습니다"})
    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "message": "파일이 없습니다"})
    # 확장자 제한
    allowed = {".jpg", ".jpeg", ".png", ".pdf"}
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in allowed:
        return jsonify({"ok": False, "message": f"허용 파일: {', '.join(allowed)}"})
    # 저장 경로
    upload_dir = os.path.join(get_path("db"), "certs")
    os.makedirs(upload_dir, exist_ok=True)
    filename = f"{username}_cert{ext}"
    filepath = os.path.join(upload_dir, filename)
    f.save(filepath)
    # DB 업데이트
    from user_db import _conn
    conn = _conn()
    try:
        conn.execute("UPDATE users SET business_cert_file = ? WHERE username = ?", (filename, username))
        conn.commit()
        logger.info(f"사업자등록증 업로드: {username} → {filename}")
        return jsonify({"ok": True, "message": "업로드 완료", "filename": filename})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/members/cert/<path:filename>")
@admin_required
def serve_cert(filename):
    """사업자등록증 파일 서빙"""
    cert_dir = os.path.join(get_path("db"), "certs")
    return send_from_directory(cert_dir, filename)


@app.route(f"{URL_PREFIX}/members/<path:username>/reset-password", methods=["POST"])
@admin_required
def reset_member_password(username):
    """관리자: 고객 비밀번호 초기화/변경"""
    data = request.json or {}
    new_pw = data.get("new_password", "").strip()
    if not new_pw:
        return jsonify({"ok": False, "message": "새 비밀번호를 입력하세요"})
    if len(new_pw) < 4:
        return jsonify({"ok": False, "message": "비밀번호는 4자 이상이어야 합니다"})
    from user_db import _conn
    conn = _conn()
    try:
        row = conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
        if not row:
            return jsonify({"ok": False, "message": "회원을 찾을 수 없습니다"})
        hashed = generate_password_hash(new_pw)
        conn.execute("UPDATE users SET password_hash=? WHERE username=?", (hashed, username))
        conn.commit()
        logger.info(f"회원 비밀번호 초기화: {username}")
        return jsonify({"ok": True, "message": f"{username} 비밀번호가 변경되었습니다"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/admin/change-password", methods=["POST"])
@admin_required
def admin_change_password():
    """관리자 비밀번호 변경"""
    data = request.json or {}
    current = data.get("current", "")
    new_pw = data.get("new_password", "")
    if not current or not new_pw:
        return jsonify({"ok": False, "message": "현재 비밀번호와 새 비밀번호를 입력하세요"})
    if len(new_pw) < 4:
        return jsonify({"ok": False, "message": "새 비밀번호는 4자 이상이어야 합니다"})
    if LOGIN_USERS.get("admin") != current:
        return jsonify({"ok": False, "message": "현재 비밀번호가 틀립니다"})
    # 비밀번호 변경
    LOGIN_USERS["admin"] = new_pw
    # .env 파일에 영구 저장
    from notifier import _save_to_env
    _save_to_env("ADMIN_PASSWORD", new_pw)
    logger.info("🔑 관리자 비밀번호 변경 완료")
    return jsonify({"ok": True, "message": "비밀번호가 변경되었습니다"})


@app.route(f"{URL_PREFIX}/vintage/translate", methods=["POST"])
@admin_required
def translate_vintage_products():
    """빈티지 상품 이름 일괄 한국어 번역"""
    from product_db import _conn
    from translator import translate_vintage_name
    conn = _conn()
    try:
        rows = conn.execute(
            "SELECT id, name, name_ko FROM products WHERE source_type='vintage' AND brand NOT LIKE '%OFF%'"
        ).fetchall()
        updated = 0
        for r in rows:
            name_ja = r["name"] or ""
            old_ko = r["name_ko"] or ""
            # 이미 번역된 것과 원본이 다르면 건너뜀 (수동 수정된 경우)
            new_ko = translate_vintage_name(name_ja)
            if new_ko != old_ko:
                conn.execute("UPDATE products SET name_ko = ? WHERE id = ?", (new_ko, r["id"]))
                updated += 1
        conn.commit()
        push_log(f"🌐 빈티지 상품 번역 완료: {updated}/{len(rows)}개 업데이트")
        return jsonify({"ok": True, "message": f"{updated}개 상품 번역 완료", "total": len(rows)})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)})
    finally:
        conn.close()


# 진행상황 브로드캐스트 (멀티 클라이언트 SSE 지원)
_log_subscribers = []          # 각 클라이언트별 queue 리스트
_log_subscribers_lock = threading.Lock()
_log_history = []              # 최근 로그 100개 보관 (새 접속 시 전송)
_LOG_HISTORY_MAX = 100

# 현재 실행 상태
status = {
    "scraping": False,
    "uploading": False,
    "last_scrape": None,
    "last_upload": None,
    "product_count": 0,
    "uploaded_count": 0,
    "paused": False,      # 일시정지 플래그
    "stop_requested": False,  # 중단 요청 플래그
}
_upload_lock = threading.Lock()  # 업로드 동시 실행 방지 락


def push_log(msg: str):
    """실시간 로그를 모든 접속 클라이언트에게 브로드캐스트"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    full_msg = f"[{timestamp}] {msg}"

    # 히스토리에 저장
    _log_history.append(full_msg)
    if len(_log_history) > _LOG_HISTORY_MAX:
        _log_history.pop(0)

    # 모든 구독자에게 전송
    with _log_subscribers_lock:
        dead = []
        for q in _log_subscribers:
            try:
                q.put_nowait(full_msg)
            except Exception:
                dead.append(q)
        for q in dead:
            _log_subscribers.remove(q)

    logger.info(msg)


def _subscribe_logs() -> queue.Queue:
    """새 SSE 클라이언트 등록 — 개별 큐 반환"""
    q = queue.Queue()
    # 최근 히스토리 전송 (접속 즉시 이전 로그 확인 가능)
    for msg in _log_history:
        q.put_nowait(msg)
    with _log_subscribers_lock:
        _log_subscribers.append(q)
    return q


def _unsubscribe_logs(q: queue.Queue):
    """SSE 클라이언트 해제"""
    with _log_subscribers_lock:
        if q in _log_subscribers:
            _log_subscribers.remove(q)


# =============================================
# 스크래핑 / 업로드 실행 함수 (백그라운드)
# =============================================

def run_scrape(site_id="xebio", category_id="sale", keyword="", pages="", brand_code=""):
    """백그라운드 스레드에서 스크래핑 실행 (사이트별 크롤러 디스패치)"""
    if status["scraping"]:
        push_log("⚠️ 이미 스크래핑이 진행 중입니다")
        push_log("   💡 이전 작업이 비정상 종료된 경우 '리셋' 버튼을 눌러주세요")
        return
    # 중단 요청 초기화
    status["stop_requested"] = False

    push_log(f"🔧 run_scrape 시작: site={site_id}, cat={category_id}, brand={brand_code}, pages={pages}")
    status["scraping"] = True
    try:
        from site_config import get_site
        site_info = get_site(site_id)
        source_type = site_info.get("source_type", "sports") if site_info else "sports"
        push_log(f"   📡 source_type={source_type}, site_info={'있음' if site_info else '없음'}")

        if source_type == "vintage":
            from secondst_crawler import scrape_2ndstreet, set_app_status as set_2nd_status
            set_2nd_status(status)
            push_log("   🚀 2ndstreet 크롤러 시작...")
            result = asyncio.run(scrape_2ndstreet(
                status_callback=push_log,
                category=category_id,
                keyword=keyword,
                pages=pages,
                brand_code=brand_code,
            ))
            if isinstance(result, dict):
                products = []
                product_count = result.get("total_saved", 0)
            else:
                products = result or []
                product_count = len(products)
        else:
            push_log("   🚀 Xebio 크롤러 시작 (Playwright 브라우저 열기)...")
            products = asyncio.run(scrape_nike_sale(
                status_callback=push_log,
                site_id=site_id,
                category_id=category_id,
                keyword=keyword,
                pages=pages,
                brand_code=brand_code,
            ))
            product_count = len(products)
        status["product_count"] = product_count
        status["last_scrape"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        push_log(f"🎉 스크래핑 완료: {product_count}개 상품 수집")
        # 수집 이력 저장
        try:
            from scrape_history import add_history
            from site_config import get_brands as get_site_brands
            brand_name = ""
            if brand_code:
                brands_map = get_site_brands(site_id)
                brand_name = brands_map.get(brand_code, brand_code)
            add_history(
                site_id=site_id,
                category_id=category_id or "전체",
                product_count=product_count,
                keyword=keyword or "",
                brand=brand_name or "",
            )
        except Exception as e:
            logger.warning(f"수집 이력 저장 실패: {e}")
    except Exception as e:
        import traceback
        push_log(f"❌ 스크래핑 오류: {e}")
        push_log(f"   📋 상세: {traceback.format_exc()[-500:]}")
        logger.error(f"스크래핑 오류 상세:\n{traceback.format_exc()}")
    finally:
        status["scraping"] = False
        push_log("🔧 run_scrape 종료 (scraping=False)")

        # [Windows] 수집 완료 후 교대 최신화 체크 실행
        import platform
        if platform.system() == "Windows" and not _freshness_status.get("running"):
            try:
                push_log("🔍 교대 실행: 기존 상품 300개 최신화 체크 시작")
                run_interleaved_check()
                push_log("🔍 교대 체크 완료")
            except Exception as e:
                push_log(f"🔍 교대 체크 오류: {e}")


def _shuffle_by_brand(products: list) -> list:
    """브랜드가 연속되지 않도록 섞기 — 라운드로빈 방식 (브랜드 내 수집 순서 유지)"""
    import random
    from collections import defaultdict

    brand_buckets = defaultdict(list)
    for p in products:
        brand = (p.get("brand_ko") or p.get("brand") or "기타").strip()
        brand_buckets[brand].append(p)

    # 브랜드 내부는 수집 순서 그대로 유지 (shuffle 안 함)
    # 브랜드 순서만 랜덤
    brand_keys = list(brand_buckets.keys())
    random.shuffle(brand_keys)

    # 라운드로빈으로 섞기
    result = []
    while brand_keys:
        empty_brands = []
        for brand in brand_keys:
            if brand_buckets[brand]:
                result.append(brand_buckets[brand].pop(0))
            else:
                empty_brands.append(brand)
        for b in empty_brands:
            brand_keys.remove(b)

    return result


def run_upload(max_upload=None, shuffle_brands=False, checked_codes=None, delay_min=13, delay_max=15, source_type="sports", filter_brand="ALL", filter_category="ALL", filter_min_grade=""):
    """백그라운드 스레드에서 업로드 실행

    우선순위:
    1. checked_codes에 있는 상품 우선
    2. 나머지는 대기(待機) 상품에서 랜덤으로 채움
    3. max_upload 수량만큼만 업로드
    """
    if not _upload_lock.acquire(blocking=False):
        push_log("⚠️ 이미 업로드가 진행 중입니다 (락)")
        return
    if status["uploading"]:
        _upload_lock.release()
        push_log("⚠️ 이미 업로드가 진행 중입니다")
        return

    if source_type == "vintage":
        # 빈티지: DB에서 직접 로드
        push_log("📦 빈티지 상품 DB에서 로드 중...")
        products = []
        try:
            from product_db import _conn
            conn = _conn()
            import json as _json
            sql = """
                SELECT * FROM products WHERE source_type='vintage' AND in_stock=1
                AND (cafe_status IS NULL OR cafe_status='' OR cafe_status='대기')
            """
            params = []
            if filter_brand and filter_brand != "ALL":
                sql += " AND brand = ?"
                params.append(filter_brand)
            if filter_category and filter_category != "ALL":
                # 카테고리는 상품명(name)에 일본어 키워드로 매칭
                cat_keywords = {
                    "bag": ["バッグ","ショルダー","トート","リュック","ハンド","ポーチ","ボストン","クラッチ","ウエスト"],
                    "clothing": ["ジャケット","コート","シャツ","ブラウス","ワンピース","パンツ","スラックス","ニット","セーター","カーディガン","パーカー","スウェット","ベスト","Tシャツ","ドレス","スカート"],
                    "shoes": ["シューズ","スニーカー","ブーツ","サンダル","パンプス","ローファー"],
                    "watch": ["時計","ウォッチ"],
                    "accessory": ["財布","ベルト","マフラー","帽子","サングラス","ネックレス","ブレスレット","リング","アクセサリー","ストール","スカーフ"],
                }
                kws = cat_keywords.get(filter_category, [])
                if kws:
                    like_clauses = " OR ".join(["name LIKE ?" for _ in kws])
                    sql += f" AND ({like_clauses})"
                    params.extend([f"%{k}%" for k in kws])
            # 등급 필터 (선택 등급 이상만)
            if filter_min_grade:
                grade_rank = {"NS": 0, "S": 1, "A": 2, "B": 3, "C": 4, "D": 5}
                threshold = grade_rank.get(filter_min_grade, 99)
                allowed = [g for g, r in grade_rank.items() if r <= threshold]
                if allowed:
                    grade_placeholders = ",".join(["?"] * len(allowed))
                    sql += f" AND condition_grade IN ({grade_placeholders})"
                    params.extend(allowed)
            sql += " ORDER BY created_at DESC"
            rows = conn.execute(sql, params).fetchall()
            for r in rows:
                p = {c: r[c] for c in r.keys()}
                p["source_type"] = "vintage"
                # internal_code 없으면 자동 생성
                if not p.get("internal_code"):
                    try:
                        from product_db import _generate_internal_code
                        new_code = _generate_internal_code(p.get("site_id", "2ndstreet"))
                        conn.execute("UPDATE products SET internal_code=? WHERE id=?", (new_code, p["id"]))
                        conn.commit()
                        p["internal_code"] = new_code
                    except Exception:
                        pass
                p["product_code"] = p.get("internal_code") or p.get("product_code", "")
                p["name_ko"] = p.get("name_ko") or p.get("name", "")
                try:
                    p["detail_images"] = _json.loads(p.get("detail_images") or "[]")
                except Exception:
                    p["detail_images"] = []
                products.append(p)
            conn.close()
            push_log(f"📦 빈티지 대기 상품: {len(products)}개 로드 완료 (브랜드={filter_brand}, 카테고리={filter_category})")
        except Exception as e:
            push_log(f"❌ 빈티지 상품 로드 실패: {e}")
            _upload_lock.release()
            return
    else:
        products = load_latest_products()
        # latest.json 상품은 모두 sports
        for p in products:
            if "source_type" not in p:
                p["source_type"] = "sports"

        # 빅데이터 DB 미업로드 상품 병합 (스포츠만)
        try:
            from product_db import get_unuploaded_products
            db_products = get_unuploaded_products(source_type="sports")
            existing_codes = {p.get("product_code", "") for p in products if p.get("product_code")}
            for dp in db_products:
                if dp.get("product_code") and dp["product_code"] not in existing_codes:
                    existing_codes.add(dp["product_code"])
                    products.append(dp)
        except Exception as e:
            logger.warning(f"DB 상품 병합 실패: {e}")

        # 빈티지 상품 제외
        products = [p for p in products if (p.get("source_type") or "sports") == "sports"]

    if not products:
        _upload_lock.release()
        push_log("⚠️ 업로드할 상품이 없습니다. 먼저 스크래핑을 실행하세요")
        return

    import random as _random

    # 체크된 상품과 대기 상품 분리 (product_code 중복 제거)
    checked_set = set(checked_codes) if checked_codes else set()
    # 빈 문자열 제거
    checked_set.discard("")
    checked_products = []
    waiting_products = []
    seen_codes = set()

    push_log(f"📋 업로드 요청: max_upload={max_upload}, checked_codes={len(checked_set)}개, shuffle={shuffle_brands}")
    if checked_set:
        push_log(f"   ✅ 체크된 품번: {', '.join(list(checked_set)[:5])}{'...' if len(checked_set) > 5 else ''}")

    for p in products:
        code = p.get("product_code", "")
        if code and code in seen_codes:
            continue  # 중복 product_code 건너뜀
        if code:
            seen_codes.add(code)
        is_waiting = (p.get("cafe_status") or "대기") == "대기"
        if code and code in checked_set:
            checked_products.append(p)
        elif is_waiting:
            waiting_products.append(p)

    push_log(f"   📊 체크 매칭: {len(checked_products)}개, 대기 상품: {len(waiting_products)}개")

    # 업로드 대상 결정
    if checked_set:
        # 체크된 상품이 있는 경우 → 체크된 것만 업로드 (랜덤 추가 안 함)
        selected = checked_products
        push_log(f"📋 체크된 상품 {len(selected)}개 업로드")
    else:
        # 체크 없음
        if max_upload:
            # 수량만 지정 → 대기 상품에서 랜덤
            _random.shuffle(waiting_products)
            selected = waiting_products[:max_upload]
            push_log(f"📋 대기 상품에서 랜덤 {len(selected)}개 업로드")
        else:
            # 체크도 수량도 없음 → 기존 방식 (선택된 상품)
            selected = [p for p in products if p.get("selected", True)]
            push_log(f"📋 선택된 상품 {len(selected)}개 업로드")

    if not selected:
        _upload_lock.release()
        push_log("⚠️ 업로드할 상품이 없습니다")
        return

    # 브랜드 랜덤 섞기
    if shuffle_brands:
        selected = _shuffle_by_brand(selected)
        brands_order = [p.get("brand_ko") or p.get("brand", "") for p in selected[:10]]
        push_log(f"🔀 브랜드 랜덤 적용: {' → '.join(brands_order[:5])}...")

    # 활성 네이버 계정의 쿠키 경로 결정
    naver_data = _load_naver_accounts()
    active_slot = naver_data.get("active", 1)
    active_cookie = _get_cookie_path(active_slot)
    active_id = naver_data.get("accounts", {}).get(str(active_slot), {}).get("naver_id", "")
    if active_id:
        push_log(f"👤 네이버 계정: {active_id} (슬롯 {active_slot})")
    else:
        push_log(f"👤 네이버 계정: 기본 쿠키 사용 (슬롯 {active_slot})")

    push_log(f"📤 총 {len(selected)}개 업로드 시작")
    status["uploading"] = True
    try:
        count = asyncio.run(upload_products(
            products=selected,
            status_callback=push_log,
            max_upload=max_upload,
            delay_min=delay_min,
            delay_max=delay_max,
            on_single_success=_on_single_upload_success,
            cookie_path=active_cookie,
        ))
        status["uploaded_count"] = count
        status["last_upload"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _save_upload_history(selected[:count])

        # 전부 "이미 게시됨"으로 0개 성공 시 → 새 상품 리스트로 재시도
        if count == 0 and len(selected) > 0 and not checked_set:
            push_log("🔄 전부 이미 게시됨 — 새 상품 리스트로 재시도")
            try:
                from product_db import get_unuploaded_products
                retry_list = get_unuploaded_products(source_type=source_type)
                # 브랜드 필터 적용
                if filter_brand and filter_brand != "ALL":
                    retry_list = [p for p in retry_list
                                  if p.get("brand", "") == filter_brand
                                  or p.get("brand_ko", "") == filter_brand]
                if shuffle_brands:
                    retry_list = _shuffle_by_brand(retry_list)
                else:
                    _random.shuffle(retry_list)
                retry_upload = retry_list[:max_upload or 5]
                if retry_upload:
                    push_log(f"🔄 재시도: {len(retry_upload)}개 새 상품 발견")
                    count2 = asyncio.run(upload_products(
                        products=retry_upload,
                        status_callback=push_log,
                        max_upload=max_upload,
                        delay_min=delay_min, delay_max=delay_max,
                        on_single_success=_on_single_upload_success,
                        cookie_path=active_cookie,
                    ))
                    status["uploaded_count"] = count2
                    _save_upload_history(retry_upload[:count2])
                    push_log(f"🎉 재시도 완료: {count2}개 성공")
                else:
                    push_log("⚠️ 재시도할 새 상품 없음 — 모든 상품이 이미 게시됨")
            except Exception as e2:
                push_log(f"🔄 재시도 오류: {e2}")
        else:
            push_log(f"🎉 업로드 완료: {count}개 성공")
    except Exception as e:
        push_log(f"❌ 업로드 오류: {e}")
    finally:
        status["uploading"] = False
        _upload_lock.release()


def _save_upload_history(uploaded_products: list):
    """업로드된 상품을 히스토리에 저장 (중복 체크용)"""
    history_path = os.path.join(DB_DIR, "uploaded_history.json")
    os.makedirs(DB_DIR, exist_ok=True)
    if os.path.exists(history_path):
        with open(history_path, "r", encoding="utf-8") as f:
            history = json.load(f)
    else:
        history = []

    for p in uploaded_products:
        history.append({
            "product_code": p.get("product_code", ""),
            "name": p.get("name", ""),
            "price_jpy": p.get("price_jpy", 0),
            "brand": p.get("brand", ""),
            "uploaded_at": datetime.now().isoformat(),
        })

    with open(history_path, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def _on_single_upload_success(product: dict):
    """상품 1개 업로드 성공 시 즉시 latest.json + DB에 완료 표시 (중복 업로드 방지)"""
    try:
        code = product.get("product_code", "")
        if not code:
            return
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

        # latest.json에 즉시 반영
        products = load_latest_products()
        changed = False
        for p in products:
            if p.get("product_code") == code:
                p["cafe_status"] = "완료"
                p["cafe_uploaded"] = True
                p["cafe_uploaded_at"] = now_str
                changed = True
        if changed:
            from xebio_search import save_products
            save_products(products)

        # DB에도 반영
        try:
            from product_db import update_cafe_status
            update_cafe_status(code, "완료", now_str)
        except Exception:
            pass

        logger.info(f"✅ 즉시 완료 표시: {code}")
    except Exception as e:
        logger.warning(f"즉시 완료 표시 실패: {e}")


def _mark_uploaded_products(uploaded_products: list):
    """업로드 완료된 상품에 cafe_status='완료' 표시 후 latest.json + DB 저장"""
    try:
        uploaded_codes = {p.get("product_code") for p in uploaded_products if p.get("product_code")}
        if not uploaded_codes:
            return

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        products = load_latest_products()
        changed = False
        for p in products:
            if p.get("product_code") in uploaded_codes:
                p["cafe_status"] = "완료"
                p["cafe_uploaded"] = True
                p["cafe_uploaded_at"] = now_str
                changed = True

        if changed:
            from xebio_search import save_products
            save_products(products)
            logger.info(f"✅ {len(uploaded_codes)}개 상품 업로드 완료 표시")

        # 빅데이터 DB에도 반영
        try:
            from product_db import update_cafe_status
            for code in uploaded_codes:
                update_cafe_status(code, "완료", now_str)
        except Exception as e:
            logger.warning(f"DB 상태 업데이트 실패: {e}")
    except Exception as e:
        logger.warning(f"업로드 완료 표시 실패: {e}")


def run_auto_pipeline():
    """자동 모드: 스크래핑 → 업로드 순서로 실행"""
    push_log("⏰ 자동 실행 시작 (스크래핑 → 업로드)")
    run_scrape()
    if status["product_count"] > 0:
        run_upload()


def run_scheduled_upload(slot_id: str, brand: str, quantity: int):
    """스케줄 슬롯에 의한 자동 카페 업로드"""
    if not _upload_lock.acquire(blocking=False):
        push_log(f"⚠️ [{slot_id}] 이미 업로드가 진행 중이라 스킵합니다 (락)")
        return
    if status["uploading"]:
        _upload_lock.release()
        push_log(f"⚠️ [{slot_id}] 이미 업로드가 진행 중이라 스킵합니다")
        return

    products = load_latest_products()
    for p in products:
        if "source_type" not in p:
            p["source_type"] = "sports"

    # 빅데이터 DB 미업로드 상품 병합 (스포츠만)
    try:
        from product_db import get_unuploaded_products
        db_products = get_unuploaded_products(source_type="sports")
        existing_codes = {p.get("product_code", "") for p in products if p.get("product_code")}
        for dp in db_products:
            if dp.get("product_code") and dp["product_code"] not in existing_codes:
                existing_codes.add(dp["product_code"])
                products.append(dp)
    except Exception as e:
        logger.warning(f"DB 상품 병합 실패: {e}")

    # 빈티지 상품 제외
    products = [p for p in products if (p.get("source_type") or "sports") == "sports"]

    if not products:
        _upload_lock.release()
        push_log(f"⏰ [{slot_id}] 업로드할 상품이 없습니다")
        return

    # 대기 상태만 필터 (product_code 중복 제거)
    seen_codes = set()
    waiting = []
    for p in products:
        code = p.get("product_code", "")
        if code and code in seen_codes:
            continue
        if code:
            seen_codes.add(code)
        if (p.get("cafe_status") or "대기") == "대기":
            waiting.append(p)

    # 브랜드 필터
    if brand and brand != "ALL":
        waiting = [p for p in waiting
                   if (p.get("brand_ko") or "").strip() == brand
                   or (p.get("brand") or "").strip() == brand]

    if not waiting:
        _upload_lock.release()
        push_log(f"⏰ [{slot_id}] 조건에 맞는 대기 상품이 없습니다 (브랜드: {brand})")
        return

    # 브랜드 ALL이면 랜덤 섞기
    if brand == "ALL":
        waiting = _shuffle_by_brand(waiting)

    # 수량 제한
    to_upload = waiting[:quantity]

    # 활성 네이버 계정의 쿠키 경로
    naver_data = _load_naver_accounts()
    active_slot = naver_data.get("active", 1)
    active_cookie = _get_cookie_path(active_slot)
    active_id = naver_data.get("accounts", {}).get(str(active_slot), {}).get("naver_id", "")
    push_log(f"⏰ [{slot_id}] 자동 업로드 시작 — {brand} {len(to_upload)}개 (계정: {active_id or '기본'})")

    status["uploading"] = True
    try:
        count = asyncio.run(upload_products(
            products=to_upload,
            status_callback=push_log,
            max_upload=quantity,
            delay_min=20,
            delay_max=30,
            on_single_success=_on_single_upload_success,
            cookie_path=active_cookie,
        ))
        status["uploaded_count"] = count
        status["last_upload"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _save_upload_history(to_upload[:count])

        # 전부 "이미 게시됨"으로 0개 성공 시 → 새 상품 리스트로 재시도
        if count == 0 and len(to_upload) > 0:
            push_log(f"⏰ [{slot_id}] 전부 이미 게시됨 — 새 상품 리스트로 재시도")
            # 이미 게시된 상품 제외하고 다시 대기 목록 조회
            try:
                from product_db import get_unuploaded_products
                retry_products = get_unuploaded_products(source_type="sports")
                # 브랜드 필터
                if brand and brand != "ALL":
                    retry_products = [p for p in retry_products
                                      if (p.get("brand_ko") or "").strip() == brand
                                      or (p.get("brand") or "").strip() == brand]
                if brand == "ALL":
                    retry_products = _shuffle_by_brand(retry_products)
                retry_upload = retry_products[:quantity]
                if retry_upload:
                    push_log(f"⏰ [{slot_id}] 재시도: {len(retry_upload)}개 새 상품 발견")
                    count2 = asyncio.run(upload_products(
                        products=retry_upload,
                        status_callback=push_log,
                        max_upload=quantity,
                        delay_min=20, delay_max=30,
                        on_single_success=_on_single_upload_success,
                        cookie_path=active_cookie,
                    ))
                    status["uploaded_count"] = count2
                    _save_upload_history(retry_upload[:count2])
                    push_log(f"⏰ [{slot_id}] 재시도 완료: {count2}개 성공")
                else:
                    push_log(f"⏰ [{slot_id}] 재시도할 새 상품 없음")
            except Exception as e2:
                push_log(f"⏰ [{slot_id}] 재시도 오류: {e2}")
        else:
            push_log(f"⏰ [{slot_id}] 자동 업로드 완료: {count}개 성공")
    except Exception as e:
        push_log(f"❌ [{slot_id}] 자동 업로드 오류: {e}")
    finally:
        status["uploading"] = False
        _upload_lock.release()


# =============================================
# 자동 스케줄러 설정
# =============================================

scheduler = BackgroundScheduler()


def _register_schedule_jobs():
    """스케줄 설정 파일을 읽어 APScheduler 잡 등록/갱신"""
    slots = load_schedule()
    for slot in slots:
        job_id = f"cafe_schedule_{slot['id']}"
        # 기존 잡 제거
        if scheduler.get_job(job_id):
            scheduler.remove_job(job_id)
        if slot.get("enabled"):
            scheduler.add_job(
                func=run_scheduled_upload,
                trigger="cron",
                hour=slot["hour"],
                minute=slot["minute"],
                id=job_id,
                name=f"카페 자동업로드 [{slot['label']}] {slot['hour']:02d}:{slot['minute']:02d}",
                args=[slot["id"], slot.get("brand", "ALL"), slot.get("quantity", 5)],
                replace_existing=True,
            )
            logger.info(f"📅 스케줄 등록: {slot['label']} {slot['hour']:02d}:{slot['minute']:02d} (브랜드={slot.get('brand','ALL')}, 수량={slot.get('quantity',5)})")


def _register_vt_schedule_jobs():
    """빈티지 카페 업로드 스케줄 잡 등록/갱신"""
    from cafe_schedule import load_vt_schedule
    slots = load_vt_schedule()
    for slot in slots:
        job_id = f"vt_cafe_{slot['id']}"
        if scheduler.get_job(job_id):
            scheduler.remove_job(job_id)
        if slot.get("enabled"):
            scheduler.add_job(
                func=run_vt_scheduled_upload,
                trigger="cron",
                hour=slot["hour"],
                minute=slot["minute"],
                id=job_id,
                name=f"빈티지 카페 [{slot['label']}] {slot['hour']:02d}:{slot['minute']:02d}",
                args=[slot["id"], slot.get("brand", "ALL"), slot.get("category", "ALL"), slot.get("min_grade", ""), slot.get("quantity", 3)],
                replace_existing=True,
            )
            logger.info(f"📅 빈티지 스케줄 등록: {slot['label']} {slot['hour']:02d}:{slot['minute']:02d} 브랜드={slot.get('brand','ALL')} 카테고리={slot.get('category','ALL')} 최소등급={slot.get('min_grade','전체')}")


def run_vt_scheduled_upload(slot_id: str, brand: str, category: str, min_grade: str, quantity: int):
    """빈티지 자동 카페 업로드"""
    grade_label = f"등급={min_grade}이상" if min_grade else "등급=전체"
    push_log(f"⏰ [빈티지/{slot_id}] 자동 업로드 시작 — 브랜드={brand}, 카테고리={category}, {grade_label}, {quantity}개")
    run_upload(max_upload=quantity, shuffle_brands=(brand == "ALL"),
               checked_codes=None, delay_min=20, delay_max=30, source_type="vintage",
               filter_brand=brand, filter_category=category, filter_min_grade=min_grade)
    push_log(f"⏰ [빈티지/{slot_id}] 자동 업로드 완료")


def _register_check_schedule_job():
    """업로드 체크 자동 확인 스케줄 잡 등록/갱신"""
    job_id = "upload_check_auto"
    if scheduler.get_job(job_id):
        scheduler.remove_job(job_id)
    sched = load_check_schedule()
    if sched.get("enabled"):
        scheduler.add_job(
            func=lambda: threading.Thread(target=_run_upload_check, daemon=True).start(),
            trigger="cron",
            hour=sched["hour"],
            minute=sched["minute"],
            id=job_id,
            name=f"업로드 체크 자동확인 {sched['hour']:02d}:{sched['minute']:02d}",
            replace_existing=True,
        )
        logger.info(f"📅 체크 스케줄 등록: {sched['hour']:02d}:{sched['minute']:02d}")


def run_scheduled_scrape(task_id, site_id, category_id, brand_code, keyword, pages):
    """스케줄에 의한 자동 수집"""
    from site_config import get_site, get_brands as get_site_brands
    brand_name = ""
    if brand_code:
        brands = get_site_brands(site_id)
        brand_name = brands.get(brand_code, brand_code)
    brand_msg = f" [{brand_name}]" if brand_name else ""
    push_log(f"⏰ [{task_id}] 자동 수집 시작{brand_msg}")
    try:
        run_scrape(
            site_id=site_id,
            category_id=category_id,
            keyword=keyword,
            pages=pages,
            brand_code=brand_code,
        )
    except Exception as e:
        push_log(f"❌ [{task_id}] 자동 수집 오류: {e}")


def run_scheduled_check(task_id, brand_name):
    """스케줄에 의한 자동 업로드 체크"""
    brand_filter = brand_name if brand_name and brand_name != "ALL" else ""
    push_log(f"⏰ [{task_id}] 자동 업로드 체크 시작 (브랜드: {brand_name or 'ALL'})")
    try:
        _run_upload_check(brand_filter=brand_filter)
    except Exception as e:
        push_log(f"❌ [{task_id}] 자동 체크 오류: {e}")


def run_scheduled_combo(task_id, site_id, category_id, brand_code, brand_name, keyword, pages):
    """스케줄에 의한 콤보 (수집 → 체크)"""
    push_log(f"⏰ [{task_id}] 콤보 시작: 수집 → 체크 (브랜드: {brand_name or 'ALL'})")
    try:
        # 1단계: 수집
        run_scrape(
            site_id=site_id,
            category_id=category_id,
            keyword=keyword,
            pages=pages,
            brand_code=brand_code,
        )
        push_log(f"⏰ [{task_id}] 수집 완료, 업로드 체크 시작...")
        # 2단계: 체크
        brand_filter = brand_name if brand_name and brand_name != "ALL" else ""
        _run_upload_check(brand_filter=brand_filter)
        push_log(f"⏰ [{task_id}] 콤보 완료!")
    except Exception as e:
        push_log(f"❌ [{task_id}] 콤보 오류: {e}")


def _register_task_schedule_jobs():
    """자동 작업 스케줄 (수집/체크/콤보) 잡 등록/갱신"""
    slots = load_task_schedule()
    for slot in slots:
        job_id = f"task_schedule_{slot['id']}"
        if scheduler.get_job(job_id):
            scheduler.remove_job(job_id)
        if slot.get("enabled"):
            task_type = slot.get("type", "scrape")
            if task_type == "scrape":
                func = lambda s=slot: threading.Thread(
                    target=run_scheduled_scrape,
                    args=(s["id"], s["site_id"], s["category_id"], s.get("brand_code", ""), s.get("keyword", ""), s.get("pages", "")),
                    daemon=True
                ).start()
            elif task_type == "check":
                func = lambda s=slot: threading.Thread(
                    target=run_scheduled_check,
                    args=(s["id"], s.get("brand_name", "ALL")),
                    daemon=True
                ).start()
            elif task_type == "combo":
                func = lambda s=slot: threading.Thread(
                    target=run_scheduled_combo,
                    args=(s["id"], s["site_id"], s["category_id"], s.get("brand_code", ""), s.get("brand_name", "ALL"), s.get("keyword", ""), s.get("pages", "")),
                    daemon=True
                ).start()
            else:
                continue

            scheduler.add_job(
                func=func,
                trigger="cron",
                hour=slot["hour"],
                minute=slot["minute"],
                id=job_id,
                name=f"자동 작업 [{slot['label']}] {slot['hour']:02d}:{slot['minute']:02d}",
                replace_existing=True,
            )
            type_label = {"scrape": "수집", "check": "체크", "combo": "콤보"}.get(task_type, task_type)
            logger.info(f"📅 작업 스케줄 등록: {slot['label']} ({type_label}) {slot['hour']:02d}:{slot['minute']:02d} 브랜드={slot.get('brand_name','ALL')}")


# 스케줄러 초기화 함수 (중복 시작 방지)
_scheduler_started = False


def _refresh_daily_rate_job():
    """자정 환율 갱신"""
    try:
        from exchange import refresh_daily_rate
        refresh_daily_rate()
    except Exception as e:
        logger.warning(f"환율 갱신 실패: {e}")


def _retry_failed_tasks_job():
    """매일 23시 — 오류 상태 작업을 큐에 자동 추가"""
    try:
        import sqlite3
        db_path = os.path.join(get_path("db"), "users.db")
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        failed = conn.execute("SELECT * FROM scrape_tasks WHERE status='오류' ORDER BY id").fetchall()
        conn.close()

        if not failed:
            logger.info("🔄 23시 자동 재시도: 오류 작업 없음")
            return

        _start_queue_worker()

        count = 0
        for r in failed:
            conn = sqlite3.connect(db_path)
            conn.execute("UPDATE scrape_tasks SET status='예약', count=0 WHERE id=?", (r["id"],))
            conn.commit()
            conn.close()
            _scrape_queue.put(r["id"])
            count += 1

        msg = f"🔄 23시 자동 재시도: {count}개 오류 작업 큐에 예약"
        logger.info(msg)
        push_log(msg)
        try:
            from notifier import send_telegram
            task_names = "\n".join(f"  {r['brand_name'] or '전체'} / {r['cat_name'] or '전체'} (p.{r['pages'] or '전체'})" for r in failed)
            send_telegram(f"🔄 <b>오류 작업 자동 재시도</b>\n{count}개 작업 큐에 예약됨\n\n{task_names}")
        except Exception:
            pass
    except Exception as e:
        logger.warning(f"오류 재시도 실패: {e}")


def _check_ai_api_job():
    """AI API 상태 확인 → 문제 시 텔레그램 알림"""
    try:
        from notifier import check_ai_api_and_notify
        check_ai_api_and_notify()
    except Exception as e:
        logger.warning(f"AI API 모니터링 오류: {e}")


def _start_scheduler_once():
    """스케줄러를 한 번만 시작 (중복 방지)
    Mac(서버): 카페/블로그/기사 업로드 + NAS 동기화 + 환율 (수집 최소화)
    Windows(수집PC): 수집 스케줄 + 오류 재시도
    """
    global _scheduler_started
    if _scheduler_started:
        return

    import platform
    is_mac = platform.system() == "Darwin"
    is_windows = platform.system() == "Windows"
    env_label = "Mac 서버" if is_mac else "Windows 수집PC"
    logger.info(f"📅 스케줄러 초기화 ({env_label})")

    if is_windows:
        # ── 윈도우 전용: 수집 + NAS 내보내기 ──
        _register_task_schedule_jobs()     # 자동 수집/체크/콤보
        _register_check_schedule_job()     # 업로드 체크
        # 오류 작업 자동 재시도 (매일 23:00)
        scheduler.add_job(
            func=_retry_failed_tasks_job,
            trigger="cron", hour=23, minute=0,
            id="retry_failed_tasks", replace_existing=True,
            name="오류 작업 자동 재시도 (23:00)",
        )
        # 매시 NAS로 products.db 내보내기
        try:
            scheduler.add_job(
                func=export_all_to_nas,
                trigger="cron", minute=0,
                id="nas_export", replace_existing=True,
                name="NAS 전체 내보내기 (매시 정각)",
            )
            logger.info("📤 [Windows] NAS 내보내기 등록 (매시 정각)")
        except NameError:
            pass
        # 상품DB 자동 업데이트 (7일 주기)
        _register_db_update_job()
        # 유휴 시간 상품 최신화 체크 (5분 간격)
        scheduler.add_job(
            func=_idle_product_check,
            trigger="interval", minutes=5,
            id="idle_product_check", replace_existing=True,
            name="유휴 시간 상품 최신화 (5분 간격)",
        )
        logger.info("🔍 [Windows] 유휴 시간 상품 최신화 등록 (5분 간격)")
        logger.info("🔄 [Windows] 수집 스케줄 등록 완료")

    if is_mac:
        # ── 맥 서버 전용: 카페 업로드(스포츠+빈티지)/기사/동기화 ──
        _register_schedule_jobs()          # 스포츠 카페 업로드
        _register_vt_schedule_jobs()       # 빈티지 카페 업로드
        try:
            _register_fb_schedule_jobs()   # 자유게시판 기사 자동 생성/업로드
        except NameError:
            pass
        # NAS 상품 동기화 (매시 30분)
        try:
            scheduler.add_job(
                func=sync_products_from_nas,
                trigger="cron", minute=30,
                id="nas_sync", replace_existing=True,
                name="NAS 상품 동기화 (매시 30분)",
            )
            logger.info("📂 [Mac] NAS 동기화 스케줄 등록 (매시 30분)")
        except NameError:
            pass
        # AI API 상태 모니터링 (5분 간격)
        scheduler.add_job(
            func=_check_ai_api_job,
            trigger="interval", minutes=5,
            id="ai_api_monitor", replace_existing=True,
            name="AI API 상태 모니터링 (5분)",
        )
        logger.info("📡 [Mac] AI API 모니터링 등록")

    # ── 공통: 환율 갱신 ──
    scheduler.add_job(
        func=_refresh_daily_rate_job,
        trigger="cron", hour=0, minute=1,
        id="daily_rate_refresh", replace_existing=True,
        name="일일 환율 갱신 (00:01)",
    )
    logger.info("💱 환율 갱신 등록 (매일 00:01)")

    # ── 공통: 회원 만료 알림 (매일 09:00) ──
    scheduler.add_job(
        func=_check_member_expiry,
        trigger="cron", hour=9, minute=0,
        id="member_expiry_check", replace_existing=True,
        name="회원 만료 알림 (09:00)",
    )
    logger.info("📅 회원 만료 알림 등록 (매일 09:00)")

    # 장바구니 품절 체크: 담을 때 개별 체크로 변경 (스케줄 제거)

    # ── 공통: Git 자동 풀 (매시 정각) ──
    scheduler.add_job(
        func=_auto_git_pull,
        trigger="cron", minute=0,
        id="auto_git_pull", replace_existing=True,
        name="Git 자동 풀 (매시 정각)",
    )
    logger.info("🔄 Git 자동 풀 등록 (매시 정각)")

    # ── 자동 백업 (OS별) ──
    if is_mac:
        # Mac → NAS: users.db(하루 2회) + config(하루 1회)
        scheduler.add_job(
            func=_backup_users_db,
            trigger="cron", hour="0,12", minute=0,
            id="backup_users", replace_existing=True,
            name="users.db 백업 (00:00, 12:00)",
        )
        scheduler.add_job(
            func=_backup_config_files,
            trigger="cron", hour=1, minute=0,
            id="backup_config", replace_existing=True,
            name="설정파일 백업 (01:00)",
        )
        logger.info("💾 [Mac] 백업: users.db(00:00,12:00) + config(01:00) → NAS")

    if is_windows:
        # Windows → NAS: products.db(하루 1회, 새벽 2시)
        # (매시 정각 NAS 내보내기 + 300개 단위 내보내기는 기존 유지)
        scheduler.add_job(
            func=_backup_products_db,
            trigger="cron", hour=2, minute=0,
            id="backup_products", replace_existing=True,
            name="products.db 백업 (02:00)",
        )
        logger.info("💾 [Windows] 백업: products.db(02:00) → NAS")

    scheduler.start()
    _scheduler_started = True
    logger.info(f"📅 스케줄러 시작 완료 ({env_label}, PID: {os.getpid()})")


def _backup_db_file(db_name, subfolder, max_backups=60):
    """DB 파일 백업 공통 함수 (로컬 + NAS)"""
    import shutil
    import sqlite3 as _sq
    db_dir = get_path("db")
    src = os.path.join(db_dir, db_name)
    if not os.path.exists(src):
        logger.warning(f"[백업] {db_name} 없음 — 스킵")
        return

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    prefix = db_name.replace(".db", "")

    # 로컬 백업
    local_bk = os.path.join(db_dir, "backups", subfolder)
    os.makedirs(local_bk, exist_ok=True)
    dst = os.path.join(local_bk, f"{prefix}_{ts}.db")
    try:
        conn = _sq.connect(src)
        bk = _sq.connect(dst)
        conn.backup(bk)
        bk.close()
        conn.close()
        logger.info(f"[백업] 로컬 {subfolder}/{prefix}_{ts}.db 완료")
    except Exception:
        shutil.copy2(src, dst)
        logger.info(f"[백업] 로컬 {subfolder}/{prefix}_{ts}.db 완료 (복사)")

    _cleanup_old_backups(local_bk, prefix, max_backups)

    # NAS 백업
    from data_manager import NAS_SHARED_PATH
    nas_bk = os.path.join(NAS_SHARED_PATH, "backups", subfolder)
    if os.path.isdir(NAS_SHARED_PATH):
        try:
            os.makedirs(nas_bk, exist_ok=True)
            shutil.copy2(dst, os.path.join(nas_bk, f"{prefix}_{ts}.db"))
            logger.info(f"[백업] NAS {subfolder}/{prefix}_{ts}.db 완료")
            _cleanup_old_backups(nas_bk, prefix, max_backups)
        except Exception as e:
            logger.warning(f"[백업] NAS {subfolder} 실패: {e}")
    else:
        logger.warning("[백업] NAS 미연결 — NAS 백업 스킵")


def _cleanup_old_backups(bk_dir, prefix, max_count):
    """오래된 백업 파일 삭제"""
    files = sorted(
        [f for f in os.listdir(bk_dir) if f.startswith(prefix + "_") and f.endswith(".db")],
        reverse=True
    )
    for f in files[max_count:]:
        try:
            os.remove(os.path.join(bk_dir, f))
        except Exception:
            pass


def _backup_users_db():
    """[Mac] users.db 백업 → NAS/backups/users/"""
    _backup_db_file("users.db", "users", max_backups=60)


def _backup_config_files():
    """[Mac] 설정파일 백업 → NAS/backups/config/"""
    import shutil
    db_dir = get_path("db")
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    max_backups = 30  # 1달분

    config_files = [
        "cafe_schedule.json", "vt_cafe_schedule.json", "check_schedule.json",
        "fb_schedule.json", "db_update_schedule.json", "task_schedule.json",
        "price_config.json", "vintage_price.json", "biz_info.json",
        "naver_accounts.json", "blog_accounts.json",
        "translation_dict.json", "uploaded_history.json",
        "scrape_history.json", "ai_settings.db",
    ]

    # 로컬 백업
    local_bk = os.path.join(db_dir, "backups", "config", ts)
    os.makedirs(local_bk, exist_ok=True)
    count = 0
    for fn in config_files:
        src = os.path.join(db_dir, fn)
        if os.path.exists(src):
            shutil.copy2(src, os.path.join(local_bk, fn))
            count += 1
    logger.info(f"[백업] 로컬 config/{ts}/ — {count}개 파일")

    # 오래된 폴더 삭제
    config_root = os.path.join(db_dir, "backups", "config")
    dirs = sorted([d for d in os.listdir(config_root) if os.path.isdir(os.path.join(config_root, d))], reverse=True)
    for d in dirs[max_backups:]:
        try:
            shutil.rmtree(os.path.join(config_root, d))
        except Exception:
            pass

    # NAS 백업
    from data_manager import NAS_SHARED_PATH
    if os.path.isdir(NAS_SHARED_PATH):
        nas_bk = os.path.join(NAS_SHARED_PATH, "backups", "config", ts)
        try:
            os.makedirs(nas_bk, exist_ok=True)
            for fn in os.listdir(local_bk):
                shutil.copy2(os.path.join(local_bk, fn), os.path.join(nas_bk, fn))
            logger.info(f"[백업] NAS config/{ts}/ — {count}개 파일")
            # NAS 오래된 폴더 삭제
            nas_root = os.path.join(NAS_SHARED_PATH, "backups", "config")
            nas_dirs = sorted([d for d in os.listdir(nas_root) if os.path.isdir(os.path.join(nas_root, d))], reverse=True)
            for d in nas_dirs[max_backups:]:
                try:
                    shutil.rmtree(os.path.join(nas_root, d))
                except Exception:
                    pass
        except Exception as e:
            logger.warning(f"[백업] NAS config 실패: {e}")


def _backup_products_db():
    """[Windows] products.db 백업 → NAS/backups/products/"""
    _backup_db_file("products.db", "products", max_backups=30)


def _check_cart_soldout():
    """장바구니 상품 품절 체크 (매일 새벽 4시)"""
    from user_db import _conn
    conn = _conn()
    try:
        rows = conn.execute("SELECT DISTINCT product_code FROM cart WHERE product_code IS NOT NULL AND product_code != ''").fetchall()
        codes = [r[0] for r in rows]
        if not codes:
            return
        logger.info(f"[장바구니 품절체크] {len(codes)}개 상품 체크 시작")
        sold_count = 0
        for code in codes:
            try:
                is_sold = _check_product_soldout(code)
                if is_sold:
                    sold_count += 1
                import time
                time.sleep(1)  # 봇 감지 방지
            except Exception:
                pass
        logger.info(f"[장바구니 품절체크] 완료: {len(codes)}개 중 {sold_count}개 품절")
    except Exception as e:
        logger.warning(f"[장바구니 품절체크] 오류: {e}")
    finally:
        conn.close()


def _check_member_expiry():
    """회원 사용기간 만료 3일전/1일전 문자 알림"""
    from datetime import timedelta
    from user_db import _conn
    conn = _conn()
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        d3 = (datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d")
        d1 = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")

        # 3일 후 만료 회원
        rows_3d = conn.execute("""
            SELECT username, name, phone, expires_at FROM users
            WHERE status='approved' AND expires_at=? AND phone IS NOT NULL AND phone != ''
        """, (d3,)).fetchall()

        # 1일 후 만료 회원
        rows_1d = conn.execute("""
            SELECT username, name, phone, expires_at FROM users
            WHERE status='approved' AND expires_at=? AND phone IS NOT NULL AND phone != ''
        """, (d1,)).fetchall()

        if not rows_3d and not rows_1d:
            return

        from aligo_sms import send_sms, load_config
        load_config()

        for r in rows_3d:
            msg = (
                f"[TheOne Vintage] {r['name'] or r['username']}님,\n"
                f"사용기간이 3일 후({r['expires_at']}) 만료됩니다.\n"
                f"연장을 원하시면 관리자에게 문의해주세요.\n"
                f"감사합니다."
            )
            send_sms(r["phone"], msg, title="TheOne Vintage")
            logger.info(f"[SMS] 만료 3일전 알림: {r['username']} ({r['phone']})")

        for r in rows_1d:
            msg = (
                f"[TheOne Vintage] {r['name'] or r['username']}님,\n"
                f"사용기간이 내일({r['expires_at']}) 만료됩니다.\n"
                f"연장을 원하시면 관리자에게 문의해주세요.\n"
                f"감사합니다."
            )
            send_sms(r["phone"], msg, title="TheOne Vintage")
            logger.info(f"[SMS] 만료 1일전 알림: {r['username']} ({r['phone']})")

    except Exception as e:
        logger.warning(f"[SMS] 만료 알림 체크 실패: {e}")
    finally:
        conn.close()


def _auto_git_pull():
    """GitHub에서 최신 코드 자동 풀"""
    import subprocess
    try:
        project_dir = os.path.dirname(os.path.abspath(__file__))
        result = subprocess.run(
            ["git", "pull", "origin", "main"],
            cwd=project_dir, capture_output=True, text=True, timeout=30
        )
        output = result.stdout.strip()
        if "Already up to date" in output:
            logger.debug("🔄 Git: 이미 최신")
        elif output:
            push_log(f"🔄 Git 자동 풀 완료: {output[:100]}")
            logger.info(f"🔄 Git pull: {output[:100]}")
    except Exception as e:
        logger.warning(f"🔄 Git pull 실패: {e}")


# use_reloader=True 시 부모(리로더) + 자식(워커) 2개 프로세스가 생성됨
# 자식(워커)에만 WERKZEUG_RUN_MAIN="true" 설정됨
# 부모에서도 스케줄러가 시작되면 같은 잡이 2번 실행 → 브라우저 2개 열림!
# → 워커 프로세스에서만 스케줄러 시작
# 스케줄러는 __main__ 블록에서만 시작 (중복 방지)

set_app_status(status)  # xebio_search에 status 딕셔너리 주입

# 카페 모니터 + 텔레그램 봇 자동 시작
try:
    _monitor_started = start_monitor(log_callback=push_log, interval=180)
    _bot_started = start_bot(log_callback=push_log)
    if _monitor_started:
        logger.info("📡 카페 모니터 자동 시작됨")
    if _bot_started:
        logger.info("🤖 텔레그램 봇 자동 시작됨")
except Exception as e:
    logger.warning(f"⚠️ 모니터/봇 자동 시작 실패: {e}")


# =============================================
# 라우트 (URL)
# =============================================

@app.route(f"{URL_PREFIX}/")
def root_redirect():
    """루트: 비로그인/일반회원 → 쇼핑몰 직접 렌더링, 관리자 → 대시보드"""
    if session.get("logged_in") and session.get("role", "admin") == "admin":
        return redirect(f"{URL_PREFIX}/admin/vintage")
    # 302 리다이렉트 대신 직접 렌더링 (네이버 검색 로봇 대응)
    return shop()


@app.route(f"{URL_PREFIX}/dashboard")
@admin_required
def dashboard_page():
    """메인 대시보드 페이지"""
    products = load_latest_products()
    rate = get_jpy_to_krw_rate()
    resp = make_response(render_template(
        "dashboard.html",
        status=status,
        rate=rate,
        product_count=len(products),
        schedule_time=f"{AUTO_SCHEDULE_HOUR:02d}:{AUTO_SCHEDULE_MINUTE:02d}",
        url_prefix=URL_PREFIX,
        env=APP_ENV,
        version=APP_VERSION,
    ))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.route(f"{URL_PREFIX}/products")
@admin_required
def get_products():
    """수집된 상품 목록 JSON 반환 (브랜드 필터, 페이지네이션)
    latest.json + 빅데이터 DB 미업로드 상품 병합
    """
    products = load_latest_products()
    # latest.json 상품은 모두 sports로 간주
    for p in products:
        if "source_type" not in p:
            p["source_type"] = "sports"

    # 빅데이터 DB에서 미업로드 상품 병합 (중복 제거)
    include_db = request.args.get("include_db", "true").lower()
    source_type_filter = request.args.get("source_type", "").strip()
    if include_db == "true":
        from product_db import get_unuploaded_products
        db_products = get_unuploaded_products(source_type=source_type_filter)
        # latest.json에 있는 품번 수집
        existing_codes = set()
        for p in products:
            code = p.get("product_code", "")
            if code:
                existing_codes.add(code)
        # DB 상품 중 latest.json에 없는 것만 추가
        for dp in db_products:
            if dp.get("product_code") and dp["product_code"] not in existing_codes:
                existing_codes.add(dp["product_code"])
                products.append(dp)

    # source_type 필터 (sports / vintage)
    if source_type_filter:
        products = [p for p in products if (p.get("source_type") or "sports") == source_type_filter]

    # 브랜드별 수량 집계 (source_type 필터 적용 후)
    brand_counts = {}
    for p in products:
        b = (p.get("brand_ko") or p.get("brand") or "").strip()
        if b:
            brand_counts[b] = brand_counts.get(b, 0) + 1
    total_all = len(products)

    # 브랜드 필터 (한국어/원문 모두 비교)
    brand_filter = request.args.get("brand", "").strip()
    search_filter = request.args.get("search", "").strip().lower()
    status_filter = request.args.get("status", "").strip()
    category_filter = request.args.get("category", "").strip()
    if brand_filter and brand_filter != "ALL":
        products = [p for p in products if
                    (p.get("brand_ko") or "").strip() == brand_filter or
                    (p.get("brand")    or "").strip() == brand_filter]
    if search_filter:
        products = [p for p in products if search_filter in p.get("name", "").lower()
                    or search_filter in p.get("brand", "").lower()
                    or search_filter in p.get("product_code", "").lower()]
    if category_filter and category_filter != "ALL":
        cat_keywords = {
            "bag": ["バッグ","ショルダー","トート","リュック","ハンド","ポーチ","ボストン","クラッチ","ウエスト"],
            "clothing": ["ジャケット","コート","シャツ","ブラウス","ワンピース","パンツ","スラックス","ニット","セーター","カーディガン","パーカー","スウェット","ベスト","Tシャツ","ドレス","スカート"],
            "shoes": ["シューズ","スニーカー","ブーツ","サンダル","パンプス","ローファー"],
            "watch": ["時計","ウォッチ"],
            "accessory": ["財布","ベルト","マフラー","帽子","サングラス","ネックレス","ブレスレット","リング","アクセサリー","ストール","スカーフ"],
        }
        kws = cat_keywords.get(category_filter, [])
        if kws:
            products = [p for p in products if any(k in (p.get("name") or "") for k in kws)]
    if status_filter and status_filter != "ALL":
        products = [p for p in products if (p.get("cafe_status") or "대기") == status_filter]
        # 완료/중복 필터 시 DB에서도 해당 상태 상품 병합
        if status_filter in ("완료", "중복"):
            try:
                from product_db import get_products_by_status
                db_status_products = get_products_by_status(status_filter)
                existing_codes = {p.get("product_code") for p in products if p.get("product_code")}
                for dp in db_status_products:
                    if dp.get("product_code") and dp["product_code"] not in existing_codes:
                        existing_codes.add(dp["product_code"])
                        products.append(dp)
            except Exception as e:
                logger.warning(f"DB 상태 조회 실패: {e}")
        # DB도 없고 latest에도 없으면 업로드 히스토리에서 완료 상품 복원
        if status_filter == "완료" and len(products) == 0:
            try:
                hist_path = os.path.join(get_path("db"), "uploaded_history.json")
                if os.path.exists(hist_path):
                    import json as _json
                    with open(hist_path, "r", encoding="utf-8") as _f:
                        hist = _json.load(_f)
                    for h in hist:
                        if h.get("product_code"):
                            products.append({
                                "product_code": h["product_code"],
                                "name": h.get("name", ""),
                                "name_ko": h.get("name", ""),
                                "brand": h.get("brand", ""),
                                "brand_ko": h.get("brand", ""),
                                "price_jpy": h.get("price_jpy", 0),
                                "cafe_status": "완료",
                                "cafe_uploaded_at": h.get("uploaded_at", ""),
                            })
            except Exception as e:
                logger.warning(f"업로드 히스토리 조회 실패: {e}")

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)
    start = (page - 1) * per_page
    end = start + per_page
    page_products = products[start:end]

    # 구매대행 가격 추가 - 환율 한 번만 조회 후 재사용
    rate = get_jpy_to_krw_rate()
    for p in page_products:
        if p.get("price_jpy"):
            p["price_info"] = calc_buying_price(p["price_jpy"], rate=rate)

    return jsonify({
        "total": len(products),
        "total_all": total_all,
        "page": page,
        "per_page": per_page,
        "products": page_products,
        "brand_counts": brand_counts,
    })


@app.route(f"{URL_PREFIX}/products/download")
@admin_required
def download_csv_products():
    """수집된 상품 CSV 다운로드"""
    import io
    import csv as csv_mod
    from flask import send_file

    products = load_latest_products()
    rate = get_cached_rate()

    # CSV 헤더
    fields = ["product_code", "brand", "brand_ko", "name_ko", "name", "price_jpy", "cost_krw"]
    buf = io.StringIO()
    writer = csv_mod.writer(buf)
    writer.writerow(fields)

    for p in products:
        cost_krw = 0
        if p.get("price_jpy"):
            from exchange import calc_buying_price
            info = calc_buying_price(p["price_jpy"], rate=rate)
            cost_krw = info["cost_krw"]

        writer.writerow([
            p.get("product_code", ""),
            p.get("brand", ""),
            p.get("brand_ko") or p.get("brand", ""),
            p.get("name_ko") or p.get("name", ""),
            p.get("name", ""),
            p.get("price_jpy", 0),
            cost_krw,
        ])

    output = io.BytesIO()
    output.write(b'\xef\xbb\xbf')  # UTF-8 BOM for Excel compatibility
    output.write(buf.getvalue().encode("utf-8"))
    output.seek(0)

    filename = f"products_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return send_file(
        output,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=filename
    )


@app.route(f"{URL_PREFIX}/products/brands")
@admin_required
def get_brands():
    """수집된 상품의 브랜드 목록 반환 (한국어 번역 우선, DB 미업로드 상품 포함)"""
    products = load_latest_products()

    # DB 미업로드 상품 병합 (스포츠만)
    from product_db import get_unuploaded_products
    db_products = get_unuploaded_products(source_type="sports")
    existing_codes = {p.get("product_code", "") for p in products if p.get("product_code")}
    for dp in db_products:
        if dp.get("product_code") and dp["product_code"] not in existing_codes:
            existing_codes.add(dp["product_code"])
            products.append(dp)

    brand_map = {}  # 원문 → 한국어 매핑
    brand_counts = {}

    for p in products:
        b_raw = (p.get("brand") or "").strip()
        b_ko  = (p.get("brand_ko") or b_raw).strip()
        if not b_raw:
            continue
        brand_map[b_raw] = b_ko
        key = b_ko or b_raw
        brand_counts[key] = brand_counts.get(key, 0) + 1

    # 한국어 기준으로 정렬
    brands_ko = sorted(set(brand_map.values()))
    return jsonify({
        "brands"   : brands_ko,          # 콤보박스에 표시할 한국어 목록
        "brand_map": brand_map,          # 원문→한국어 (필터 시 역매핑용)
        "counts"   : brand_counts,
    })


@app.route(f"{URL_PREFIX}/products/update", methods=["POST"])
@admin_required
def update_products():
    """상품 선택 상태 업데이트 (체크박스)"""
    data = request.json or {}
    # product_code 기반 선택 (우선) 또는 인덱스 기반 (하위호환)
    selected_codes = set(data.get("selected_codes", []))
    selected_ids = set(data.get("selected", []))

    products = load_latest_products()
    if selected_codes:
        for p in products:
            p["selected"] = p.get("product_code", "") in selected_codes
    else:
        for i, p in enumerate(products):
            p["selected"] = i in selected_ids

    from xebio_search import save_products
    save_products(products)
    count = sum(1 for p in products if p.get("selected"))
    return jsonify({"ok": True, "selected_count": count})


@app.route(f"{URL_PREFIX}/products/delete", methods=["POST"])
@admin_required
def delete_products():
    """상품 삭제 (인덱스 기준, 복수 가능)"""
    data = request.json or {}
    indices = data.get("indices", [])
    if not indices:
        return jsonify({"ok": False, "message": "삭제할 인덱스가 없습니다"})

    products = load_latest_products()
    # 내림차순 정렬 후 삭제 (인덱스 밀림 방지)
    valid = sorted(set(int(i) for i in indices if 0 <= int(i) < len(products)), reverse=True)
    for i in valid:
        products.pop(i)

    from xebio_search import save_products
    save_products(products)
    msg = f"상품 {len(valid)}개 삭제 완료 (남은 상품: {len(products)}개)"
    push_log(f"🗑️ " + msg)
    return jsonify({"ok": True, "deleted": len(valid), "remaining": len(products), "message": msg})


@app.route(f"{URL_PREFIX}/products/check-duplicate", methods=["POST"])
@admin_required
def check_duplicate():
    """업로드 전 중복 체크 (품번 + 가격 기준)"""
    data = request.json or {}
    selected_indices = data.get("indices", [])
    products = load_latest_products()

    uploaded_path = os.path.join(DB_DIR, "uploaded_history.json")
    if os.path.exists(uploaded_path):
        with open(uploaded_path, "r", encoding="utf-8") as f:
            history = json.load(f)
    else:
        history = []

    results = {"block": [], "price_changed": [], "ok": []}
    for idx in selected_indices:
        if idx >= len(products):
            continue
        p = products[idx]
        code = p.get("product_code", "")
        price = p.get("price_jpy", 0)

        matched = [h for h in history if h.get("product_code") == code and code]
        if not matched:
            results["ok"].append(idx)
        else:
            old_price = matched[-1].get("price_jpy", 0)
            if old_price == price:
                results["block"].append({"idx": idx, "name": p.get("name"), "reason": "동일 가격 중복"})
            else:
                results["price_changed"].append({
                    "idx": idx, "name": p.get("name"),
                    "old_price": old_price, "new_price": price,
                    "diff": price - old_price
                })

    return jsonify(results)


@app.route(f"{URL_PREFIX}/products/status", methods=["POST"])
@admin_required
def update_product_status():
    """개별 상품의 cafe_status 변경 (대기/완료/중복)"""
    data = request.json or {}
    product_code = data.get("product_code", "").strip()
    new_status = data.get("status", "").strip()

    if not product_code or new_status not in ("대기", "완료", "중복"):
        return jsonify({"ok": False, "message": "잘못된 요청입니다"})

    products = load_latest_products()
    found = False
    for p in products:
        if p.get("product_code") == product_code:
            p["cafe_status"] = new_status
            if new_status == "완료":
                p["cafe_uploaded"] = True
                p["cafe_uploaded_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            elif new_status == "대기":
                p["cafe_uploaded"] = False
                p.pop("cafe_uploaded_at", None)
            found = True
            break

    if not found:
        return jsonify({"ok": False, "message": "상품을 찾을 수 없습니다"})

    from xebio_search import save_products
    save_products(products)
    # DB에도 반영
    try:
        from product_db import update_cafe_status
        uploaded_at = datetime.now().strftime("%Y-%m-%d %H:%M") if new_status == "완료" else ""
        update_cafe_status(product_code, new_status, uploaded_at)
    except Exception as e:
        logger.warning(f"DB 상태 업데이트 실패: {e}")
    return jsonify({"ok": True, "product_code": product_code, "status": new_status})


@app.route(f"{URL_PREFIX}/products/bulk-status", methods=["POST"])
@admin_required
def bulk_update_product_status():
    """체크된 상품의 cafe_status 일괄 변경"""
    data = request.json or {}
    codes = data.get("codes", [])
    new_status = data.get("status", "").strip()

    if not codes or new_status not in ("대기", "완료", "중복"):
        return jsonify({"ok": False, "message": "잘못된 요청입니다"})

    products = load_latest_products()
    code_set = set(codes)
    count = 0
    for p in products:
        if p.get("product_code") in code_set:
            p["cafe_status"] = new_status
            if new_status == "완료":
                p["cafe_uploaded"] = True
                p["cafe_uploaded_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            elif new_status == "대기":
                p["cafe_uploaded"] = False
                p.pop("cafe_uploaded_at", None)
            count += 1

    from xebio_search import save_products
    save_products(products)
    # DB에도 반영
    try:
        from product_db import update_cafe_status
        uploaded_at = datetime.now().strftime("%Y-%m-%d %H:%M") if new_status == "완료" else ""
        for code in codes:
            update_cafe_status(code, new_status, uploaded_at)
    except Exception as e:
        logger.warning(f"DB 일괄 상태 업데이트 실패: {e}")
    return jsonify({"ok": True, "count": count, "status": new_status})


@app.route(f"{URL_PREFIX}/status")
@admin_required
def get_status():
    """현재 실행 상태 반환"""
    products = load_latest_products()
    # 현재 AI 모델
    try:
        from translator import get_current_ai_model
        ai_model = get_current_ai_model()
    except Exception:
        ai_model = ""
    return jsonify({
        **status,
        "product_count": len(products),
        "rate": get_cached_rate(),
        "margin": get_margin_rate(),
        "schedule_time": f"{AUTO_SCHEDULE_HOUR:02d}:{AUTO_SCHEDULE_MINUTE:02d}",
        "ai_model": ai_model,
    })


# ── 수동 실행 API ──────────────────────────

@app.route(f"{URL_PREFIX}/run/scrape", methods=["POST"])
@admin_required
def manual_scrape():
    """수동 스크래핑 실행"""
    data = request.json or {}
    site_id = data.get("site_id", "xebio")
    category_id = data.get("category_id", "sale")
    keyword = data.get("keyword", "")
    pages = data.get("pages", "")
    brand_code = data.get("brand_code", "")

    # 이미 진행 중이면 즉시 알림
    if status["scraping"]:
        return jsonify({"ok": False, "message": "⚠️ 이미 스크래핑이 진행 중입니다. 리셋 후 다시 시도해주세요."})

    push_log(f"🚀 수동 스크래핑 요청: site={site_id}, cat={category_id}, brand={brand_code}")
    thread = threading.Thread(
        target=run_scrape,
        args=(site_id, category_id, keyword, pages, brand_code),
        daemon=True,
    )
    thread.start()
    desc = f"{site_id} › {category_id}"
    if brand_code:
        desc += f" [{brand_code}]"
    if keyword:
        desc += f" [{keyword}]"
    if pages:
        desc += f" (p.{pages})"
    return jsonify({"ok": True, "message": f"스크래핑 시작됨 ({desc})"})


# ── 수집 작업리스트 DB ──────────────────────────
def _init_scrape_tasks_db():
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS scrape_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            site TEXT DEFAULT '2ndstreet',
            site_name TEXT DEFAULT '',
            cat TEXT DEFAULT '',
            cat_name TEXT DEFAULT '',
            brand TEXT DEFAULT '',
            brand_name TEXT DEFAULT '',
            pages TEXT DEFAULT '',
            total_items INTEGER DEFAULT 0,
            total_pages INTEGER DEFAULT 0,
            count INTEGER DEFAULT 0,
            status TEXT DEFAULT '대기',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )""")
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()


# ── 키워드 분석 API ─────────────────────
@app.route(f"{URL_PREFIX}/api/keyword/config", methods=["POST"])
@admin_required
def keyword_config_save():
    """네이버 검색광고 API 키 저장"""
    data = request.get_json() or {}
    from naver_keyword import save_api_keys
    save_api_keys(data.get("api_key", ""), data.get("secret_key", ""), data.get("customer_id", ""))
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/api/keyword/search", methods=["POST"])
@admin_required
def keyword_search():
    """키워드 검색량 조회"""
    data = request.get_json() or {}
    keywords = data.get("keywords", [])
    if not keywords:
        return jsonify({"ok": False, "message": "키워드를 입력해주세요"})
    try:
        from naver_keyword import get_keyword_stats, load_api_keys
        load_api_keys()
        results = get_keyword_stats(keywords)
        # 검색량 합계 기준 정렬
        results.sort(key=lambda x: x.get("total", 0), reverse=True)
        return jsonify({"ok": True, "results": results})
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)})
    except Exception as e:
        return jsonify({"ok": False, "message": f"조회 실패: {e}"})


# ── 상품 최신화 체크 (Windows 전용) ─────────────────────
_freshness_status = {
    "running": False,
    "mode": "",        # "idle" 또는 "interleaved"
    "log": [],
    "last_result": {},
}

def _freshness_log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    _freshness_status["log"].append(f"[{ts}] {msg}")
    if len(_freshness_status["log"]) > 500:
        _freshness_status["log"] = _freshness_status["log"][-300:]
    logger.info(f"[최신화] {msg}")


def _idle_product_check():
    """5분 간격 호출 — 수집 작업이 없으면 자동으로 상태 체크 실행 (Windows 전용)"""
    import platform
    if platform.system() != "Windows":
        return

    # 이미 실행 중이면 스킵
    if _freshness_status["running"]:
        return
    # 수집 작업 진행 중이면 스킵 (교대 실행은 수집 함수 내에서 처리)
    if status.get("scraping"):
        return
    if _db_update_status.get("running"):
        return

    # 유휴 시간 → 최신화 체크 실행
    t = threading.Thread(target=_run_freshness_check, args=("idle",), daemon=True)
    t.start()


def _run_freshness_check(mode="idle"):
    """상품 최신화 체크 실행 (백그라운드)

    mode="idle": 유휴 시간 — 계속 300개씩 체크 (다른 작업 시작되면 중단)
    mode="interleaved": 교대 실행 — 1배치(300개)만 체크 후 반환
    """
    import time
    from product_checker import run_check_batch, get_check_stats, checker_status, CHUNK_SIZE

    _freshness_status["running"] = True
    _freshness_status["mode"] = mode
    checker_status["stop_requested"] = False

    try:
        if mode == "idle":
            _freshness_log("유휴 시간 최신화 시작")
            batch_count = 0
            while True:
                # 다른 작업이 시작되면 즉시 중단
                if status.get("scraping") or _db_update_status.get("running"):
                    _freshness_log("수집 작업 감지 — 최신화 일시 중단")
                    break
                if checker_status["stop_requested"]:
                    _freshness_log("중지 요청 — 최신화 중단")
                    break

                result = run_check_batch(CHUNK_SIZE, _freshness_log)
                if result["checked"] == 0 and result["sold_out"] == 0:
                    _freshness_log("체크할 상품 없음 — 전체 최신화 완료")
                    break

                batch_count += 1
                total = result["checked"] + result["sold_out"]
                _freshness_log(f"배치 #{batch_count} 완료: 체크 {result['checked']}, 품절 {result['sold_out']}, 가격변동 {result['price_changed']}")

                # 배치 간 대기 (30초)
                _freshness_log("30초 대기...")
                for _ in range(30):
                    if status.get("scraping") or _db_update_status.get("running"):
                        break
                    if checker_status["stop_requested"]:
                        break
                    time.sleep(1)

        elif mode == "interleaved":
            _freshness_log("교대 실행: 체크 300개 시작")
            result = run_check_batch(CHUNK_SIZE, _freshness_log)
            _freshness_log(f"교대 체크 완료: 체크 {result['checked']}, 품절 {result['sold_out']}, 가격변동 {result['price_changed']}")
            _freshness_status["last_result"] = result

    except Exception as e:
        _freshness_log(f"최신화 오류: {e}")
    finally:
        _freshness_status["running"] = False
        _freshness_status["mode"] = ""


def run_interleaved_check():
    """수집 작업에서 호출 — 1배치 교대 체크 실행 (동기)"""
    import time
    from product_checker import run_check_batch, CHUNK_SIZE

    _freshness_log("교대 실행: 체크 300개 시작")
    result = run_check_batch(CHUNK_SIZE, _freshness_log)
    _freshness_log(f"교대 체크 완료: 체크 {result['checked']}, 품절 {result['sold_out']}, 가격변동 {result['price_changed']}")
    return result


@app.route(f"{URL_PREFIX}/api/freshness/status", methods=["GET"])
@admin_required
def freshness_status_api():
    """최신화 체크 상태 조회"""
    from product_checker import get_check_stats
    try:
        stats = get_check_stats()
    except Exception:
        stats = {}
    return jsonify({
        "ok": True,
        "running": _freshness_status["running"],
        "mode": _freshness_status["mode"],
        "stats": stats,
        "log": "\n".join(_freshness_status["log"][-50:]),
    })


@app.route(f"{URL_PREFIX}/api/freshness/run", methods=["POST"])
@admin_required
def freshness_run_api():
    """최신화 체크 수동 실행"""
    if _freshness_status["running"]:
        return jsonify({"ok": False, "message": "이미 실행 중입니다"})
    if status.get("scraping"):
        return jsonify({"ok": False, "message": "수집 작업 진행 중 — 수집 완료 후 자동 실행됩니다"})
    t = threading.Thread(target=_run_freshness_check, args=("idle",), daemon=True)
    t.start()
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/api/freshness/stop", methods=["POST"])
@admin_required
def freshness_stop_api():
    """최신화 체크 중지"""
    from product_checker import checker_status
    checker_status["stop_requested"] = True
    return jsonify({"ok": True})


# ── To Do List ─────────────────────
@app.route(f"{URL_PREFIX}/admin/vintage")
@admin_required
def admin_vintage():
    """빈티지 관리 페이지 (경량)"""
    return render_template("admin_vintage.html", url_prefix=URL_PREFIX, env=APP_ENV, active_menu="vintage", default_page="analytics")


@app.route(f"{URL_PREFIX}/admin/brand")
@admin_required
def admin_brand():
    """브랜드 관리 페이지 (경량)"""
    return render_template("admin_brand.html", url_prefix=URL_PREFIX, env=APP_ENV, active_menu="brand", default_page="brand-dashboard")


@app.route(f"{URL_PREFIX}/admin/kabinet")
@admin_required
def admin_kabinet():
    """캐비넷 관리 페이지 (경량)"""
    return render_template("admin_kabinet.html", url_prefix=URL_PREFIX, env=APP_ENV, active_menu="kabinet", default_page="kv-dashboard")


@app.route(f"{URL_PREFIX}/admin/setting")
@admin_required
def admin_setting():
    """설정 페이지 (경량)"""
    return render_template("admin_setting.html", url_prefix=URL_PREFIX, env=APP_ENV, active_menu="setting", default_page="st-account")


@app.route(f"{URL_PREFIX}/orders-page")
@admin_required
def orders_light_page():
    """주문확인 경량 페이지"""
    return render_template("orders_light.html", url_prefix=URL_PREFIX)


@app.route(f"{URL_PREFIX}/members-page")
@admin_required
def members_light_page():
    """회원관리 경량 페이지"""
    return render_template("members_light.html", url_prefix=URL_PREFIX)


@app.route(f"{URL_PREFIX}/analytics")
@admin_required
def analytics_page():
    """접속 통계 전용 페이지 (경량)"""
    return render_template("analytics.html", url_prefix=URL_PREFIX)


@app.route(f"{URL_PREFIX}/todo")
@admin_required
def todo_page():
    """To Do List 전용 페이지"""
    return render_template("todo.html", url_prefix=URL_PREFIX)


@app.route(f"{URL_PREFIX}/api/todo", methods=["GET"])
@admin_required
def get_todos():
    path = os.path.join(get_path("db"), "todos.json")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return jsonify({"ok": True, "todos": json.load(f)})
    return jsonify({"ok": True, "todos": []})


@app.route(f"{URL_PREFIX}/api/todo", methods=["POST"])
@admin_required
def save_todos():
    data = request.get_json() or {}
    todos = data.get("todos", [])
    # 이미지 제거하여 용량 절약 (이미지는 별도 저장)
    path = os.path.join(get_path("db"), "todos.json")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(todos, f, ensure_ascii=False, indent=2)
    return jsonify({"ok": True})


# ── IP 지역 조회 ─────────────────────
_ip_region_cache = {}  # {ip: {"region": "서울", "ts": timestamp}}

def _get_ip_region(ip: str) -> str:
    """IP 주소에서 지역 정보 조회 (캐시 24시간)"""
    import time as _time
    if not ip or ip in ("127.0.0.1", "::1", "localhost"):
        return "로컬"

    # 사설 IP 체크
    if ip.startswith(("10.", "192.168.", "172.")):
        return "내부망"

    # 캐시 확인 (24시간)
    cached = _ip_region_cache.get(ip)
    if cached and (_time.time() - cached["ts"]) < 86400:
        return cached["region"]

    region = ""
    try:
        r = requests.get(f"http://ip-api.com/json/{ip}?fields=status,country,regionName,city&lang=ko", timeout=3)
        if r.status_code == 200:
            d = r.json()
            if d.get("status") == "success":
                city = d.get("city", "")
                region_name = d.get("regionName", "")
                country = d.get("country", "")
                if city:
                    region = city
                elif region_name:
                    region = region_name
                elif country:
                    region = country
    except Exception:
        pass

    if not region:
        region = "알수없음"

    _ip_region_cache[ip] = {"region": region, "ts": _time.time()}
    # 캐시 크기 제한
    if len(_ip_region_cache) > 1000:
        oldest = sorted(_ip_region_cache.items(), key=lambda x: x[1]["ts"])[:500]
        for k, _ in oldest:
            del _ip_region_cache[k]

    return region


# ── 접속 통계 ─────────────────────
@app.route(f"{URL_PREFIX}/api/analytics/log", methods=["POST"])
def analytics_log():
    """접속 로그 수집"""
    data = request.get_json() or {}
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()
    ua = data.get("ua", "")
    referrer = data.get("referrer", "")
    path = data.get("path", "/shop")

    # 디바이스 판별
    ua_lower = ua.lower()
    if "mobile" in ua_lower or "android" in ua_lower or "iphone" in ua_lower:
        device = "모바일"
    elif "tablet" in ua_lower or "ipad" in ua_lower:
        device = "태블릿"
    else:
        device = "PC"

    # 검색 키워드 추출
    keyword = ""
    ref_source = ""
    if referrer:
        from urllib.parse import urlparse, parse_qs
        try:
            parsed = urlparse(referrer)
            host = parsed.hostname or ""
            if "google" in host:
                ref_source = "Google"
                keyword = parse_qs(parsed.query).get("q", [""])[0]
            elif "naver" in host:
                ref_source = "Naver"
                keyword = parse_qs(parsed.query).get("query", [""])[0]
            elif "daum" in host or "kakao" in host:
                ref_source = "Daum/Kakao"
                keyword = parse_qs(parsed.query).get("q", [""])[0]
            elif "bing" in host:
                ref_source = "Bing"
                keyword = parse_qs(parsed.query).get("q", [""])[0]
            else:
                ref_source = host
        except Exception:
            ref_source = referrer[:50]

    # 회원 정보
    member_username = session.get("username", "") if session.get("logged_in") else ""
    member_name = session.get("name", "") if session.get("logged_in") else ""
    member_type = "회원" if session.get("logged_in") else "비회원"

    # 로그 저장 (JSON 파일)
    log_path = os.path.join(get_path("db"), "analytics.json")
    logs = []
    if os.path.exists(log_path):
        try:
            with open(log_path, "r", encoding="utf-8") as f:
                logs = json.load(f)
        except Exception:
            logs = []

    # IP 지역 조회 (캐시 사용)
    region = _get_ip_region(ip)

    logs.append({
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "ip": ip,
        "region": region,
        "path": path,
        "referrer": ref_source or referrer[:80],
        "keyword": keyword,
        "device": device,
        "ua": ua[:100],
        "username": member_username,
        "name": member_name,
        "member_type": member_type,
    })

    # 최대 5000건 유지
    if len(logs) > 5000:
        logs = logs[-3000:]

    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(logs, f, ensure_ascii=False)

    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/api/analytics", methods=["GET"])
@admin_required
def analytics_stats():
    """접속 통계 조회"""
    period = request.args.get("period", "7d")
    log_path = os.path.join(get_path("db"), "analytics.json")
    logs = []
    if os.path.exists(log_path):
        try:
            with open(log_path, "r", encoding="utf-8") as f:
                logs = json.load(f)
        except Exception:
            pass

    from datetime import timedelta
    now = datetime.now()
    if period == "today":
        cutoff = now.strftime("%Y-%m-%d")
        filtered = [l for l in logs if l.get("time", "").startswith(cutoff)]
    elif period == "yesterday":
        yesterday = (now - timedelta(days=1)).strftime("%Y-%m-%d")
        filtered = [l for l in logs if l.get("time", "").startswith(yesterday)]
    elif period == "30d":
        cutoff = (now - timedelta(days=30)).strftime("%Y-%m-%d")
        filtered = [l for l in logs if l.get("time", "") >= cutoff]
    else:
        cutoff = (now - timedelta(days=7)).strftime("%Y-%m-%d")
        filtered = [l for l in logs if l.get("time", "") >= cutoff]

    today_str = now.strftime("%Y-%m-%d")
    today_count = sum(1 for l in logs if l.get("time", "").startswith(today_str))
    unique_ips = len(set(l.get("ip", "") for l in filtered))

    # 유입 경로
    from collections import Counter
    ref_counter = Counter(l.get("referrer", "") or "직접 접속" for l in filtered)
    referrers = [{"name": k, "count": v} for k, v in ref_counter.most_common(15)]

    # 검색 키워드
    kw_counter = Counter(l["keyword"] for l in filtered if l.get("keyword"))
    keywords = [{"keyword": k, "count": v} for k, v in kw_counter.most_common(20)]
    search_count = sum(1 for l in filtered if l.get("keyword"))

    # 디바이스
    dev_counter = Counter(l.get("device", "기타") for l in filtered)
    devices = [{"name": k, "count": v} for k, v in dev_counter.most_common(5)]

    # 일별
    day_counter = Counter(l.get("time", "")[:10] for l in filtered)
    daily = [{"date": k, "count": v} for k, v in sorted(day_counter.items(), reverse=True)[:14]]

    # 회원/비회원 분류
    member_count = sum(1 for l in filtered if l.get("member_type") == "회원")
    guest_count = len(filtered) - member_count

    # 회원별 접속 이력
    from collections import defaultdict
    filtered_names = {}
    member_visits = defaultdict(list)
    for l in filtered:
        if l.get("username"):
            member_visits[l["username"]].append(l.get("time", ""))
            if l.get("name"):
                filtered_names[l["username"]] = l["name"]
    members = [{"username": k, "name": filtered_names.get(k, k), "count": len(v), "last": max(v) if v else ""}
               for k, v in sorted(member_visits.items(), key=lambda x: -len(x[1]))]

    # 최근 로그 (페이지네이션)
    log_page = int(request.args.get("log_page", 1))
    log_per = 30
    all_recent = list(reversed(filtered))
    log_total = len(all_recent)
    log_pages = (log_total + log_per - 1) // log_per
    log_offset = (log_page - 1) * log_per
    paged = all_recent[log_offset:log_offset + log_per]

    log_list = [{"time": l.get("time", "")[11:19], "ip": l.get("ip", ""),
                 "region": l.get("region", ""), "path": l.get("path", ""),
                 "referrer": l.get("referrer", ""), "device": l.get("device", ""),
                 "username": l.get("username", ""), "name": l.get("name", ""),
                 "member_type": l.get("member_type", "비회원")} for l in paged]

    return jsonify({
        "ok": True,
        "total": len(filtered),
        "today": today_count,
        "unique": unique_ips,
        "search_count": search_count,
        "member_count": member_count,
        "guest_count": guest_count,
        "referrers": referrers,
        "keywords": keywords,
        "devices": devices,
        "daily": daily,
        "members": members,
        "logs": log_list,
        "log_page": log_page,
        "log_pages": log_pages,
        "log_total": log_total,
    })


# ── 견적서 다운로드 ─────────────────────
@app.route(f"{URL_PREFIX}/api/quote/download")
@admin_required
def download_quote():
    """선택 주문 견적서 Excel 다운로드"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    ids_str = request.args.get("ids", "")
    if not ids_str:
        return "주문을 선택해주세요", 400
    ids = [int(x) for x in ids_str.split(",") if x.strip().isdigit()]

    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        placeholders = ",".join(["?"] * len(ids))
        rows = conn.execute(f"SELECT * FROM orders WHERE id IN ({placeholders}) ORDER BY id", ids).fetchall()
        orders = [{c: r[c] for c in r.keys()} for r in rows]
        if not orders:
            return "주문을 찾을 수 없습니다", 404

        # 고객 정보
        username = orders[0].get("username", "")
        customer = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        cust = {c: customer[c] for c in customer.keys()} if customer else {}
        user_level = cust.get("level", "b2c")
    finally:
        conn.close()

    # 상품 링크 + 레벨별 가격 재계산
    from product_db import _conn as prod_conn
    pconn = prod_conn()
    for o in orders:
        code = o.get("product_code", "")
        if code:
            pr = pconn.execute("SELECT link, price_jpy FROM products WHERE internal_code=? OR product_code=? LIMIT 1", (code, code)).fetchone()
            o["product_link"] = pr["link"] if pr else ""
            # 레벨별 가격 재계산
            jpy = pr["price_jpy"] if pr else (o.get("price_jpy", 0) or 0)
            if jpy > 0:
                o["quote_price"] = _calc_vintage_price(jpy, user_level)
            else:
                o["quote_price"] = int("".join(ch for ch in str(o.get("price", "0")) if ch.isdigit()) or 0)
        else:
            o["quote_price"] = int("".join(ch for ch in str(o.get("price", "0")) if ch.isdigit()) or 0)
    pconn.close()

    # Excel 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"

    # 스타일 정의
    thin = Side(style="thin")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center")
    right_al = Alignment(horizontal="right", vertical="center")
    title_font = Font(name="맑은 고딕", size=22, bold=True)
    company_font = Font(name="맑은 고딕", size=14, bold=True)
    label_font = Font(name="맑은 고딕", size=10, bold=True)
    val_font = Font(name="맑은 고딕", size=10)
    hdr_font = Font(name="맑은 고딕", size=10, bold=True, color="000000")
    hdr_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    total_font = Font(name="맑은 고딕", size=11, bold=True)
    note_font = Font(name="맑은 고딕", size=13, bold=True)
    label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 열 너비
    for c, w in [("A", 12), ("B", 22), ("C", 20), ("D", 14), ("E", 30), ("F", 14)]:
        ws.column_dimensions[c].width = w

    # 행 높이
    def set_row(r, h=20):
        ws.row_dimensions[r].height = h

    def cell_style(r, c, val, font=val_font, align=left_al, border=bd, fill=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font; cell.alignment = align; cell.border = border
        if fill: cell.fill = fill
        return cell

    def merge_border(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(row=r, column=c).border = bd

    today = datetime.now().strftime("%Y-%m-%d")

    # ── 제목 ──
    set_row(2, 36)
    merge_border(2, 1, 2, 6)
    cell_style(2, 1, "견 적 서", title_font, center, bd)

    # ── 주문일자 + 판매자 구분 ──
    set_row(3, 18)
    cell_style(3, 1, f"주문일자: {today}", val_font, left_al, Border())
    set_row(4, 18)
    level_text = "B2B" if user_level == "b2b" else "B2C"
    cell_style(4, 1, f"판매자 구분 : {level_text}", Font(name="맑은 고딕", size=10, bold=True, color="0000FF" if user_level=="b2b" else "008000"), left_al, Border())

    # ── 회사명 ──
    set_row(5, 28)
    merge_border(5, 1, 5, 6)
    cell_style(5, 1, "더원 빈티지", company_font, center, bd)

    # ── 고객 정보 ──
    info = [
        ("이름", cust.get("name", "")),
        ("연락처", cust.get("phone", "")),
        ("주소", f'{cust.get("address", "")} {cust.get("address_detail", "")}'),
        ("사업자 명", cust.get("business_name", cust.get("name", ""))),
        ("사업자 번호", cust.get("business_number", "")),
        ("견적내용", "빈티지 명품 구매대행"),
    ]
    for i, (label, value) in enumerate(info):
        r = 7 + i
        set_row(r, 22)
        cell_style(r, 1, label, label_font, center, bd, label_fill)
        merge_border(r, 2, r, 6)
        cell_style(r, 2, value, val_font, left_al, bd)

    # 비고
    set_row(13, 22)
    set_row(14, 22)
    cell_style(13, 1, "비고", label_font, center, bd, label_fill)
    ws.merge_cells("A13:A14")
    merge_border(13, 2, 14, 6)

    # ── 견적내용 제목 ──
    set_row(16, 28)
    merge_border(16, 1, 16, 6)
    cell_style(16, 1, "견 적 내 용", company_font, center, bd)

    # 빈 행
    set_row(17, 6)

    # ── 테이블 헤더 ──
    set_row(18, 24)
    hdrs = ["번호", "주문번호", "상품번호", "브랜드", "상품명", "판매가"]
    for c, h in enumerate(hdrs, 1):
        cell_style(18, c, h, hdr_font, center, bd, hdr_fill)

    # ── 주문 데이터 ──
    total = 0
    for i, o in enumerate(orders):
        r = 19 + i
        set_row(r, 22)
        price_num = o.get("quote_price", 0) or 0
        total += price_num

        cell_style(r, 1, i + 1, val_font, center, bd)
        cell_style(r, 2, o.get("order_number", ""), val_font, center, bd)
        cell_style(r, 3, o.get("product_code", ""), val_font, center, bd)
        cell_style(r, 4, o.get("brand", ""), val_font, center, bd)
        cell_style(r, 5, o.get("product_name", ""), val_font, left_al, bd)
        pc = cell_style(r, 6, price_num, val_font, right_al, bd)
        pc.number_format = '₩#,##0'

    # ── 합계 행 ──
    tr = 19 + len(orders)
    set_row(tr, 24)
    cell_style(tr, 1, "합계", total_font, center, bd, label_fill)
    for c in range(2, 6):
        cell_style(tr, c, "", total_font, center, bd, label_fill)
    tc = cell_style(tr, 6, total, total_font, right_al, bd, label_fill)
    tc.number_format = '₩#,##0'

    # ── Note ──
    nr = tr + 2
    set_row(nr, 18)
    cell_style(nr, 1, "*Note", label_font, left_al, Border())

    set_row(nr+1, 26)
    cell_style(nr+1, 1, "합계", note_font, left_al, Border())
    cell_style(nr+1, 2, f"₩{total:,}", note_font, left_al, Border())

    set_row(nr+3, 18)
    merge_border(nr+3, 1, nr+3, 6)
    cell_style(nr+3, 1, "계좌번호", label_font, left_al, bd)

    set_row(nr+4, 18)
    merge_border(nr+4, 1, nr+4, 6)
    cell_style(nr+4, 1, "하나은행 박수현 307-910169-58305", val_font, left_al, bd)

    # 파일 저장 후 전송
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    customer_name = cust.get("name", username)
    short_date = datetime.now().strftime("%y-%m-%d")
    filename = f"{short_date} {customer_name}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{__import__('urllib.parse', fromlist=['quote']).quote(filename)}"}
    )


# ── 관리자 장바구니 조회 ─────────────────────
@app.route(f"{URL_PREFIX}/api/admin/carts", methods=["GET"])
@admin_required
def admin_carts():
    """전체 고객 장바구니 조회 (고객명 + 품절 체크)"""
    from user_db import _conn
    conn = _conn()
    try:
        rows = conn.execute("""
            SELECT c.*, u.name as customer_name
            FROM cart c LEFT JOIN users u ON c.username = u.username
            ORDER BY c.created_at DESC
        """).fetchall()
        carts = [{col: r[col] for col in r.keys()} for r in rows]

        # 품절 체크는 장바구니 담을 때 개별 실행 (자동 전체 체크 제거)

        # cart 테이블의 is_sold_out 값 사용
        for c in carts:
            is_sold = c.get("is_sold_out", 0)
            checked = c.get("checked_at", "")
            c["sold_out"] = bool(is_sold)
            c["order_status"] = "품절" if is_sold else ("주문가능" if checked else "확인중")
        return jsonify({"ok": True, "carts": carts})
    except Exception as e:
        return jsonify({"ok": True, "carts": []})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/api/admin/cart/<int:item_id>", methods=["DELETE"])
@admin_required
def admin_delete_cart(item_id):
    """관리자: 장바구니 항목 삭제"""
    from user_db import _conn
    conn = _conn()
    try:
        conn.execute("DELETE FROM cart WHERE id=?", (item_id,))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


# ── 문자 발송 API (알리고) ─────────────────────
@app.route(f"{URL_PREFIX}/api/sms/status", methods=["GET"])
@admin_required
def sms_status():
    from aligo_sms import load_config, check_balance, _config
    load_config()
    balance = check_balance() if _config.get("api_key") else {}
    return jsonify({
        "ok": True,
        "config": {"api_key": bool(_config.get("api_key")), "user_id": _config.get("user_id", ""), "sender": _config.get("sender", "")},
        "balance": balance if balance.get("ok") else {},
    })


@app.route(f"{URL_PREFIX}/api/sms/config", methods=["POST"])
@admin_required
def sms_config_save():
    from aligo_sms import save_config, load_config, _config
    data = request.get_json() or {}
    load_config()
    api_key = data.get("api_key", _config.get("api_key", ""))
    user_id = data.get("user_id", "")
    sender = data.get("sender", "")
    save_config(api_key, user_id, sender)
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/api/sms/members", methods=["GET"])
@admin_required
def sms_members():
    """문자 발송 가능한 회원 목록 (전화번호 있는 회원)"""
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        rows = conn.execute("""
            SELECT username, name, phone FROM users
            WHERE phone IS NOT NULL AND phone != '' AND phone != '-'
            ORDER BY name
        """).fetchall()
        members = [{"name": r["name"] or r["username"], "phone": r["phone"]} for r in rows]
        return jsonify({"ok": True, "members": members})
    except Exception as e:
        return jsonify({"ok": False, "members": [], "message": str(e)})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/api/sms/send", methods=["POST"])
@admin_required
def sms_send():
    """문자 발송"""
    from aligo_sms import send_bulk, load_config
    load_config()
    data = request.get_json() or {}
    receivers = data.get("receivers", [])
    msg = data.get("msg", "")
    title = data.get("title", "")
    if not receivers or not msg:
        return jsonify({"ok": False, "message": "수신자와 메시지를 입력해주세요"})
    result = send_bulk(receivers, msg, title)
    return jsonify(result)


# ── Kabinet API (무신사 → 바이마 워크플로우) ─────────────────────
_kabinet_status = {"running": False, "stop_requested": False, "log": []}

# 마진 설정 파일
_MUSINSA_CONFIG_PATH = os.path.join(get_path("db"), "musinsa_config.json")

def _load_musinsa_config() -> dict:
    """무신사→바이마 마진 설정 로드"""
    defaults = {
        "margin_pct": 30.0,           # 마진율 (%)
        "krw_to_jpy_rate": 0.11,      # 원→엔 환율 (1원 = 0.11엔 기본)
        "shipping_jpy": 800,          # 배송비 (엔)
        "buyma_fee_pct": 5.0,         # 바이마 수수료 (%)
        "category_id": "",            # 바이마 カテゴリ ID
        "brand_id": "",               # 바이마 ブランド ID
        "shipping_method": "1062886_1061293",  # 配送方法 ID
        "buying_area": "2002003",     # 買付エリア (한국)
        "buying_city": "000",         # 買付都市 (한국)
        "shipping_area": "2002003",   # 発送エリア (한국)
        "shipping_city": "001",       # 発送都市 (서울)
        "purchase_deadline": 10,      # 구매기한 (일)
        "tariff_included": 0,         # 関税込み (0=なし, 1=込み)
        "tags": "",                   # タグ ID (언더스코어 구분)
        "quantity": 100,              # 買付可数量
        "control": "下書き",           # コントロール (下書き/公開)
        "comment_template": """KABINETのすべての商品は、韓国の現地バイヤーと各ブランドの担当者を通じて
直接買い付けを行いた100%正規品となりますのでご安心してお買い物くださいませ。

※あまりにも安価な商品につきましては、
トラブル防止のため、正規品であるかを十分にご確認の上ご購入ください。
当店では、鑑定基準を満たした正規品のみを取り扱っております。

※安心してお買い物いただくために
価格が安すぎる商品にはご注意ください。
当店は正規店・公式ルートからのみ買い付けを行っており、
全商品100%正規品を保証しております。""",
        "max_items": 50,              # 무신사 수집 최대 개수
    }
    if os.path.exists(_MUSINSA_CONFIG_PATH):
        try:
            with open(_MUSINSA_CONFIG_PATH, "r", encoding="utf-8") as f:
                saved = json.load(f)
                defaults.update(saved)
        except Exception:
            pass
    return defaults

def _save_musinsa_config(cfg: dict):
    """무신사→바이마 마진 설정 저장"""
    os.makedirs(os.path.dirname(_MUSINSA_CONFIG_PATH), exist_ok=True)
    with open(_MUSINSA_CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def _kv_log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    _kabinet_status["log"].append(f"[{ts}] {msg}")
    if len(_kabinet_status["log"]) > 300:
        _kabinet_status["log"] = _kabinet_status["log"][-200:]
    logger.info(f"[Musinsa] {msg}")


def _calc_buyma_price(price_krw: int, cfg: dict) -> int:
    """KRW 가격 → 바이마 출품가 (JPY) 계산"""
    margin = 1 + cfg.get("margin_pct", 30) / 100
    fee = 1 + cfg.get("buyma_fee_pct", 5) / 100
    rate = cfg.get("krw_to_jpy_rate", 0.11)
    shipping = cfg.get("shipping_jpy", 800)
    jpy = math.ceil(price_krw * margin * fee * rate + shipping)
    # 100엔 단위 올림
    return math.ceil(jpy / 100) * 100


@app.route(f"{URL_PREFIX}/api/kabinet/products", methods=["GET"])
@admin_required
def kabinet_products():
    """무신사 수집 상품 목록"""
    from product_db import _conn
    conn = _conn()
    try:
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 100))
        q = request.args.get("q", "").strip()
        offset = (page - 1) * per_page

        where = "WHERE site_id='musinsa'"
        params = []
        if q:
            where += " AND (name LIKE ? OR brand LIKE ? OR product_code LIKE ?)"
            params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])

        total = conn.execute(f"SELECT COUNT(*) FROM products {where}", params).fetchone()[0]
        today = conn.execute(
            f"SELECT COUNT(*) FROM products {where} AND date(created_at)=date('now','localtime')", params
        ).fetchone()[0]
        rows = conn.execute(
            f"SELECT * FROM products {where} ORDER BY created_at DESC LIMIT ? OFFSET ?",
            params + [per_page, offset]
        ).fetchall()
        products = []
        cfg = _load_musinsa_config()
        for r in rows:
            price_krw = r["price_jpy"]  # musinsa는 KRW를 price_jpy 컬럼에 저장
            products.append({
                "id": r["id"],
                "name": r["name"],
                "brand": r["brand"],
                "price_krw": price_krw,
                "price_jpy": _calc_buyma_price(price_krw, cfg),
                "img_url": r["img_url"],
                "link": r["link"],
                "created_at": r["created_at"],
                "product_code": r["product_code"],
                "sizes": json.loads(r["sizes"]) if r["sizes"] and r["sizes"] != "[]" else [],
                "category_id": r["category_id"],
                "description": r["description"] or "",
            })
        return jsonify({
            "ok": True, "products": products,
            "total": total, "today": today,
            "page": page, "pages": (total + per_page - 1) // per_page,
        })
    except Exception as e:
        logger.error(f"무신사 상품 목록 오류: {e}")
        return jsonify({"ok": True, "products": [], "total": 0, "today": 0, "page": 1, "pages": 0})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/api/kabinet/config", methods=["GET"])
@admin_required
def kabinet_config_get():
    """무신사→바이마 마진 설정 조회"""
    return jsonify({"ok": True, "config": _load_musinsa_config()})


@app.route(f"{URL_PREFIX}/api/kabinet/config", methods=["POST"])
@admin_required
def kabinet_config_set():
    """무신사→바이마 마진 설정 저장"""
    data = request.get_json() or {}
    cfg = _load_musinsa_config()
    for key in ["margin_pct", "krw_to_jpy_rate", "shipping_jpy", "buyma_fee_pct",
                "category_id", "brand_id", "shipping_method", "buying_area", "buying_city",
                "shipping_area", "shipping_city", "purchase_deadline", "tariff_included",
                "tags", "quantity", "control", "comment_template", "max_items"]:
        if key in data:
            val = data[key]
            if key in ("margin_pct", "krw_to_jpy_rate", "buyma_fee_pct"):
                val = float(val)
            elif key in ("shipping_jpy", "purchase_deadline", "tariff_included", "quantity", "max_items"):
                val = int(val)
            cfg[key] = val
    _save_musinsa_config(cfg)
    return jsonify({"ok": True, "config": cfg})


@app.route(f"{URL_PREFIX}/api/kabinet/scrape", methods=["POST"])
@admin_required
def kabinet_scrape():
    """무신사 수집 시작"""
    if _kabinet_status["running"]:
        return jsonify({"ok": False, "message": "이미 수집 중입니다"})
    data = request.get_json() or {}
    keyword = data.get("keyword", "").strip()
    search_mode = data.get("search_mode", "keyword")  # keyword / ranking / url
    url = data.get("url", "").strip()
    cfg = _load_musinsa_config()
    max_items = int(data.get("max_items", cfg.get("max_items", 50)))
    if search_mode == "url":
        if not url:
            return jsonify({"ok": False, "message": "URL을 입력해주세요"})
        if "musinsa.com" not in url:
            return jsonify({"ok": False, "message": "무신사 URL을 입력해주세요"})
    elif search_mode == "keyword" and not keyword:
        return jsonify({"ok": False, "message": "키워드를 입력해주세요"})
    t = threading.Thread(target=_run_musinsa_scrape, args=(keyword, max_items, search_mode, url), daemon=True)
    t.start()
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/api/kabinet/stop", methods=["POST"])
@admin_required
def kabinet_stop():
    """무신사 수집 중지"""
    _kabinet_status["stop_requested"] = True
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/api/kabinet/status", methods=["GET"])
@admin_required
def kabinet_scrape_status():
    """무신사 수집 상태"""
    return jsonify({
        "ok": True,
        "running": _kabinet_status["running"],
        "log": "\n".join(_kabinet_status["log"][-100:]),
    })


@app.route(f"{URL_PREFIX}/api/kabinet/delete", methods=["POST"])
@admin_required
def kabinet_delete():
    """무신사 수집 상품 삭제"""
    data = request.get_json() or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"ok": False, "message": "삭제할 상품을 선택해주세요"})
    from product_db import delete_by_ids
    deleted = delete_by_ids(ids)
    return jsonify({"ok": True, "deleted": deleted})


@app.route(f"{URL_PREFIX}/api/kabinet/csv", methods=["POST"])
@admin_required
def kabinet_csv():
    """바이마 업로드용 CSV 생성 (items.csv + colorsizes.csv → ZIP) — 실제 바이마 포맷"""
    import csv
    import io
    import zipfile

    data = request.get_json() or {}
    ids = data.get("ids", [])

    from product_db import _conn
    conn = _conn()
    try:
        if ids:
            placeholders = ",".join("?" for _ in ids)
            rows = conn.execute(
                f"SELECT * FROM products WHERE id IN ({placeholders}) ORDER BY created_at DESC", ids
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM products WHERE site_id='musinsa' ORDER BY created_at DESC LIMIT 500"
            ).fetchall()

        if not rows:
            return jsonify({"ok": False, "message": "상품이 없습니다"})

        cfg = _load_musinsa_config()
        deadline_days = int(cfg.get("purchase_deadline", 10))
        from datetime import timedelta
        deadline = (datetime.now() + timedelta(days=deadline_days)).strftime("%Y-%m-%d")
        control = cfg.get("control", "下書き")
        comment_tpl = cfg.get("comment_template", "")

        # ── items.csv (바이마 실제 포맷) ──
        ITEMS_HEADER = [
            "商品ID", "商品管理番号", "コントロール", "公開ステータス", "商品名",
            "ブランド", "ブランド名", "モデル", "カテゴリ", "シーズン", "テーマ",
            "単価", "買付可数量", "購入期限", "参考価格/通常出品価格", "参考価格",
            "商品コメント", "色サイズ補足", "タグ", "配送方法",
            "買付エリア", "買付都市", "買付ショップ", "発送エリア", "発送都市",
            "関税込み", "出品メモ",
            "商品イメージ1", "商品イメージ2", "商品イメージ3", "商品イメージ4", "商品イメージ5",
            "商品イメージ6", "商品イメージ7", "商品イメージ8", "商品イメージ9", "商品イメージ10",
            "商品イメージ11", "商品イメージ12", "商品イメージ13", "商品イメージ14", "商品イメージ15",
            "商品イメージ16", "商品イメージ17", "商品イメージ18", "商品イメージ19", "商品イメージ20",
            "ブランド品番1", "ブランド品番識別メモ1",
            "買付先名1", "買付先URL1", "買付先説明1",
        ]
        CS_HEADER = [
            "商品ID", "商品管理番号", "商品名", "並び順",
            "サイズ名称", "サイズ単位", "検索用サイズ", "色名称", "色系統",
            "在庫ステータス", "手元に在庫あり数量", "色サイズリプレイス",
        ]

        items_buf = io.StringIO()
        items_writer = csv.writer(items_buf, quoting=csv.QUOTE_ALL)
        items_writer.writerow(ITEMS_HEADER)

        cs_buf = io.StringIO()
        cs_writer = csv.writer(cs_buf, quoting=csv.QUOTE_ALL)
        cs_writer.writerow(CS_HEADER)

        for r in rows:
            price_krw = r["price_jpy"]
            buyma_price = _calc_buyma_price(price_krw, cfg)
            mgmt_no = f"MS-{r['product_code']}"
            name = r["name"] or ""
            brand_name = r["brand"] or ""
            brand_id = cfg.get("brand_id", "")
            img = r["img_url"] or ""
            link = r["link"] or ""
            desc = r["description"] or ""
            product_code = r["product_code"] or ""
            detail_images = json.loads(r["detail_images"]) if r.get("detail_images") and r["detail_images"] != "[]" else []

            # 상품 코멘트 구성
            comment = comment_tpl
            if desc:
                comment = f"{comment}\n\n{desc}" if comment else desc
            if not comment:
                comment = name

            # 이미지 리스트 (최대 20개)
            all_images = [img] if img else []
            for di in detail_images[:19]:
                if isinstance(di, str) and di:
                    all_images.append(di)
            while len(all_images) < 20:
                all_images.append("")

            items_writer.writerow([
                "",                                     # 商品ID (신규=공백)
                mgmt_no,                                # 商品管理番号
                control,                                # コントロール
                "",                                     # 公開ステータス
                name,                                   # 商品名
                brand_id,                               # ブランド (ID)
                brand_name,                             # ブランド名
                "0",                                    # モデル
                cfg.get("category_id", ""),              # カテゴリ
                "0",                                    # シーズン
                "0",                                    # テーマ
                buyma_price,                            # 単価
                cfg.get("quantity", 100),                # 買付可数量
                deadline,                               # 購入期限
                "0",                                    # 参考価格/通常出品価格
                "",                                     # 参考価格
                comment,                                # 商品コメント
                "",                                     # 色サイズ補足
                cfg.get("tags", ""),                     # タグ
                cfg.get("shipping_method", "1062886_1061293"),  # 配送方法
                cfg.get("buying_area", "2002003"),       # 買付エリア
                cfg.get("buying_city", "000"),            # 買付都市
                "",                                     # 買付ショップ
                cfg.get("shipping_area", "2002003"),      # 発送エリア
                cfg.get("shipping_city", "001"),          # 発送都市
                cfg.get("tariff_included", 0),            # 関税込み
                "",                                     # 出品メモ
                *all_images,                            # 商品イメージ1~20
                product_code,                           # ブランド品番1
                "",                                     # ブランド品番識別メモ1
                "MUSINSA",                              # 買付先名1
                link,                                   # 買付先URL1
                "",                                     # 買付先説明1
            ])

            # ── colorsizes.csv ──
            sizes = json.loads(r["sizes"]) if r["sizes"] and r["sizes"] != "[]" else []
            if not sizes:
                sizes = [{"name": "f", "search_size": "0"}]
            for idx, sz in enumerate(sizes):
                if isinstance(sz, str):
                    sz_name = sz
                    sz_search = "0"
                else:
                    sz_name = sz.get("name", str(sz))
                    sz_search = sz.get("search_size", "0")
                cs_writer.writerow([
                    "",             # 商品ID
                    mgmt_no,        # 商品管理番号
                    name,           # 商品名
                    idx + 1,        # 並び順
                    sz_name,        # サイズ名称
                    "",             # サイズ単位
                    sz_search,      # 検索用サイズ
                    "",             # 色名称
                    "",             # 色系統
                    2,              # 在庫ステータス (2=買付可能)
                    "",             # 手元に在庫あり数量
                    "",             # 色サイズリプレイス
                ])

        # ZIP 생성 (UTF-8)
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("items.csv", items_buf.getvalue())
            zf.writestr("colorsizes.csv", cs_buf.getvalue())
        zip_buf.seek(0)

        resp = make_response(zip_buf.read())
        resp.headers["Content-Type"] = "application/zip"
        resp.headers["Content-Disposition"] = f"attachment; filename=buyma_upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        return resp
    except Exception as e:
        logger.error(f"바이마 CSV 생성 오류: {e}")
        return jsonify({"ok": False, "message": str(e)})
    finally:
        conn.close()


def _run_musinsa_scrape(keyword, max_items=50, search_mode="keyword", url=""):
    """무신사 크롤링 — 포이즌서치 방식 (키워드/랭킹 → 무한스크롤 → 상세 페이지 방문)"""
    _kabinet_status["running"] = True
    _kabinet_status["stop_requested"] = False
    _kabinet_status["log"] = []

    mode_label = {"keyword": "키워드 검색", "ranking": "랭킹", "url": "URL 직접"}
    _kv_log(f"무신사 수집 시작 ({mode_label.get(search_mode, search_mode)})")
    if search_mode == "keyword":
        _kv_log(f"키워드: {keyword}")
    elif search_mode == "url":
        _kv_log(f"URL: {url}")
    _kv_log(f"최대 {max_items}개")

    collected = []
    try:
        from playwright.sync_api import sync_playwright
        import time
        import re as re_mod
        import random

        MUSINSA_SEARCH_URL = "https://www.musinsa.com/search/musinsa/goods"
        MUSINSA_RANKING_URL = (
            "https://www.musinsa.com/main/musinsa/ranking"
            "?gf=A&storeCode=musinsa&sectionId=200&contentsId="
            "&categoryCode=103000&ageBand=AGE_BAND_ALL&subPan=product"
        )
        COOKIE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "musinsa_data", "musinsa_cookies.json")
        if not os.path.exists(COOKIE_FILE):
            COOKIE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "musinsa_cookies.json")

        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True, channel="chrome",
                args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
            )
            context = browser.new_context(
                viewport={"width": 960, "height": 648},
                locale="ko-KR", timezone_id="Asia/Seoul",
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            )
            # 쿠키 로드
            if os.path.exists(COOKIE_FILE):
                try:
                    with open(COOKIE_FILE, "r", encoding="utf-8") as f:
                        cookies = json.load(f)
                    musinsa_cookies = [c for c in cookies if "musinsa.com" in c.get("domain", "")]
                    if musinsa_cookies:
                        context.add_cookies(cookies)
                        _kv_log("쿠키 로드 완료")
                except Exception:
                    pass

            page = context.new_page()
            page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
                Object.defineProperty(navigator, 'languages', {get: () => ['ko-KR','ko','en-US','en']});
            """)

            # 검색 URL 결정
            if search_mode == "url":
                search_url = url
            elif search_mode == "ranking":
                search_url = MUSINSA_RANKING_URL
            else:
                search_url = f"{MUSINSA_SEARCH_URL}?q={keyword}"

            _kv_log(f"검색 페이지 로딩...")
            page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
            time.sleep(2)

            # ── 1단계: 무한 스크롤로 상품 링크 수집 ──
            _kv_log("무한 스크롤로 상품 링크 수집 중...")
            product_links = []
            scroll_count = 0
            max_scrolls = 200
            no_new = 0

            while len(product_links) < max_items and scroll_count < max_scrolls:
                if _kabinet_status["stop_requested"]:
                    _kv_log("사용자 요청으로 중지")
                    break
                scroll_count += 1

                current_links = page.evaluate("""() => {
                    const links = [];
                    document.querySelectorAll('a[href*="/products/"]').forEach(a => {
                        if (a.href && !links.includes(a.href)) links.push(a.href);
                    });
                    return links;
                }""")
                new_links = [lk for lk in current_links if lk not in product_links]

                if new_links:
                    product_links.extend(new_links)
                    no_new = 0
                    _kv_log(f"  스크롤 {scroll_count}: {len(product_links)}/{max_items}개 (+{len(new_links)})")
                else:
                    no_new += 1
                    if no_new >= 5:
                        _kv_log("  더 이상 새 상품 없음")
                        break

                if len(product_links) >= max_items:
                    break

                # 단계적 스크롤
                try:
                    sh = page.evaluate("document.body.scrollHeight")
                    cy = page.evaluate("window.pageYOffset")
                    vh = page.evaluate("window.innerHeight")
                    step = cy
                    while step < sh:
                        step = min(step + vh, sh)
                        page.evaluate(f"window.scrollTo(0, {step})")
                        time.sleep(0.5)
                except Exception:
                    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(random.uniform(2.0, 3.0))

            product_links = product_links[:max_items]
            _kv_log(f"수집 링크: {len(product_links)}개")

            # ── 2단계: 각 상품 상세 페이지 방문하여 정보 추출 ──
            _kv_log("상세 페이지 방문 시작...")
            for idx, product_url in enumerate(product_links, 1):
                if _kabinet_status["stop_requested"]:
                    _kv_log("사용자 요청으로 중지")
                    break

                try:
                    page.goto(product_url, wait_until="domcontentloaded", timeout=30000)
                    time.sleep(3)

                    # 팝업 닫기
                    try:
                        btn = page.locator("button:has-text('오늘 그만보기')").first
                        if btn.count() > 0 and btn.is_visible():
                            btn.click()
                            time.sleep(0.3)
                    except Exception:
                        pass

                    # 상품 정보 추출 (포이즌 서치 원본 로직 — Playwright locator)
                    info = {"name": "", "price": 0, "brand": "", "product_code": "", "image_url": ""}

                    # ── 제품명 추출 ──
                    try:
                        ne = page.locator('span[class*="GoodsName"]').first
                        if ne.count() > 0:
                            info["name"] = ne.text_content().strip()
                    except Exception:
                        pass
                    if not info["name"]:
                        try:
                            ne2 = page.locator('span.text-title_18px_med').first
                            if ne2.count() > 0:
                                info["name"] = ne2.text_content().strip()
                        except Exception:
                            pass

                    # ── 가격 추출 (포이즌 서치 4단계 로직) ──
                    price = None
                    # 1순위: span.text-title_18px_semi.text-red (최대혜택가)
                    try:
                        pe = page.locator('span.text-title_18px_semi.text-red').first
                        if pe.count() > 0 and pe.is_visible(timeout=3000):
                            pt = pe.text_content()
                            if pt and '원' in pt:
                                nums = re_mod.findall(r'\d+', pt.replace(',', ''))
                                if nums:
                                    price = int(''.join(nums))
                    except Exception:
                        pass
                    # 2순위: "XX,XXX원 최대혜택가" 텍스트 패턴
                    if not price:
                        try:
                            price = page.evaluate("""() => {
                                const t = document.body.innerText;
                                const m = t.match(/([0-9]{1,3}(?:,?[0-9]{3})*)\\s*원\\s*최대혜택가/);
                                if (m && m[1]) return parseInt(m[1].replace(/,/g,''));
                                return null;
                            }""")
                            if not price or price <= 0:
                                price = None
                        except Exception:
                            pass
                    # 3순위: span.text-red 중에서 "원" 포함
                    if not price:
                        try:
                            reds = page.locator('span.text-red').all()
                            for sp in reds:
                                try:
                                    t = sp.text_content()
                                    if t and '원' in t:
                                        nums = re_mod.findall(r'\d+', t.replace(',', ''))
                                        if nums:
                                            tp = int(''.join(nums))
                                            if 10000 <= tp <= 10000000:
                                                price = tp
                                                break
                                except Exception:
                                    continue
                        except Exception:
                            pass
                    # 4순위: 모든 span에서 "원" 포함
                    if not price:
                        try:
                            spans = page.locator('span').all()
                            for sp in spans[:100]:
                                try:
                                    t = sp.text_content()
                                    if t and '원' in t and len(t) < 20:
                                        nums = re_mod.findall(r'\d+', t.replace(',', ''))
                                        if nums:
                                            tp = int(''.join(nums))
                                            if 10000 <= tp <= 10000000:
                                                price = tp
                                                break
                                except Exception:
                                    continue
                        except Exception:
                            pass
                    info["price"] = price or 0

                    # ── 브랜드 추출 ──
                    try:
                        be = page.locator('a[href*="/brands/"]').first
                        if be.count() > 0:
                            info["brand"] = be.text_content().strip()
                    except Exception:
                        pass

                    # ── 품번 추출 ──
                    try:
                        all_text = page.evaluate("() => document.body.innerText")
                        cm = re_mod.search(r'(품번|모델번호|상품코드)\s*[:：\s]+([A-Z0-9][A-Z0-9-]+)', all_text, re_mod.IGNORECASE)
                        if cm and cm.group(2) and len(cm.group(2)) >= 5:
                            info["product_code"] = cm.group(2).strip()
                    except Exception:
                        pass

                    # ── 이미지 추출 ──
                    try:
                        img1 = page.locator('img[alt="Thumbnail 0"]').first
                        if img1.count() > 0:
                            src = img1.get_attribute('src')
                            if src and 'image.msscdn.net' in src:
                                info["image_url"] = src.split('?')[0]
                    except Exception:
                        pass
                    if not info["image_url"]:
                        try:
                            sw = page.locator('div[class*="Swiper"] img').first
                            if sw.count() > 0:
                                src = sw.get_attribute('src')
                                if src:
                                    info["image_url"] = src.split('?')[0]
                        except Exception:
                            pass

                    if not info or not info.get("name"):
                        _kv_log(f"  [{idx}/{len(product_links)}] 정보 추출 실패 — 건너뜀")
                        continue

                    # product_code가 없으면 URL에서 추출
                    code = info.get("product_code", "")
                    if not code:
                        m = re_mod.search(r"/products/(\d+)", product_url)
                        code = m.group(1) if m else str(idx)

                    # 이미지 URL 정제
                    img_url = info.get("image_url", "")
                    if img_url:
                        if "/thumbnails/" in img_url:
                            img_url = img_url.replace("/thumbnails/", "/images/")

                    collected.append({
                        "site_id": "musinsa",
                        "source_type": "musinsa",
                        "product_code": code,
                        "name": info.get("name", ""),
                        "brand": info.get("brand", ""),
                        "price_jpy": info.get("price", 0),
                        "img_url": img_url,
                        "link": product_url,
                        "category_id": "",
                        "scraped_at": datetime.now().isoformat(),
                    })
                    _kv_log(f"  [{idx}/{len(product_links)}] {info.get('brand','')} {info.get('name','')[:40]} — {info.get('price',0):,}원")

                    time.sleep(random.uniform(1.0, 2.0))

                except Exception as e:
                    _kv_log(f"  [{idx}/{len(product_links)}] 오류: {e}")

            browser.close()

        # DB 저장
        if collected:
            _kv_log(f"DB 저장 시작: {len(collected)}개")
            from product_db import insert_products
            saved = insert_products(collected)
            _kv_log(f"DB 저장 완료: {saved}개")
        else:
            _kv_log("수집된 상품 없음")

    except Exception as e:
        _kv_log(f"오류: {e}")
        import traceback
        _kv_log(traceback.format_exc())
    finally:
        _kabinet_status["running"] = False
        _kv_log(f"수집 종료 (총 {len(collected)}개)")


@app.route(f"{URL_PREFIX}/api/vintage-db-stats", methods=["GET"])
@admin_required
def vintage_db_stats():
    """빈티지 DB 총 수집량 통계"""
    from product_db import _conn
    conn = _conn()
    try:
        total = conn.execute("SELECT COUNT(*) FROM products WHERE source_type='vintage'").fetchone()[0]
        brands = conn.execute("""
            SELECT brand, COUNT(*) as cnt FROM products
            WHERE source_type='vintage' GROUP BY brand ORDER BY cnt DESC
        """).fetchall()
        brand_list = [{"name": r[0], "count": r[1]} for r in brands]
        return jsonify({"ok": True, "total": total, "brands": brand_list})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/scrape/tasks", methods=["GET"])
@admin_required
def get_scrape_tasks():
    _init_scrape_tasks_db()
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        rows = conn.execute("SELECT * FROM scrape_tasks ORDER BY id ASC").fetchall()
        return jsonify({"ok": True, "tasks": [{c: r[c] for c in r.keys()} for r in rows]})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/scrape/tasks", methods=["POST"])
@admin_required
def add_scrape_task():
    _init_scrape_tasks_db()
    data = request.json or {}
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("""INSERT INTO scrape_tasks (site, site_name, cat, cat_name, brand, brand_name, pages, total_items, total_pages)
                        VALUES (?,?,?,?,?,?,?,?,?)""",
                     (data.get("site","2ndstreet"), data.get("siteName",""), data.get("cat",""),
                      data.get("catName",""), data.get("brand",""), data.get("brandName",""),
                      data.get("pages",""), data.get("totalItems",0), data.get("totalPages",0)))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/scrape/tasks/<int:task_id>", methods=["PATCH"])
@admin_required
def update_scrape_task(task_id):
    data = request.json or {}
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        updates, params = [], []
        for k in ["status", "count"]:
            if k in data:
                updates.append(f"{k}=?")
                params.append(data[k])
        if updates:
            params.append(task_id)
            conn.execute(f"UPDATE scrape_tasks SET {','.join(updates)} WHERE id=?", params)
            conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/scrape/tasks/<int:task_id>", methods=["DELETE"])
@admin_required
def delete_scrape_task(task_id):
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("DELETE FROM scrape_tasks WHERE id=?", (task_id,))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/scrape/tasks/clear", methods=["DELETE"])
@admin_required
def clear_scrape_tasks():
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("DELETE FROM scrape_tasks")
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


# ── 수집 큐 (예약 실행) ──────────────────────────
import queue as _queue_mod
_scrape_queue = _queue_mod.Queue()
_queue_worker_started = False

def _start_queue_worker():
    """큐 워커: 큐에 작업이 들어오면 순차 실행"""
    global _queue_worker_started
    if _queue_worker_started:
        return
    _queue_worker_started = True

    # 서버 재시작 시 "예약"/"수집중" 상태 작업을 큐에 복구
    try:
        import sqlite3 as _sq
        _db = os.path.join(get_path("db"), "users.db")
        _c = _sq.connect(_db)
        _c.row_factory = _sq.Row
        _stuck = _c.execute("SELECT id FROM scrape_tasks WHERE status IN ('예약','수집중') ORDER BY id").fetchall()
        for _r in _stuck:
            _c.execute("UPDATE scrape_tasks SET status='예약' WHERE id=?", (_r["id"],))
            _scrape_queue.put(_r["id"])
        _c.commit()
        _c.close()
        if _stuck:
            push_log(f"🔄 서버 재시작: {len(_stuck)}개 예약 작업 큐에 복구")
    except Exception:
        pass

    def _worker():
        while True:
            task_id = _scrape_queue.get()
            try:
                import sqlite3
                db_path = os.path.join(get_path("db"), "users.db")
                conn = sqlite3.connect(db_path)
                conn.row_factory = sqlite3.Row
                r = conn.execute("SELECT * FROM scrape_tasks WHERE id=?", (task_id,)).fetchone()
                conn.close()
                if not r:
                    continue

                push_log(f"📋 큐 실행: {r['brand_name'] or '전체'} / {r['cat_name'] or '전체'} (p.{r['pages'] or '전체'})")

                # 상태 → 수집중
                conn = sqlite3.connect(db_path)
                conn.execute("UPDATE scrape_tasks SET status='수집중' WHERE id=?", (task_id,))
                conn.commit()
                conn.close()

                status["stop_requested"] = False
                status["paused"] = False
                status["scraping"] = True

                import asyncio
                from secondst_crawler import scrape_2ndstreet, set_app_status as set_2nd_status
                set_2nd_status(status)
                result = asyncio.run(scrape_2ndstreet(
                    status_callback=push_log,
                    category=r["cat"],
                    pages=r["pages"] or "",
                    brand_code=r["brand"],
                ))
                count = result.get("total_saved", 0) if isinstance(result, dict) else 0

                conn = sqlite3.connect(db_path)
                conn.execute("UPDATE scrape_tasks SET status='완료', count=? WHERE id=?", (count, task_id))
                conn.commit()
                conn.close()

                push_log(f"✅ 큐 완료: {r['brand_name'] or '전체'} — {count}개")

                # 수집 완료 → 300개마다 NAS 자동 내보내기 (윈도우에서만)
                import platform as _pf
                if _pf.system() == "Windows" and count > 0:
                    _nas_export_acc = getattr(_worker, '_nas_acc', 0) + count
                    _worker._nas_acc = _nas_export_acc
                    if _nas_export_acc >= 300:
                        try:
                            export_all_to_nas()
                            push_log(f"📤 {_nas_export_acc}개 수집 → NAS 자동 내보내기")
                            _worker._nas_acc = 0
                        except Exception:
                            pass

            except Exception as e:
                conn = sqlite3.connect(db_path)
                conn.execute("UPDATE scrape_tasks SET status='오류' WHERE id=?", (task_id,))
                conn.commit()
                conn.close()
                push_log(f"❌ 큐 오류: {e}")
            finally:
                status["scraping"] = False
                _scrape_queue.task_done()

    threading.Thread(target=_worker, daemon=True).start()


@app.route(f"{URL_PREFIX}/scrape/queue", methods=["POST"])
@admin_required
def enqueue_tasks():
    """선택한 작업을 큐에 추가"""
    _start_queue_worker()
    data = request.json or {}
    task_ids = data.get("ids", [])
    if not task_ids:
        return jsonify({"ok": False, "message": "작업을 선택해주세요"})

    import sqlite3
    db_path = os.path.join(get_path("db"), "users.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    added = 0
    for tid in task_ids:
        r = conn.execute("SELECT status FROM scrape_tasks WHERE id=?", (tid,)).fetchone()
        if r and r["status"] in ("대기", "예약", "오류"):
            conn.execute("UPDATE scrape_tasks SET status='예약' WHERE id=?", (tid,))
            _scrape_queue.put(tid)
            added += 1
    conn.commit()
    conn.close()

    queue_size = _scrape_queue.qsize()
    push_log(f"⏰ {added}개 작업 큐에 예약됨 (대기 {queue_size}개)")
    return jsonify({"ok": True, "message": f"{added}개 예약 완료 (큐 {queue_size}개)", "queue_size": queue_size})


@app.route(f"{URL_PREFIX}/scrape/queue/status")
@admin_required
def queue_status():
    """큐 상태 조회"""
    return jsonify({"ok": True, "queue_size": _scrape_queue.qsize()})


@app.route(f"{URL_PREFIX}/scrape/stop-all", methods=["POST"])
@admin_required
def stop_all_tasks():
    """현재 수집 중지 + 큐 비우기 + 예약 상태 → 대기로"""
    import sqlite3
    # 1) 수집 강제 중지
    status["scraping"] = False
    status["stop_requested"] = True
    status["paused"] = False
    try:
        import asyncio
        from secondst_crawler import force_close_browser as fc
        asyncio.run(fc())
    except Exception:
        pass

    # 2) 큐 비우기
    while not _scrape_queue.empty():
        try:
            _scrape_queue.get_nowait()
            _scrape_queue.task_done()
        except Exception:
            break

    # 3) 예약/수집중 상태 → 대기로
    db_path = os.path.join(get_path("db"), "users.db")
    conn = sqlite3.connect(db_path)
    conn.execute("UPDATE scrape_tasks SET status='대기' WHERE status IN ('예약','수집중')")
    conn.commit()
    conn.close()

    push_log("⏹ 전체 멈춤: 수집 중지 + 큐 비우기 + 예약 → 대기")
    try:
        from notifier import send_telegram
        send_telegram("⏹ <b>전체 멈춤</b>\n수집 중지 + 큐 비우기 완료")
    except Exception:
        pass
    return jsonify({"ok": True, "message": "전체 멈춤 완료"})


@app.route(f"{URL_PREFIX}/scrape/check-count")
@admin_required
def scrape_check_count():
    """2ndstreet 검색 결과 상품 수량 확인 (Playwright 사용)"""
    category = request.args.get("category", "")
    brand = request.args.get("brand", "")
    try:
        import asyncio
        from playwright.async_api import async_playwright

        async def _check():
            pw = await async_playwright().start()
            browser = await pw.chromium.launch(headless=False, args=["--disable-translate", "--lang=ja"])
            ctx = await browser.new_context(locale="ja-JP", extra_http_headers={"Accept-Language": "ja"})
            page = await ctx.new_page()
            params = []
            if category:
                params.append(f"category={category}")
            if brand and brand.startswith("kw:"):
                params.append(f"keyword={brand[3:]}")
            elif brand:
                params.append(f"brand%5B%5D={brand}")
            params.append("sortBy=recommend&page=1")
            url = "https://www.2ndstreet.jp/search?" + "&".join(params)
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(6)
            # 팝업 제거
            await page.evaluate("""() => {
                document.body.classList.remove('zigzag-worldshopping-style-body-lock');
                document.body.style.overflow = 'auto';
                const btn = document.querySelector('#onetrust-accept-btn-handler');
                if (btn) btn.click();
            }""")
            await asyncio.sleep(3)
            # 검색 결과 수량 파싱 (최대 3회 재시도)
            result = ''
            for attempt in range(3):
                result = await page.evaluate(r"""() => {
                    const els = document.querySelectorAll('*');
                    for (const el of els) {
                        const t = el.innerText || '';
                        const m = t.match(/検索結果[：:]\s*([\d,]+)\s*点/);
                        if (m) return m[1];
                    }
                    return '';
                }""")
                if result:
                    break
                await page.wait_for_timeout(3000)
            await browser.close()
            await pw.stop()
            return result

        # Flask 내부 이벤트 루프 충돌 방지: 별도 스레드에서 실행
        import concurrent.futures
        with concurrent.futures.ThreadPoolExecutor() as pool:
            total_text = pool.submit(lambda: asyncio.run(_check())).result(timeout=60)
        if total_text:
            total_items = int(total_text.replace(",", ""))
            total_pages = (total_items + 29) // 30
            return jsonify({"ok": True, "total_items": total_items, "total_pages": total_pages})
        return jsonify({"ok": False, "total_items": 0, "total_pages": 0, "message": "수량 파싱 실패"})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e), "total_items": 0, "total_pages": 0})


@app.route(f"{URL_PREFIX}/scrape", methods=["POST"])
@admin_required
def api_scrape_sync():
    """작업리스트용 수집 API (완료까지 대기 후 결과 반환)"""
    site_id = request.args.get("site", "2ndstreet")
    category_id = request.args.get("category", "")
    brand_code = request.args.get("brand", "")
    pages = request.args.get("pages", "")
    max_items = request.args.get("max_items", 0, type=int)

    if status["scraping"]:
        return jsonify({"ok": False, "message": "이미 수집 진행 중", "count": 0})

    # 이전 중지 요청 리셋
    status["stop_requested"] = False
    status["paused"] = False
    push_log(f"📋 작업리스트 수집 시작: {site_id} / {category_id} / {brand_code}")
    status["scraping"] = True
    count = 0
    try:
        import asyncio
        from secondst_crawler import scrape_2ndstreet, set_app_status as set_2nd_status
        set_2nd_status(status)
        kwargs = dict(
            status_callback=push_log,
            category=category_id,
            pages=pages,
            brand_code=brand_code,
        )
        if max_items > 0:
            kwargs["max_items"] = max_items
        result = asyncio.run(scrape_2ndstreet(**kwargs))
        if isinstance(result, dict):
            count = result.get("total_saved", 0)
        else:
            count = len(result) if result else 0
        status["product_count"] = count
        # 이력 저장
        try:
            from scrape_history import add_history
            from site_config import get_brands as get_site_brands
            brand_name = ""
            if brand_code:
                brands_map = get_site_brands(site_id)
                brand_name = brands_map.get(brand_code, brand_code)
            add_history(site_id=site_id, category_id=category_id or "전체",
                       product_count=count, brand=brand_name)
        except Exception:
            pass
    except Exception as e:
        push_log(f"❌ 수집 오류: {e}")
        return jsonify({"ok": False, "message": str(e), "count": 0})
    finally:
        status["scraping"] = False

    # 수집 완료 → NAS로 자동 내보내기 (윈도우에서만)
    import platform
    if platform.system() == "Windows":
        try:
            export_all_to_nas()
            push_log(f"📤 수집 완료 → NAS 자동 내보내기 완료")
        except Exception as e:
            push_log(f"⚠️ NAS 내보내기 실패: {e}")

    return jsonify({"ok": True, "count": count, "message": f"수집 완료: {count}개"})


# ── 사이트/카테고리 API ────────────────────────

@app.route(f"{URL_PREFIX}/sites", methods=["GET"])
@admin_required
def api_sites():
    """사이트/카테고리 트리 반환"""
    return jsonify(get_sites_for_ui())


@app.route(f"{URL_PREFIX}/scrape-history", methods=["GET"])
@admin_required
def api_scrape_history():
    """수집 이력 반환"""
    limit = request.args.get("limit", 50, type=int)
    return jsonify(get_scrape_history(limit))


@app.route(f"{URL_PREFIX}/scrape-history", methods=["DELETE"])
@admin_required
def api_scrape_history_clear():
    """수집 이력 전체 삭제"""
    from scrape_history import _save
    _save([])
    return jsonify({"ok": True})


# ── 빅데이터 관리 API ──────────────────────────

@app.route(f"{URL_PREFIX}/bigdata/stats", methods=["GET"])
@admin_required
def api_bigdata_stats():
    """빅데이터 통계"""
    return jsonify(bigdata_get_stats())


@app.route(f"{URL_PREFIX}/bigdata/products", methods=["GET"])
@admin_required
def api_bigdata_products():
    """빅데이터 상품 검색"""
    return jsonify(bigdata_search(
        query=request.args.get("q", ""),
        site_id=request.args.get("site_id", ""),
        category_id=request.args.get("category_id", ""),
        brand=request.args.get("brand", ""),
        cafe_status=request.args.get("cafe_status", ""),
        page=request.args.get("page", 1, type=int),
        per_page=request.args.get("per_page", 50, type=int),
    ))


@app.route(f"{URL_PREFIX}/bigdata/delete-selected", methods=["POST"])
@admin_required
def api_bigdata_delete_selected():
    """선택된 상품 삭제 (ID 리스트)"""
    data = request.json or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"ok": False, "message": "삭제할 상품을 선택하세요"})
    count = bigdata_delete_ids(ids)
    return jsonify({"ok": True, "deleted": count, "message": f"{count}개 삭제 완료"})


@app.route(f"{URL_PREFIX}/bigdata/brands", methods=["GET"])
@admin_required
def api_bigdata_brands():
    """빅데이터 브랜드 목록"""
    return jsonify(bigdata_get_brands())


@app.route(f"{URL_PREFIX}/bigdata/delete", methods=["POST"])
@admin_required
def api_bigdata_delete():
    """빅데이터 삭제"""
    data = request.json or {}
    scope = data.get("scope", "")
    if scope == "all":
        count = bigdata_delete_all()
        return jsonify({"ok": True, "deleted": count, "message": f"전체 {count}개 삭제"})
    elif scope == "site":
        site_id = data.get("site_id", "")
        if not site_id:
            return jsonify({"ok": False, "message": "site_id 필요"})
        count = bigdata_delete_site(site_id)
        return jsonify({"ok": True, "deleted": count, "message": f"{site_id} {count}개 삭제"})
    return jsonify({"ok": False, "message": "scope 지정 필요 (all 또는 site)"})


@app.route(f"{URL_PREFIX}/bigdata/download")
@admin_required
def api_bigdata_download():
    """빅데이터 CSV 다운로드"""
    import io
    import csv as csv_mod
    from flask import send_file

    q = request.args.get("q", "")
    site_id = request.args.get("site_id", "")
    brand = request.args.get("brand", "")

    products = bigdata_export_csv(query=q, site_id=site_id, brand=brand)

    # CSV 헤더 (병합 시 이 컬럼명 그대로 사용)
    fields = [
        "site_id", "category_id", "product_code", "brand", "brand_ko",
        "name", "name_ko", "price_jpy", "original_price", "discount_rate",
        "in_stock", "link", "img_url", "cafe_status", "cafe_uploaded_at", "created_at",
    ]

    buf = io.StringIO()
    writer = csv_mod.DictWriter(buf, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for p in products:
        writer.writerow(p)

    output = io.BytesIO()
    output.write(b'\xef\xbb\xbf')  # UTF-8 BOM
    output.write(buf.getvalue().encode("utf-8"))
    output.seek(0)

    suffix = f"_{site_id}" if site_id else ""
    filename = f"bigdata{suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return send_file(
        output,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=filename,
    )


@app.route(f"{URL_PREFIX}/bigdata/merge", methods=["POST"])
@admin_required
def api_bigdata_merge():
    """CSV 파일 업로드 + 병합 (created_at 기준 최신 데이터 우선)"""
    import csv as csv_mod
    import io

    if "file" not in request.files:
        return jsonify({"ok": False, "message": "파일이 없습니다"})

    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith(".csv"):
        return jsonify({"ok": False, "message": "CSV 파일만 업로드 가능합니다"})

    try:
        content = file.read().decode("utf-8-sig")  # BOM 처리
        reader = csv_mod.DictReader(io.StringIO(content))
        rows = list(reader)

        if not rows:
            return jsonify({"ok": False, "message": "CSV 파일이 비어있습니다"})

        # 필수 컬럼 확인
        required = {"site_id", "product_code", "price_jpy"}
        header_set = set(reader.fieldnames or [])
        missing = required - header_set
        if missing:
            return jsonify({"ok": False, "message": f"필수 컬럼 누락: {', '.join(missing)}"})

        result = bigdata_merge(rows)
        msg = f"병합 완료: 신규 {result['inserted']}개, 업데이트 {result['updated']}개, 스킵 {result['skipped']}개"
        push_log(f"📥 CSV {msg} (파일: {file.filename})")
        return jsonify({"ok": True, "message": msg, **result})
    except Exception as e:
        logger.error(f"CSV 병합 오류: {e}")
        return jsonify({"ok": False, "message": f"병합 실패: {str(e)}"})


# ── 카페 모니터 & 텔레그램 봇 API ─────────────

@app.route(f"{URL_PREFIX}/monitor/status", methods=["GET"])
@admin_required
def api_monitor_status():
    """모니터/봇 상태"""
    return jsonify({
        "monitor_running": is_monitoring(),
        "bot_running": is_bot_running(),
    })


@app.route(f"{URL_PREFIX}/monitor/start", methods=["POST"])
@admin_required
def api_monitor_start():
    """카페 모니터 + 텔레그램 봇 시작"""
    data = request.json or {}
    interval = data.get("interval", 180)

    monitor_ok = start_monitor(log_callback=push_log, interval=interval)
    bot_ok = start_bot(log_callback=push_log)

    return jsonify({
        "ok": True,
        "monitor": "시작됨" if monitor_ok else "이미 실행중",
        "bot": "시작됨" if bot_ok else "이미 실행중",
    })


@app.route(f"{URL_PREFIX}/monitor/stop", methods=["POST"])
@admin_required
def api_monitor_stop():
    """카페 모니터 + 텔레그램 봇 종료"""
    stop_monitor()
    stop_bot()
    return jsonify({"ok": True, "message": "모니터 & 봇 종료"})


# ── 카페 업로드 스케줄 API ────────────────────

@app.route(f"{URL_PREFIX}/cafe-schedule", methods=["GET"])
@admin_required
def api_get_schedule():
    """스케줄 설정 조회"""
    slots = load_schedule()
    # 현재 등록된 잡 상태도 포함
    for slot in slots:
        job_id = f"cafe_schedule_{slot['id']}"
        job = scheduler.get_job(job_id)
        slot["registered"] = job is not None
        if job and job.next_run_time:
            slot["next_run"] = job.next_run_time.strftime("%Y-%m-%d %H:%M")
        else:
            slot["next_run"] = None
    return jsonify({"ok": True, "slots": slots})


@app.route(f"{URL_PREFIX}/cafe-schedule", methods=["POST"])
@admin_required
def api_save_schedule():
    """스케줄 설정 저장 + 잡 재등록"""
    data = request.json or {}
    slots = data.get("slots", [])
    if not isinstance(slots, list) or len(slots) != 4:
        return jsonify({"ok": False, "error": "4개 슬롯 필요"}), 400

    save_schedule(slots)
    _register_schedule_jobs()
    push_log("📅 카페 업로드 스케줄 설정이 저장되었습니다")
    return jsonify({"ok": True})


# ── 빈티지 카페 스케줄 API ────────────────────

@app.route(f"{URL_PREFIX}/vt-cafe-schedule", methods=["GET"])
@admin_required
def api_get_vt_schedule():
    from cafe_schedule import load_vt_schedule
    slots = load_vt_schedule()
    for slot in slots:
        job_id = f"vt_cafe_{slot['id']}"
        job = scheduler.get_job(job_id)
        slot["registered"] = job is not None
        slot["next_run"] = job.next_run_time.strftime("%Y-%m-%d %H:%M") if job and job.next_run_time else None
    return jsonify({"ok": True, "slots": slots})


@app.route(f"{URL_PREFIX}/vt-cafe-schedule", methods=["POST"])
@admin_required
def api_save_vt_schedule():
    from cafe_schedule import save_vt_schedule
    data = request.json or {}
    slots = data.get("slots", [])
    if not isinstance(slots, list) or len(slots) != 4:
        return jsonify({"ok": False, "error": "4개 슬롯 필요"}), 400
    save_vt_schedule(slots)
    _register_vt_schedule_jobs()
    push_log("📅 빈티지 카페 스케줄 설정이 저장되었습니다")
    return jsonify({"ok": True})


# ── 업로드 체크 자동 확인 스케줄 API ──────────────

@app.route(f"{URL_PREFIX}/check-schedule", methods=["GET"])
@admin_required
def api_get_check_schedule():
    """업로드 체크 스케줄 설정 조회"""
    sched = load_check_schedule()
    job = scheduler.get_job("upload_check_auto")
    sched["registered"] = job is not None
    if job and job.next_run_time:
        sched["next_run"] = job.next_run_time.strftime("%Y-%m-%d %H:%M")
    else:
        sched["next_run"] = None
    return jsonify({"ok": True, "schedule": sched})


@app.route(f"{URL_PREFIX}/check-schedule", methods=["POST"])
@admin_required
def api_save_check_schedule():
    """업로드 체크 스케줄 설정 저장 + 잡 재등록"""
    data = request.json or {}
    sched = {
        "enabled": bool(data.get("enabled", False)),
        "hour": int(data.get("hour", 9)),
        "minute": int(data.get("minute", 0)),
    }
    save_check_schedule(sched)
    _register_check_schedule_job()
    push_log(f"📅 업로드 체크 자동확인 설정 저장됨: {'활성' if sched['enabled'] else '비활성'} {sched['hour']:02d}:{sched['minute']:02d}")
    return jsonify({"ok": True})


# ── 자동 작업 스케줄 API (수집/체크/콤보) ──────────

@app.route(f"{URL_PREFIX}/task-schedule", methods=["GET"])
@admin_required
def api_get_task_schedule():
    """자동 작업 스케줄 설정 조회"""
    slots = load_task_schedule()
    for slot in slots:
        job_id = f"task_schedule_{slot['id']}"
        job = scheduler.get_job(job_id)
        slot["registered"] = job is not None
        if job and job.next_run_time:
            slot["next_run"] = job.next_run_time.strftime("%Y-%m-%d %H:%M")
        else:
            slot["next_run"] = None
    return jsonify({"ok": True, "slots": slots})


@app.route(f"{URL_PREFIX}/task-schedule", methods=["POST"])
@admin_required
def api_save_task_schedule():
    """자동 작업 스케줄 설정 저장 + 잡 재등록"""
    data = request.json or {}
    slots = data.get("slots", [])
    if not isinstance(slots, list) or len(slots) != 3:
        return jsonify({"ok": False, "error": "3개 슬롯 필요"}), 400

    save_task_schedule(slots)
    _register_task_schedule_jobs()
    push_log("📅 자동 작업 스케줄 설정이 저장되었습니다")
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/upload-status-summary", methods=["GET"])
@admin_required
def api_upload_status_summary():
    """업로드 상태 요약 — 대기/완료/중복/전체 수량 + 예상 시간"""
    products = load_latest_products()
    waiting = 0
    uploaded = 0
    duplicate = 0
    total = len(products)
    for p in products:
        st = (p.get("cafe_status") or "대기")
        if st == "대기":
            waiting += 1
        elif st == "완료":
            uploaded += 1
        elif st == "중복":
            duplicate += 1
    return jsonify({
        "ok": True,
        "waiting_count": waiting,
        "uploaded_count": uploaded,
        "duplicate_count": duplicate,
        "total_count": total,
        "avg_minutes_per_post": 10,
    })


@app.route(f"{URL_PREFIX}/run/upload", methods=["POST"])
@admin_required
def manual_upload():
    """수동 업로드 실행"""
    data = request.json or {}
    max_upload = data.get("max_upload")
    shuffle_brands = data.get("shuffle_brands", False)
    checked_codes = data.get("checked_codes")  # 체크된 상품 코드 배열
    delay_min = data.get("delay_min", 8)
    delay_max = data.get("delay_max", 13)
    source_type = data.get("source_type", "sports")  # vintage 지원
    thread = threading.Thread(
        target=run_upload,
        args=(max_upload, shuffle_brands, checked_codes, delay_min, delay_max, source_type),
        daemon=True,
    )
    thread.start()
    return jsonify({"ok": True, "message": "업로드 시작됨"})


@app.route(f"{URL_PREFIX}/run/test", methods=["POST"])
@admin_required
def run_test():
    """테스트 버튼 핸들러"""
    push_log("🧪 테스트 버튼 클릭됨 — 정상 작동 확인")
    return jsonify({"ok": True, "message": "테스트 성공"})


@app.route(f"{URL_PREFIX}/run/upload-preview", methods=["POST"])
@admin_required
def upload_preview():
    """업로드 전 미리보기 — 번역 결과 포함 리스트 반환"""
    data = request.json or {}
    checked_codes = data.get("checked_codes", [])
    if not checked_codes:
        return jsonify({"ok": False, "items": []})

    from post_generator import make_title, _has_japanese

    products = load_latest_products()
    # DB 상품 병합
    try:
        from product_db import get_unuploaded_products
        db_products = get_unuploaded_products()
        existing_codes = {p.get("product_code", "") for p in products if p.get("product_code")}
        for dp in db_products:
            if dp.get("product_code") and dp["product_code"] not in existing_codes:
                existing_codes.add(dp["product_code"])
                products.append(dp)
    except Exception:
        pass

    checked_set = set(checked_codes)
    items = []
    for p in products:
        code = p.get("product_code", "")
        if code not in checked_set:
            continue
        name = p.get("name_ko") or p.get("name", "")
        brand = p.get("brand_ko") or p.get("brand", "")
        title = make_title(p)
        has_jp = _has_japanese(title)
        items.append({
            "product_code": code,
            "brand": brand,
            "name": name[:50],
            "title": title[:60],
            "has_japanese": has_jp,
        })
    return jsonify({"ok": True, "items": items})


# ── 블로그 업로드 실행/중지 ────────────────
def run_blog_upload(checked_codes=None):
    """블로그 업로드 백그라운드 실행"""
    from blog_uploader import blog_upload_products, request_blog_upload_stop

    products = load_latest_products()
    # DB 상품 병합
    try:
        from product_db import get_unuploaded_products
        db_products = get_unuploaded_products()
        existing_codes = {p.get("product_code", "") for p in products if p.get("product_code")}
        for dp in db_products:
            if dp.get("product_code") and dp["product_code"] not in existing_codes:
                existing_codes.add(dp["product_code"])
                products.append(dp)
    except Exception:
        pass

    checked_set = set(checked_codes) if checked_codes else set()
    checked_set.discard("")
    selected = [p for p in products if p.get("product_code", "") in checked_set]

    if not selected:
        push_log("⚠️ 블로그 업로드할 상품이 없습니다")
        return

    # 블로그 계정 쿠키 경로
    blog_data = _load_blog_accounts()
    active_slot = blog_data.get("active", 1)
    blog_cookie = _get_blog_cookie_path(active_slot)
    active_id = blog_data.get("accounts", {}).get(str(active_slot), {}).get("naver_id", "")
    push_log(f"👤 블로그 계정: {active_id or '미설정'} (슬롯 {active_slot})")
    push_log(f"📝 블로그 업로드 {len(selected)}개 시작")

    try:
        count = asyncio.run(blog_upload_products(
            products=selected,
            status_callback=push_log,
            cookie_path=blog_cookie,
        ))
        push_log(f"🎉 블로그 업로드 완료: {count}개 성공")
    except Exception as e:
        push_log(f"❌ 블로그 업로드 오류: {e}")


@app.route(f"{URL_PREFIX}/run/blog-upload", methods=["POST"])
@admin_required
def manual_blog_upload():
    data = request.json or {}
    checked_codes = data.get("checked_codes")
    thread = threading.Thread(
        target=run_blog_upload,
        args=(checked_codes,),
        daemon=True,
    )
    thread.start()
    return jsonify({"ok": True, "message": "블로그 업로드 시작됨"})


@app.route(f"{URL_PREFIX}/run/blog-upload-stop", methods=["POST"])
@admin_required
def blog_upload_stop():
    from blog_uploader import request_blog_upload_stop
    request_blog_upload_stop()
    push_log("⏹ 블로그 업로드 중지 요청됨")
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/run/upload-stop", methods=["POST"])
@admin_required
def upload_stop():
    """업로드 중지 요청"""
    request_upload_stop()
    push_log("⏹ 업로드 중지 요청됨 — 현재 작업 완료 후 중지됩니다")
    return jsonify({"ok": True, "message": "업로드 중지 요청됨"})


@app.route(f"{URL_PREFIX}/run/upload-reset", methods=["POST"])
@admin_required
def upload_reset():
    """업로드 중지 + 상태 초기화"""
    request_upload_stop()
    status["uploading"] = False
    status["stop_requested"] = False
    # 락 강제 해제
    try:
        _upload_lock.release()
    except RuntimeError:
        pass
    push_log("🔄 업로드 리셋 — 작업 중지 및 상태 초기화 완료")
    return jsonify({"ok": True, "message": "업로드 리셋 완료"})


_upload_check_stop = False

def _run_upload_check(brand_filter=""):
    """백그라운드에서 카페 중복 체크 실행 (Playwright 브라우저 사용)"""
    global _upload_check_stop
    _upload_check_stop = False

    from config import CAFE_MY_NICKNAME

    products = load_latest_products()
    waiting = [p for p in products if (p.get("cafe_status") or "대기") == "대기"]

    # 브랜드 필터 적용
    if brand_filter and brand_filter != "ALL":
        waiting = [p for p in waiting if
                   (p.get("brand_ko") or "").strip() == brand_filter or
                   (p.get("brand") or "").strip() == brand_filter]

    if not waiting:
        push_log("⚠️ 대기 상품이 없습니다" + (f" (브랜드: {brand_filter})" if brand_filter else ""))
        return

    brand_msg = f" [브랜드: {brand_filter}]" if brand_filter and brand_filter != "ALL" else ""
    push_log(f"🔍 업로드 체크 시작: {len(waiting)}개 상품{brand_msg} — 브라우저로 카페 검색 중...")

    from xebio_search import save_products

    def stop_check():
        return _upload_check_stop

    try:
        checked, duplicates = asyncio.run(
            batch_check_cafe_duplicates(
                products=waiting,
                nickname=CAFE_MY_NICKNAME,
                days=30,
                log=push_log,
                save_callback=lambda: save_products(products),
                stop_check=stop_check,
            )
        )

        # 최종 저장
        save_products(products)

        if _upload_check_stop:
            push_log(f"⏹ 체크 중지됨: {checked}개 확인, {duplicates}개 중복 발견")
        else:
            push_log(f"✅ 체크 완료: {checked}개 확인, {duplicates}개 중복 발견")

    except Exception as e:
        push_log(f"❌ 체크 오류: {e}")
    finally:
        _upload_check_stop = False


@app.route(f"{URL_PREFIX}/ai/verify", methods=["POST"])
@admin_required
def ai_verify():
    """AI API 키 정상 작동 여부 확인"""
    result = verify_ai_key()
    return jsonify(result)


@app.route(f"{URL_PREFIX}/run/upload-check", methods=["POST"])
@admin_required
def upload_check():
    """대기 상품을 카페에서 검색하여 중복 여부 체크 (백그라운드)"""
    data = request.json or {}
    brand_filter = data.get("brand", "")
    thread = threading.Thread(target=_run_upload_check, args=(brand_filter,), daemon=True)
    thread.start()
    msg = "카페 중복 체크 시작됨"
    if brand_filter and brand_filter != "ALL":
        msg += f" (브랜드: {brand_filter})"
    return jsonify({"ok": True, "message": msg + " — 로그를 확인하세요"})


@app.route(f"{URL_PREFIX}/run/upload-check-stop", methods=["POST"])
@admin_required
def upload_check_stop():
    """업로드 체크 중지"""
    global _upload_check_stop
    _upload_check_stop = True
    push_log("⏹ 업로드 체크 중지 요청됨")
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/run/auto", methods=["POST"])
@admin_required
def manual_auto():
    """수동으로 자동 파이프라인(스크래핑+업로드) 실행"""
    thread = threading.Thread(target=run_auto_pipeline, daemon=True)
    thread.start()
    return jsonify({"ok": True, "message": "자동 파이프라인 시작됨"})


@app.route(f"{URL_PREFIX}/products/translate", methods=["POST"])
@admin_required
def translate_products():
    """기존 수집 데이터 일괄 번역"""
    try:
        from translator import translate_ja_ko, translate_brand, TRANSLATE_AVAILABLE
        if not TRANSLATE_AVAILABLE:
            return jsonify({"ok": False, "message": "googletrans 미설치 — pip install googletrans==4.0.0-rc1"})

        products = load_latest_products()
        if not products:
            return jsonify({"ok": False, "message": "수집된 상품이 없습니다"})

        push_log(f"🌐 번역 시작: 총 {len(products)}개 상품")
        count = 0
        for p in products:
            changed = False
            # 상품명 번역
            if p.get("name") and not p.get("name_ko"):
                p["name_ko"] = translate_ja_ko(p["name"])
                changed = True
            # 브랜드 번역
            if p.get("brand") and not p.get("brand_ko"):
                p["brand_ko"] = translate_brand(p["brand"])
                changed = True
            # 상세 설명 번역
            if p.get("description") and not p.get("description_ko"):
                p["description_ko"] = translate_ja_ko(p["description"])
                changed = True
            if changed:
                count += 1

        from xebio_search import save_products
        save_products(products)
        msg = f"번역 완료: {count}개 상품"
        push_log(f"✅ " + msg)
        return jsonify({"ok": True, "message": msg, "count": count})

    except Exception as e:
        push_log(f"❌ 번역 오류: {e}")
        return jsonify({"ok": False, "message": str(e)})


@app.route(f"{URL_PREFIX}/products/translate-missing", methods=["POST"])
@admin_required
def translate_missing_fields():
    """DB에 일본어가 남아있는 color/material 필드 일괄 번역"""
    import re as _re, threading
    _ja = _re.compile(r'[\u3040-\u30FF\u4E00-\u9FFF]')

    def _run():
        from product_db import _conn
        from translator import translate_ja_ko
        conn = _conn()
        try:
            rows = conn.execute("SELECT id, color, material FROM products WHERE source_type='vintage'").fetchall()
            updated = 0
            for r in rows:
                changed = False
                color = r["color"] or ""
                material = r["material"] or ""
                if color and _ja.search(color):
                    color = translate_ja_ko(color)
                    changed = True
                if material and _ja.search(material):
                    material = translate_ja_ko(material)
                    changed = True
                if changed:
                    conn.execute("UPDATE products SET color=?, material=? WHERE id=?", (color, material, r["id"]))
                    updated += 1
                    if updated % 50 == 0:
                        conn.commit()
                        push_log(f"🔄 번역 진행중... {updated}건 완료")
            conn.commit()
            push_log(f"✅ 일본어 번역 완료: {updated}건 업데이트")
        except Exception as e:
            push_log(f"❌ 번역 오류: {e}")
        finally:
            conn.close()

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "message": "일본어 번역 시작 (백그라운드)"})


@app.route(f"{URL_PREFIX}/products/rescrape-details", methods=["POST"])
@admin_required
def rescrape_details_api():
    """2ndstreet 설명 없는 상품의 상세 페이지 재수집"""
    from secondst_crawler import is_rescrape_running, rescrape_details
    if is_rescrape_running():
        return jsonify({"ok": False, "message": "이미 재수집 진행 중입니다"})

    def _run():
        import asyncio
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(rescrape_details(log=push_log))
        loop.close()

    import threading
    t = threading.Thread(target=_run, daemon=True)
    t.start()
    push_log("🔄 2ndstreet 상세 페이지 재수집 시작...")
    return jsonify({"ok": True, "message": "상세 페이지 재수집을 시작합니다"})


@app.route(f"{URL_PREFIX}/products/rescrape-details/stop", methods=["POST"])
@admin_required
def rescrape_details_stop():
    """재수집 중지"""
    from secondst_crawler import stop_rescrape, is_rescrape_running
    if not is_rescrape_running():
        return jsonify({"ok": False, "message": "진행 중인 재수집이 없습니다"})
    stop_rescrape()
    push_log("⛔ 재수집 중지 요청")
    return jsonify({"ok": True, "message": "재수집 중지 요청됨"})


@app.route(f"{URL_PREFIX}/settings/dict", methods=["GET"])
@admin_required
def get_dict():
    """커스텀 단어장 조회"""
    from translator import CUSTOM_DICT
    return jsonify({"dict": CUSTOM_DICT})


@app.route(f"{URL_PREFIX}/settings/dict", methods=["POST"])
@admin_required
def update_dict():
    """커스텀 단어장 단어 추가/수정"""
    from translator import CUSTOM_DICT
    data = request.json or {}
    ja = data.get("ja", "").strip()
    ko = data.get("ko", "").strip()
    if not ja or not ko:
        return jsonify({"ok": False, "message": "일본어와 한국어를 모두 입력해주세요"})
    CUSTOM_DICT[ja] = ko
    push_log(f"📖 단어 추가: {ja} → {ko}")
    return jsonify({"ok": True, "message": f"{ja} → {ko} 추가 완료"})


@app.route(f"{URL_PREFIX}/settings/dict/<path:ja>", methods=["DELETE"])
@admin_required
def delete_dict(ja):
    """커스텀 단어장 단어 삭제"""
    from translator import CUSTOM_DICT
    ja = ja.strip()
    if ja in CUSTOM_DICT:
        del CUSTOM_DICT[ja]
        push_log(f"🗑️ 단어 삭제: {ja}")
        return jsonify({"ok": True})
    return jsonify({"ok": False, "message": "단어 없음"})


@app.route(f"{URL_PREFIX}/settings/margin", methods=["POST"])
@admin_required
def update_margin():
    """마진율 변경 (하위 호환)"""
    data = request.json or {}
    pct = data.get("margin_pct", 20)   # 퍼센트로 받기 (예: 20 → 1.2)
    rate = 1 + (pct / 100)
    rate = max(1.0, min(rate, 3.0))    # 0~200% 범위 제한
    set_margin_rate(rate)
    msg = f"마진율 변경: {pct}% (x{round(rate,2)})"
    push_log("💰 " + msg)
    return jsonify({"ok": True, "margin_pct": pct, "margin_rate": round(rate, 2), "message": msg})


@app.route(f"{URL_PREFIX}/settings/price", methods=["GET"])
@admin_required
def get_price_settings():
    """현재 가격 설정 조회"""
    return jsonify({"ok": True, **get_price_config()})


@app.route(f"{URL_PREFIX}/settings/price", methods=["POST"])
@admin_required
def update_price_settings():
    """가격 계산 변수 일괄 변경"""
    data = request.json or {}
    jp_fee   = data.get("jp_fee_pct")       # % 단위 (예: 3 → 0.03)
    markup   = data.get("buy_markup_pct")   # % 단위 (예: 2 → 0.02)
    margin   = data.get("margin_pct")       # % 단위 (예: 10 → 0.10)
    shipping = data.get("intl_shipping_krw")# 원화 (예: 15000)

    set_price_config(
        jp_fee   = jp_fee   / 100 if jp_fee   is not None else None,
        buy_markup = markup / 100 if markup   is not None else None,
        margin   = margin   / 100 if margin   is not None else None,
        shipping = shipping if shipping is not None else None,
    )
    cfg = get_price_config()
    msg = (f"가격설정 변경: 수수료={cfg['jp_fee_pct']}% "
           f"환율추가={cfg['buy_markup_pct']}% "
           f"마진={cfg['margin_pct']}% "
           f"배송={cfg['intl_shipping_krw']:,}원")
    push_log("💰 " + msg)
    return jsonify({"ok": True, **cfg, "message": msg})


# ── 자유게시판 (중고명품 기사) ──────────────────

def _init_free_board_db():
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS free_board (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            url TEXT DEFAULT '',
            image_path TEXT DEFAULT '',
            status TEXT DEFAULT '대기',
            cafe_menu_id TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )""")
        # 기존 테이블에 컬럼 없으면 추가
        cols = {r[1] for r in conn.execute("PRAGMA table_info(free_board)").fetchall()}
        if "image_path" not in cols:
            conn.execute("ALTER TABLE free_board ADD COLUMN image_path TEXT DEFAULT ''")
        if "tags" not in cols:
            conn.execute("ALTER TABLE free_board ADD COLUMN tags TEXT DEFAULT ''")
        if "article_type" not in cols:
            conn.execute("ALTER TABLE free_board ADD COLUMN article_type TEXT DEFAULT ''")
        conn.commit()
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/api/free-board", methods=["GET"])
@admin_required
def get_free_board():
    _init_free_board_db()
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        rows = conn.execute("SELECT * FROM free_board ORDER BY id DESC LIMIT 50").fetchall()
        posts = []
        for r in rows:
            p = {"id": r["id"], "title": r["title"], "content": r["content"],
                 "url": r["url"] or "", "status": r["status"] or "대기",
                 "created_at": r["created_at"] or ""}
            try:
                p["image_path"] = r["image_path"] or ""
            except Exception:
                p["image_path"] = ""
            try:
                p["tags"] = r["tags"] or ""
            except Exception:
                p["tags"] = ""
            try:
                p["article_type"] = r["article_type"] or ""
            except Exception:
                p["article_type"] = ""
            posts.append(p)
        return jsonify({"ok": True, "posts": posts})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/api/free-board", methods=["POST"])
@admin_required
def create_free_board():
    _init_free_board_db()
    data = request.json or {}
    title = (data.get("title") or "").strip()
    content = (data.get("content") or "").strip()
    url = (data.get("url") or "").strip()
    image_path = (data.get("image_path") or "").strip()
    tags = (data.get("tags") or "").strip()
    article_type = (data.get("article_type") or "").strip()
    if not title or not content:
        return jsonify({"ok": False, "message": "제목과 내용을 입력해주세요"})
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("INSERT INTO free_board (title, content, url, image_path, tags, article_type) VALUES (?,?,?,?,?,?)",
                     (title, content, url, image_path, tags, article_type))
        conn.commit()
        return jsonify({"ok": True, "message": "등록 완료"})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/api/free-board/<int:post_id>/status", methods=["POST"])
@admin_required
def update_free_board_status(post_id):
    _init_free_board_db()
    data = request.json or {}
    new_status = data.get("status", "대기")
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("UPDATE free_board SET status=? WHERE id=?", (new_status, post_id))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


_FB_SCHEDULE_PATH = os.path.join(get_path("db"), "fb_schedule.json")

@app.route(f"{URL_PREFIX}/api/free-board/schedule", methods=["GET"])
@admin_required
def get_fb_schedule():
    import json as _json
    data = {"enabled": False, "eco_hour": 8, "brand_hour": 17}
    if os.path.exists(_FB_SCHEDULE_PATH):
        try:
            with open(_FB_SCHEDULE_PATH, "r") as f:
                data.update(_json.load(f))
        except Exception:
            pass
    return jsonify({"ok": True, "schedule": data})


@app.route(f"{URL_PREFIX}/api/free-board/schedule", methods=["POST"])
@admin_required
def save_fb_schedule():
    import json as _json
    data = request.json or {}
    sched = {
        "enabled": bool(data.get("enabled")),
        "eco_hour": int(data.get("eco_hour", 8)),
        "brand_hour": int(data.get("brand_hour", 17)),
    }
    os.makedirs(os.path.dirname(_FB_SCHEDULE_PATH), exist_ok=True)
    with open(_FB_SCHEDULE_PATH, "w") as f:
        _json.dump(sched, f)
    # 스케줄러 등록/해제
    _register_fb_schedule_jobs()
    return jsonify({"ok": True})


def _register_fb_schedule_jobs():
    """자유게시판 자동 기사 생성+업로드 스케줄 등록"""
    import json as _json
    sched = {"enabled": False, "eco_hour": 8, "brand_hour": 17}
    if os.path.exists(_FB_SCHEDULE_PATH):
        try:
            with open(_FB_SCHEDULE_PATH, "r") as f:
                sched.update(_json.load(f))
        except Exception:
            pass

    # 기존 잡 제거
    for jid in ["fb_auto_generate", "fb_auto_upload_eco", "fb_auto_upload_brand"]:
        if scheduler.get_job(jid):
            scheduler.remove_job(jid)

    if sched.get("enabled"):
        # 새벽 6시: 기사 자동 생성 (경제2 + 브랜드2)
        scheduler.add_job(
            func=_fb_auto_generate_articles,
            trigger="cron", hour=6, minute=0,
            id="fb_auto_generate", replace_existing=True,
            name="자유게시판 자동 기사 생성 (06:00)",
        )
        # 경제 기사 업로드 (예약시간 - 컨펌 없으면 자동)
        eco_hour = sched.get("eco_hour", 8)
        scheduler.add_job(
            func=lambda: _fb_auto_upload("economy"),
            trigger="cron", hour=eco_hour, minute=0,
            id="fb_auto_upload_eco", replace_existing=True,
            name=f"자유게시판 경제기사 업로드 ({eco_hour:02d}:00)",
        )
        # 브랜드 기사 업로드
        brand_hour = sched.get("brand_hour", 17)
        scheduler.add_job(
            func=lambda: _fb_auto_upload("brand"),
            trigger="cron", hour=brand_hour, minute=0,
            id="fb_auto_upload_brand", replace_existing=True,
            name=f"자유게시판 브랜드기사 업로드 ({brand_hour:02d}:00)",
        )
        logger.info(f"📅 자유게시판 스케줄 등록: 생성 06:00, 경제 {eco_hour:02d}:00, 브랜드 {brand_hour:02d}:00")


def _fb_auto_generate_articles():
    """새벽 6시: 경제 기사 2개 + 브랜드 기사 2개 자동 생성"""
    push_log("📰 [자동] 자유게시판 기사 생성 시작 (경제2 + 브랜드2)")
    import requests as _req
    base = f"http://localhost:{3002}"
    # 세션 쿠키 없이 내부 호출이므로 직접 함수 호출
    try:
        from post_generator import get_ai_config, _call_gemini, _call_claude, _call_openai
        config = get_ai_config()
        from exchange import get_cached_rate
        rate = get_cached_rate() or 9.23
        import random, json as _json
        from datetime import datetime as _dt
        from user_db import _conn as user_conn

        today = _dt.now().strftime("%Y년 %m월 %d일")
        weekday = ["월","화","수","목","금","토","일"][_dt.now().weekday()]
        provider = config.get("provider", "gemini")

        articles = []
        for atype in ["economy", "economy", "brand", "brand"]:
            try:
                # 간단한 프롬프트로 기사 생성
                if atype == "economy":
                    topics = ["일본환율 변동과 명품 소싱 전략", "글로벌 럭셔리 시장 전망", "엔저 시대 구매대행 기회", "해외직구 관세 변화"]
                    topic = random.choice(topics)
                    prompt = f"명품 구매대행 카페 경제 기사. 주제: {topic}. 오늘: {today}. 엔화: ¥100={rate*100:.0f}원. 500자 내외, 단락별 구조, 원화 표기. JSON: {{\"title\":\"제목\",\"content\":\"본문\",\"keywords\":[\"kw1\"]}}"
                else:
                    brands = ["롤렉스","에르메스","샤넬","루이비통","구찌","프라다"]
                    brand = random.choice(brands)
                    prompt = f"명품 브랜드 기사. 브랜드: {brand}. 오늘: {today}. 500자 내외, 단락별, 원화 표기. JSON: {{\"title\":\"제목\",\"content\":\"본문\",\"keywords\":[\"kw1\"]}}"

                if provider == "gemini":
                    result = _call_gemini(prompt)
                elif provider == "claude":
                    result = _call_claude(prompt)
                else:
                    result = _call_openai(prompt)

                cleaned = result.strip()
                if "```" in cleaned:
                    cleaned = cleaned.split("```json")[-1].split("```")[0].strip() if "```json" in cleaned else cleaned.split("```")[1].split("```")[0].strip()
                parsed = _json.loads(cleaned)
                title = parsed.get("title", "")
                content = parsed.get("content", "")

                # 컨설팅 안내 추가
                content += "\n\n\n🚀 중고명품창업 컨설팅 안내\n\n현지 소싱부터 실무 운영까지, 성공적인 창업을 지원합니다.\n\n👉 [컨설팅 상세 내용 확인하기]\nhttps://cafe.naver.com/sohosupport/2972\n\n━━━━━━━━━━━━━━━━━━━━"

                # DB 저장
                conn = user_conn()
                conn.execute("INSERT INTO free_board (title, content, article_type, status) VALUES (?,?,?,?)",
                             (title, content, atype, "대기"))
                conn.commit()
                conn.close()
                articles.append({"type": atype, "title": title})
                push_log(f"📰 [자동] {atype} 기사 생성: {title[:30]}...")
                import time; time.sleep(3)
            except Exception as e:
                push_log(f"❌ [자동] 기사 생성 실패: {e}")
                logger.warning(f"자동 기사 생성 실패: {e}")

        push_log(f"📰 [자동] 총 {len(articles)}개 기사 생성 완료 — 컨펌 대기 중")
    except Exception as e:
        push_log(f"❌ [자동] 기사 생성 오류: {e}")


def _fb_auto_upload(article_type: str):
    """예약 시간: 승인된 기사 업로드, 미승인이면 자동 업로드"""
    push_log(f"📰 [자동] {article_type} 기사 업로드 확인...")
    try:
        from user_db import _conn as user_conn
        _init_free_board_db()
        conn = user_conn()
        # 승인된 기사 우선
        row = conn.execute(
            "SELECT id, title FROM free_board WHERE article_type=? AND status='승인' ORDER BY id DESC LIMIT 1",
            (article_type,)).fetchone()
        if not row:
            # 승인 없으면 대기 기사 자동 업로드 (1시간 전까지 컨펌 없음)
            row = conn.execute(
                "SELECT id, title FROM free_board WHERE article_type=? AND status='대기' ORDER BY id DESC LIMIT 1",
                (article_type,)).fetchone()
            if row:
                push_log(f"📰 [자동] 컨펌 없음 → 자동 업로드: {row['title'][:30]}...")
        conn.close()

        if row:
            post_id = row["id"]
            # 업로드 실행 (기존 upload_free_board_to_cafe 로직 재사용)
            import threading
            def _do():
                import asyncio
                from cafe_uploader import upload_article_to_cafe
                conn2 = user_conn()
                r = conn2.execute("SELECT * FROM free_board WHERE id=?", (post_id,)).fetchone()
                conn2.close()
                if not r:
                    return
                title = r["title"]
                content = r["content"]
                try:
                    img_path = r["image_path"] or ""
                except Exception:
                    img_path = ""
                try:
                    tags = r["tags"] or ""
                except Exception:
                    tags = ""
                naver_data = _load_naver_accounts()
                active_slot = naver_data.get("active", 1)
                cookie_path = _get_cookie_path(active_slot)
                push_log(f"📰 [자동] 업로드 시작: {title[:30]}...")
                result = asyncio.run(upload_article_to_cafe(
                    title=title, content=content, menu_id="126",
                    board_name="자유게시판", log=push_log, cookie_path=cookie_path,
                    image_path=img_path, tags=tags,
                ))
                conn3 = user_conn()
                conn3.execute("UPDATE free_board SET status=? WHERE id=?",
                              ("완료" if result else "실패", post_id))
                conn3.commit()
                conn3.close()
                push_log(f"📰 [자동] {'완료' if result else '실패'}: {title[:30]}")
            threading.Thread(target=_do, daemon=True).start()
        else:
            push_log(f"📰 [자동] {article_type} 업로드할 기사 없음")
    except Exception as e:
        push_log(f"❌ [자동] 업로드 오류: {e}")


@app.route(f"{URL_PREFIX}/api/free-board/<int:post_id>", methods=["DELETE"])
@admin_required
def delete_free_board(post_id):
    _init_free_board_db()
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        conn.execute("DELETE FROM free_board WHERE id=?", (post_id,))
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route(f"{URL_PREFIX}/api/free-board/generate", methods=["POST"])
@admin_required
def generate_free_board_article():
    """AI로 중고명품 관련 기사 자동 생성"""
    data = request.json or {}
    article_type = data.get("type", "economy")  # economy / brand
    user_keyword = (data.get("keyword") or "").strip()
    image_source = data.get("image_source", "gemini")  # gemini / gemini_edit / pexels / none
    user_image_prompt = (data.get("image_prompt") or "").strip()

    try:
        from post_generator import get_ai_config, _call_gemini, _call_claude, _call_openai
        config = get_ai_config()
        provider = config.get("provider", "none")
        if provider == "none":
            return jsonify({"ok": False, "message": "AI가 설정되지 않았습니다"})

        from exchange import get_cached_rate
        rate = get_cached_rate() or 9.23
        import random
        from datetime import datetime as _dt
        today = _dt.now().strftime("%Y년 %m월 %d일")
        weekday = ["월","화","수","목","금","토","일"][_dt.now().weekday()]

        # 실시간 트렌드 수집 (Google 인기검색어 + 명품 관련 추천)
        trending_keywords = []
        luxury_suggestions = []
        try:
            import urllib.request
            import xml.etree.ElementTree as ET
            # 1) Google 트렌드 실시간 인기검색어
            req = urllib.request.Request(
                "https://trends.google.co.kr/trending/rss?geo=KR",
                headers={"User-Agent": "Mozilla/5.0"}
            )
            rss_data = urllib.request.urlopen(req, timeout=5).read().decode("utf-8")
            root = ET.fromstring(rss_data)
            for item in root.findall(".//item/title"):
                if item.text:
                    trending_keywords.append(item.text.strip())
            trending_keywords = trending_keywords[:15]
        except Exception as e:
            logger.warning(f"트렌드 인기검색어 로드 실패: {e}")

        try:
            # 2) 명품/경제 관련 Google 추천 검색어
            import re as _re2
            search_seeds = ["명품 시세", "중고명품 가격", "일본환율 명품", "미국환율 관세", "명품 구매대행", "빈티지 명품"]
            for seed in search_seeds[:3]:
                encoded = urllib.request.quote(seed)
                url = f"https://www.google.com/complete/search?q={encoded}&client=gws-wiz&hl=ko"
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                resp = urllib.request.urlopen(req, timeout=3).read().decode("utf-8")
                # 추천 검색어 파싱
                matches = _re2.findall(r'\["([^"]+)"', resp)
                for m in matches[:3]:
                    clean = _re2.sub(r'<[^>]+>', '', m).strip()
                    if clean and clean not in luxury_suggestions and len(clean) < 30:
                        luxury_suggestions.append(clean)
            luxury_suggestions = luxury_suggestions[:10]
        except Exception as e:
            logger.warning(f"추천 검색어 로드 실패: {e}")

        all_trend_info = trending_keywords + luxury_suggestions
        logger.info(f"📊 트렌드: 인기 {len(trending_keywords)}개, 명품추천 {len(luxury_suggestions)}개")

        # 실시간 뉴스 수집 (Google News RSS — 사실 기반 기사 작성용)
        news_text = ""
        news_items = []
        try:
            import xml.etree.ElementTree as _ET_news
            # 키워드별 뉴스 검색
            news_queries = [user_keyword] if user_keyword else ["명품 시세", "일본 엔화 환율", "럭셔리 브랜드"]
            for nq in news_queries[:2]:
                try:
                    news_url = f"https://news.google.com/rss/search?q={_req_lib.utils.quote(nq)}&hl=ko&gl=KR&ceid=KR:ko"
                    nr = requests.get(news_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=5)
                    if nr.status_code == 200:
                        nroot = _ET_news.fromstring(nr.text)
                        for nitem in nroot.findall(".//item")[:5]:
                            ntitle = nitem.find("title").text if nitem.find("title") is not None else ""
                            npub = nitem.find("pubDate").text[:16] if nitem.find("pubDate") is not None else ""
                            if ntitle and ntitle not in [n["title"] for n in news_items]:
                                news_items.append({"title": ntitle, "date": npub})
                except Exception:
                    pass
            if news_items:
                news_lines = [f"- [{n['date']}] {n['title']}" for n in news_items[:10]]
                news_text = "\n\n[사실 확인용 최신 뉴스 — 반드시 이 뉴스를 참고하여 사실 기반으로 작성]\n" + "\n".join(news_lines)
                logger.info(f"📰 뉴스 {len(news_items)}건 수집: {', '.join([n['title'][:20] for n in news_items[:3]])}...")
        except Exception as e:
            logger.warning(f"뉴스 수집 실패: {e}")

        trending_text = ""
        parts = []
        if trending_keywords:
            parts.append(f"실시간 인기검색어: {', '.join(trending_keywords[:10])}")
        if luxury_suggestions:
            parts.append(f"명품 관련 추천검색어: {', '.join(luxury_suggestions[:8])}")
        if parts:
            trending_text = "\n\n[참고: 오늘의 검색 트렌드]\n" + "\n".join(parts)
        trending_text += "\n환율 언급 시 반드시 '일본환율' 또는 '미국환율'로만 표기하세요."
        trending_text += news_text

        if article_type == "economy":
            topics = [
                # 글로벌 경제/투자
                "글로벌 럭셔리 시장 분기별 실적과 투자 전망",
                "LVMH, 케어링, 에르메스 등 럭셔리 그룹 주가 동향",
                "명품을 재테크 수단으로? 리셀 시장의 투자 가치 분석",
                "Z세대가 바꾸는 명품 소비 트렌드",
                "온라인 명품 플랫폼(머스트잇, 트렌비, 발란) 시장 경쟁 현황",
                # 창업/사업
                "명품 구매대행 창업 가이드 — 초기 자금부터 첫 매출까지",
                "1인 명품 셀러로 월 500만원 수익 구조 만들기",
                "명품 위탁판매 vs 직접소싱, 어떤 모델이 맞을까?",
                "네이버 스마트스토어 명품 판매 노하우와 주의점",
                "인스타그램/유튜브 명품 마케팅 성공 사례",
                # 트렌드/문화
                "파리/밀라노 패션위크 하이라이트와 다음 시즌 트렌드",
                "지속가능한 패션 — 명품 브랜드의 친환경 전략",
                "빈티지 명품이 뜨는 이유 — MZ세대의 가치 소비",
                "일본 중고 명품 시장이 세계에서 주목받는 이유",
                "명품 감정사가 알려주는 진품 구별법 TOP 5",
                # 실무/관세
                "해외 직구 vs 구매대행, 관부가세 절감 전략 비교",
                "사업자 통관 vs 개인통관, 어떤 게 유리할까?",
                "일본/미국/유럽 구매대행 루트별 장단점 비교",
                # 환율 (비중 낮춤)
                "이번 주 일본환율/미국환율 변동과 소싱 타이밍",
            ]
            topic = user_keyword if user_keyword else random.choice(topics)
            prompt = f"""당신은 럭셔리 비즈니스 & 라이프스타일 매거진 에디터입니다.
오늘 날짜: {today} ({weekday}요일)
참고 환율: ¥100 = {rate*100:.0f}원

주제: {topic}
※ 환율 이야기가 주제가 아니면 환율은 언급하지 마세요.

[작성 규칙]
1. 제목: 흥미로운 헤드라인 (20~35자)
2. 본문 구조를 반드시 아래 5개 단락으로 나누어 작성:

   📌 [핵심 요약] (2~3줄 요약)
   --- 여기에 이미지 1 ---

   📊 [시장 분석] (현재 상황, 데이터 기반 분석)
   --- 여기에 이미지 2 ---

   💡 [실전 전략] (구매대행 사업자/직구족을 위한 구체적 팁 3가지, 불릿 포인트)
   --- 여기에 이미지 3 ---

   📈 [전망 & 인사이트] (앞으로의 전망, 주의점)

   💬 [참여 질문] (회원 참여 유도 질문 1개)

   ✅ 실무 팁: (한 줄 팁)

3. 각 단락은 "--- 여기에 이미지 N ---" 줄로 구분 (이미지 삽입 위치 표시)
4. 환율은 반드시 '일본환율' 또는 '미국환율'로 표기
5. 서술식 장문 금지, 짧은 문장 + 불릿 포인트 위주
6. 이모지는 섹션 제목에만 최소한으로 사용 (본문 내 이모지 남용 금지)
7. 대표키워드 5~7개 선정 (제목/본문에 자연스럽게 포함)
8. 모든 가격/금액은 반드시 한국 원화(원)로 표기 (달러($) 사용 금지)
9. 반드시 사실 기반 정보만 작성 — 아래 [최신 뉴스]를 참고하여 실제 있었던 사건/수치만 언급
10. 셀럽/인물 언급 시 실제 확인된 뉴스가 있는 경우만 (지어내기 절대 금지)
11. 확인되지 않은 수치는 "약", "추정" 등을 붙여 구분
12. 가격/시세 수치를 넣을 때는 출처나 시점을 명시 (예: "2026년 4월 기준 약 150만원")

반드시 아래 JSON 형식으로만 응답:
{{"title":"기사 제목","content":"기사 본문","keywords":["키워드1","키워드2","키워드3","키워드4","키워드5"]}}{trending_text}"""

        else:  # brand
            brands = [
                ("롤렉스", "Rolex", "시계"),
                ("에르메스", "Hermès", "가방/소품"),
                ("샤넬", "CHANEL", "가방/의류"),
                ("루이비통", "Louis Vuitton", "가방/소품"),
                ("구찌", "GUCCI", "가방/의류"),
                ("프라다", "PRADA", "가방/의류"),
                ("디올", "Dior", "가방/의류"),
                ("보테가 베네타", "Bottega Veneta", "가방"),
                ("셀린느", "CELINE", "가방"),
                ("로에베", "LOEWE", "가방"),
            ]
            brand_ko, brand_en, category = random.choice(brands)
            keyword_line = f"\n특히 다음 키워드를 중심으로 작성: {user_keyword}" if user_keyword else ""

            # 브랜드 관련 최신 뉴스 추가 수집
            try:
                import xml.etree.ElementTree as _ET_b
                brand_news_url = f"https://news.google.com/rss/search?q={_req_lib.utils.quote(brand_ko + ' 2026')}&hl=ko&gl=KR&ceid=KR:ko"
                bnr = requests.get(brand_news_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=5)
                if bnr.status_code == 200:
                    bnroot = _ET_b.fromstring(bnr.text)
                    for bni in bnroot.findall(".//item")[:5]:
                        bnt = bni.find("title").text if bni.find("title") is not None else ""
                        bnp = bni.find("pubDate").text[:16] if bni.find("pubDate") is not None else ""
                        if bnt and bnt not in [n["title"] for n in news_items]:
                            news_items.append({"title": bnt, "date": bnp})
                    # trending_text 갱신
                    if news_items:
                        news_lines = [f"- [{n['date']}] {n['title']}" for n in news_items[:12]]
                        news_text = "\n\n[사실 확인용 최신 뉴스 — 반드시 참고]\n" + "\n".join(news_lines)
                        trending_text = trending_text.split("[사실 확인용")[0] + news_text
                        logger.info(f"📰 {brand_ko} 뉴스 추가 수집")
            except Exception:
                pass

            prompt = f"""당신은 럭셔리 패션 매거진 에디터입니다.
오늘 날짜: {today} ({weekday}요일)
브랜드: {brand_ko} ({brand_en}) - {category}{keyword_line}

[핵심 방향]
- 이 브랜드의 최신 동향, 트렌드, 뉴스를 중심으로 작성
- 중고명품 얘기는 전체의 20~30%만. 나머지는 브랜드 자체의 이야기 (신제품, 컬렉션, 디자이너, 패션쇼, 셀럽 착용, 글로벌 이슈 등)
- 억지로 중고/구매대행과 연결하지 마세요

[작성 규칙]
1. 제목: "[{brand_ko}] ..." 형식, 매거진 스타일 (20~35자)
2. 본문 구조를 반드시 아래 단락으로 작성:

   🏷 [{brand_ko} 최신 뉴스] (최근 컬렉션/디자이너/캠페인/셀럽 이슈 등 브랜드 자체 소식)
   --- 여기에 이미지 1 ---

   🔥 [주목할 아이템] (이번 시즌 인기 아이템 3가지, 리스트 형식)
   --- 여기에 이미지 2 ---

   📊 [가격 동향] (신품 가격 변동, 중고 시세는 간단히 참고 수준만)
   --- 여기에 이미지 3 ---

   💡 [스타일링 & 팁] (착용법, 관리법, 또는 구매 시 체크 포인트)

   💬 [참여 질문] (회원 참여 유도 질문 1개)

3. 각 단락은 "--- 여기에 이미지 N ---" 줄로 구분
4. 서술식 장문 금지, 짧은 문장 + 불릿 포인트 위주
5. 이모지는 섹션 제목에만 최소한으로 사용
6. 대표키워드 5~7개 선정 (예: {brand_ko}신상, {brand_ko}트렌드, {brand_ko}컬렉션)
7. 모든 가격은 원화(원)로 표기
8. 반드시 사실 기반 정보만 — 아래 [최신 뉴스]를 참고하여 실제 사건/수치만 언급
9. 셀럽/인물 언급 시 실제 확인된 뉴스가 있는 경우만 (지어내기 절대 금지)
10. 확인 안 된 수치는 "약", "추정" 표기, 가격은 시점 명시

반드시 아래 JSON 형식으로만 응답:
{{"title":"기사 제목","content":"기사 본문","keywords":["키워드1","키워드2","키워드3","키워드4","키워드5"]}}{trending_text}"""

        # AI 호출
        if provider == "gemini" and config.get("gemini_key"):
            result = _call_gemini(prompt)
        elif provider == "claude" and config.get("claude_key"):
            result = _call_claude(prompt)
        elif provider == "openai" and config.get("openai_key"):
            result = _call_openai(prompt)
        else:
            return jsonify({"ok": False, "message": "AI API 키가 설정되지 않았습니다"})

        # JSON 파싱
        import json as _json
        cleaned = result.strip()
        if "```" in cleaned:
            cleaned = cleaned.split("```json")[-1].split("```")[0].strip() if "```json" in cleaned else cleaned.split("```")[1].split("```")[0].strip()
        try:
            parsed = _json.loads(cleaned)
        except Exception:
            parsed = {"title": f"[{today}] 럭셔리 경제 브리핑", "content": cleaned[:800]}

        content = parsed.get("content", "")

        # 하단 컨설팅 안내 삽입
        content += """


🚀 중고명품창업 컨설팅 안내

현지 소싱부터 실무 운영까지, 성공적인 창업을 지원합니다.

👉 [컨설팅 상세 내용 확인하기]
https://cafe.naver.com/sohosupport/2972

━━━━━━━━━━━━━━━━━━━━"""

        keywords = parsed.get("keywords", [])
        # 태그 10개 채우기: AI 키워드 + 자동 보충
        base_tags = ["일본구매대행", "명품구매대행", "중고명품", "일본직구", "빈티지명품"]
        for kw in keywords:
            if kw not in base_tags:
                base_tags.insert(0, kw)
        tags = list(dict.fromkeys(base_tags))[:10]  # 중복 제거, 최대 10개

        # ── 이미지 수집 (소스 선택) ──
        image_path = ""
        PEXELS_API_KEY = "ZMFMszrhmZ9oy5UTEC0XKa7h8JGytGpnLWkoFDcE4bdqxLv7r507JHEe"
        UNSPLASH_ACCESS_KEY = ""  # 미설정 시 Pexels 폴백

        # 기사 유형별 검색어
        if article_type == "economy":
            search_queries = ["일본 엔화 환율", "도쿄 명품거리 긴자", "명품 구매대행 시장", "일본 환전소", "명품 쇼핑백"]
        else:
            brand_ko_map = {"롤렉스":"롤렉스 시계 2026","에르메스":"에르메스 가방 최신","샤넬":"샤넬 2026 컬렉션",
                "루이비통":"루이비통 최신 컬렉션","구찌":"구찌 2026 컬렉션","프라다":"프라다 최신 가방",
                "디올":"디올 2026 컬렉션","보테가":"보테가 베네타 최신","셀린느":"셀린느 최신 가방","로에베":"로에베 2026 컬렉션"}
            brand_query = "명품 브랜드 최신 컬렉션"
            for bk, bq in brand_ko_map.items():
                if bk in (parsed.get("title","") + content):
                    brand_query = bq
                    break
            search_queries = [brand_query, f"{brand_query} 패션쇼", f"{brand_query} 셀럽", f"{brand_query} 매장", f"{brand_query} 신상"]

        def _overlay_text_on_image(img_bytes, caption_text):
            """PIL로 이미지 위에 정확한 한국어 텍스트 합성 (절대 깨지지 않음)"""
            if not caption_text:
                return img_bytes
            try:
                from PIL import Image, ImageDraw, ImageFont
                import io
                img = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
                w, h = img.size
                bar_h = int(h * 0.16)
                overlay = Image.new("RGBA", (w, h), (0, 0, 0, 0))
                d = ImageDraw.Draw(overlay)
                d.rectangle([(0, h - bar_h), (w, h)], fill=(0, 0, 0, 170))
                img = Image.alpha_composite(img, overlay).convert("RGB")
                draw = ImageDraw.Draw(img)
                font_path = "/System/Library/Fonts/AppleSDGothicNeo.ttc"
                if not os.path.exists(font_path):
                    font_path = "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"
                font_size = max(int(w * 0.038), 18)
                font = ImageFont.truetype(font_path, font_size)
                text_y = h - bar_h + int(bar_h * 0.3)
                draw.text((int(w * 0.04), text_y), caption_text, fill="white", font=font)
                buf = io.BytesIO()
                img.save(buf, format="PNG", quality=92)
                return buf.getvalue()
            except Exception as te:
                logger.warning(f"텍스트 합성 실패: {te}")
                return img_bytes

        # 모든 이미지 모드 공통: 캡션 추출
        _gem_captions = []
        if image_source != "none":
            try:
                hl_prompt = f"""아래 기사에서 이미지 캡션용 핵심 문장을 정확히 5개 뽑아주세요.
각 문장은 10~15자 이내의 정확한 한국어.
반드시 JSON 배열로만 응답: ["문장1","문장2","문장3","문장4","문장5"]

제목: {parsed.get('title','')}
본문: {content[:800]}"""
                if provider == "gemini":
                    hl_r = _call_gemini(hl_prompt)
                elif provider == "claude":
                    hl_r = _call_claude(hl_prompt)
                else:
                    hl_r = _call_openai(hl_prompt)
                hl_c = hl_r.strip()
                if "```" in hl_c:
                    hl_c = hl_c.split("```json")[-1].split("```")[0].strip() if "```json" in hl_c else hl_c.split("```")[1].split("```")[0].strip()
                _gem_captions = _json.loads(hl_c)
                logger.info(f"🖼 캡션 추출: {_gem_captions}")
            except Exception as ce:
                logger.warning(f"캡션 추출 실패: {ce}")
                _gem_captions = []

        if image_source != "none":
            try:
                import requests as _req_lib
                img_dir = os.path.join(get_path("db"), "article_images")
                os.makedirs(img_dir, exist_ok=True)
                image_paths = []

                _image_sources = []  # 이미지 출처 URL 저장

                def _search_images(query, count=8):
                    """DuckDuckGo 이미지 검색 — 고화질, 다양한 소스, 매번 새로운 이미지"""
                    import re as _re_g
                    try:
                        headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
                        # 1단계: 검색 토큰 획득
                        token_url = f"https://duckduckgo.com/?q={_req_lib.utils.quote(query)}&iax=images&ia=images"
                        r = _req_lib.get(token_url, headers=headers, timeout=10)
                        vqd = _re_g.search(r'vqd=\"([^\"]+)\"', r.text)
                        if not vqd:
                            vqd = _re_g.search(r'vqd=([^&\"]+)', r.text)
                        if not vqd:
                            return []
                        # 2단계: 이미지 결과 가져오기 (최대 100개)
                        img_url = f"https://duckduckgo.com/i.js?l=ko-kr&o=json&q={_req_lib.utils.quote(query)}&vqd={vqd.group(1)}&f=,,,,,&p=1"
                        r2 = _req_lib.get(img_url, headers=headers, timeout=10)
                        results = r2.json().get("results", [])
                        # 큰 이미지만 필터 (최소 400px)
                        valid = [r for r in results if r.get("width", 0) >= 400 and r.get("height", 0) >= 400]
                        if not valid:
                            valid = results
                        # 랜덤 셔플로 매번 다른 이미지
                        random.shuffle(valid)
                        # 도메인 다양성 확보
                        seen_domains = set()
                        selected = []
                        for r in valid:
                            img = r.get("image", "")
                            try:
                                domain = img.split("/")[2]
                            except Exception:
                                domain = ""
                            if domain not in seen_domains and img:
                                selected.append({"url": img, "source": r.get("source", ""), "domain": domain})
                                seen_domains.add(domain)
                            if len(selected) >= count:
                                break
                        # 도메인 다양성으로 부족하면 나머지 추가
                        if len(selected) < count:
                            for r in valid:
                                img = r.get("image", "")
                                if img and not any(s["url"] == img for s in selected):
                                    try:
                                        domain = img.split("/")[2]
                                    except Exception:
                                        domain = ""
                                    selected.append({"url": img, "source": r.get("source", ""), "domain": domain})
                                if len(selected) >= count:
                                    break
                        return selected
                    except Exception as e:
                        logger.warning(f"이미지 검색 실패: {e}")
                        return []

                def _get_gemini_key():
                    gk = config.get("gemini_key", "")
                    if not gk or len(gk) <= 20:
                        try:
                            env_path = os.path.join(os.path.dirname(__file__), ".env")
                            with open(env_path, encoding="utf-8") as _ef:
                                for _el in _ef:
                                    if _el.strip().startswith("GEMINI_API_KEY="):
                                        gk = _el.strip().split("=", 1)[1].strip()
                                        break
                        except Exception:
                            pass
                    return gk

                # 구글검색+AI편집: 구글 이미지 검색 → Gemini 편집
                if image_source == "google_edit":
                    try:
                        gemini_key = _get_gemini_key()
                        from google import genai
                        from google.genai import types as _gtypes
                        gclient = genai.Client(api_key=gemini_key)

                        import re as _re_img
                        sections = _re_img.split(r'-*\s*여기에 이미지\s*\d+\s*-*', content)

                        # 모든 검색어 합쳐서 이미지 풀 확보 (매번 새로운 이미지)
                        all_search_imgs = []
                        for sq in search_queries:
                            if user_image_prompt:
                                sq = user_image_prompt
                            results = _search_images(sq, 5)
                            all_search_imgs.extend(results)
                            if user_image_prompt:
                                break
                        random.shuffle(all_search_imgs)
                        img_pool_idx = 0

                        for idx in range(min(len(sections)-1, 7)):
                            try:
                                if img_pool_idx >= len(all_search_imgs):
                                    break

                                # 이미지 다운로드
                                photo_data = None
                                while img_pool_idx < len(all_search_imgs):
                                    img_info = all_search_imgs[img_pool_idx]
                                    img_pool_idx += 1
                                    try:
                                        resp = _req_lib.get(img_info["url"], headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
                                        if resp.status_code == 200 and len(resp.content) > 5000:
                                            photo_data = resp.content
                                            _image_sources.append(img_info["url"])
                                            break
                                    except Exception:
                                        continue
                                if not photo_data:
                                    continue

                                # Gemini 편집 (텍스트 없이)
                                img_part = _gtypes.Part.from_bytes(data=photo_data, mime_type="image/jpeg")
                                if user_image_prompt:
                                    edit_p = f"이 사진을 기반으로: {user_image_prompt}\n규칙: 원본 현실감 유지. IMPORTANT: Do NOT include ANY text in the image."
                                else:
                                    edit_p = f"이 사진을 고급 매거진 스타일로 보정. 색감과 조명만 개선. IMPORTANT: Do NOT include ANY text in the image."
                                resp = gclient.models.generate_content(
                                    model="gemini-2.5-flash-image",
                                    contents=[edit_p, img_part],
                                    config=_gtypes.GenerateContentConfig(response_modalities=["TEXT", "IMAGE"]),
                                )
                                for part in resp.candidates[0].content.parts:
                                    if part.inline_data:
                                        img_data = part.inline_data.data
                                        caption = _gem_captions[idx] if idx < len(_gem_captions) else ""
                                        img_data = _overlay_text_on_image(img_data, caption)
                                        img_filename = f"article_{_dt.now().strftime('%Y%m%d_%H%M%S')}_{idx+1}.png"
                                        img_path = os.path.join(img_dir, img_filename)
                                        with open(img_path, "wb") as out:
                                            out.write(img_data)
                                        image_paths.append(img_path)
                                        logger.info(f"🖼 구글+AI편집 이미지 {idx+1} 완료")
                                        break
                            except Exception as ge:
                                logger.warning(f"🖼 구글+AI편집 이미지 {idx+1} 실패: {ge}")
                            import time; time.sleep(2)
                        logger.info(f"🖼 구글+AI편집 이미지 {len(image_paths)}장 완료")
                    except Exception as e:
                        logger.warning(f"🖼 구글+AI편집 실패: {e}")

                # 실사+AI편집: Pexels 사진 → Gemini 텍스트 합성
                elif image_source == "gemini_edit":
                    try:
                        gemini_key = _get_gemini_key()
                        from google import genai
                        from google.genai import types as _gtypes
                        gclient = genai.Client(api_key=gemini_key)

                        import re as _re_img
                        sections = _re_img.split(r'-*\s*여기에 이미지\s*\d+\s*-*', content)

                        headline_lines = _gem_captions  # 공통 캡션 사용

                        for idx in range(min(len(sections)-1, 5)):
                            try:
                                # Pexels에서 실사 사진 검색
                                sq = search_queries[idx] if idx < len(search_queries) else search_queries[0]
                                r = _req_lib.get("https://api.pexels.com/v1/search",
                                    params={"query": sq, "per_page": 3, "orientation": "landscape"},
                                    headers={"Authorization": PEXELS_API_KEY}, timeout=10)
                                if r.status_code != 200:
                                    continue
                                photos = r.json().get("photos", [])
                                if not photos:
                                    continue
                                photo_url = random.choice(photos)["src"]["large"]
                                photo_data = _req_lib.get(photo_url, timeout=10).content

                                # 캡션 텍스트 (미리 추출한 것 사용)
                                caption = headline_lines[idx] if idx < len(headline_lines) else ""

                                # Gemini에 사진 + 정확한 텍스트 전달
                                img_part = _gtypes.Part.from_bytes(data=photo_data, mime_type="image/jpeg")
                                if user_image_prompt:
                                    edit_prompt = f"이 실제 사진을 기반으로: {user_image_prompt}\n규칙: 원본 사진의 현실감 유지. IMPORTANT: Do NOT include ANY text, letters, words, numbers, or typography in the image. The image must contain ZERO text. Only visual elements."
                                else:
                                    edit_prompt = f"이 실제 사진의 분위기를 더 고급스럽고 매거진 느낌으로 보정해주세요.\n규칙: 원본 사진 최대한 유지. 색감/조명만 살짝 보정. IMPORTANT: Do NOT include ANY text, letters, words, numbers, or typography in the image. The image must contain ZERO text. Only visual elements."
                                resp = gclient.models.generate_content(
                                    model="gemini-2.5-flash-image",
                                    contents=[edit_prompt, img_part],
                                    config=_gtypes.GenerateContentConfig(response_modalities=["TEXT", "IMAGE"]),
                                )
                                for part in resp.candidates[0].content.parts:
                                    if part.inline_data:
                                        img_data = part.inline_data.data
                                        # PIL로 캡션 합성
                                        img_data = _overlay_text_on_image(img_data, caption)
                                        img_filename = f"article_{_dt.now().strftime('%Y%m%d_%H%M%S')}_{idx+1}.png"
                                        img_path = os.path.join(img_dir, img_filename)
                                        with open(img_path, "wb") as out:
                                            out.write(img_data)
                                        image_paths.append(img_path)
                                        logger.info(f"🖼 실사+편집 이미지 {idx+1} 완료")
                                        break
                            except Exception as ge:
                                logger.warning(f"🖼 실사+편집 이미지 {idx+1} 실패: {ge}")
                            import time; time.sleep(2)
                        logger.info(f"🖼 실사+AI편집 이미지 {len(image_paths)}장 완료")
                    except Exception as e:
                        logger.warning(f"🖼 실사+편집 실패: {e}")

                # Gemini AI 순수 생성
                elif image_source == "gemini":
                    try:
                        gemini_key = _get_gemini_key()
                        from google import genai
                        from google.genai import types as _gtypes
                        gclient = genai.Client(api_key=gemini_key)

                        # 본문에서 섹션 제목 추출하여 각 섹션에 맞는 이미지 프롬프트 생성
                        import re as _re_img
                        sections = _re_img.split(r'-*\s*여기에 이미지\s*\d+\s*-*', content)

                        img_prompts = []
                        for i, sec in enumerate(sections[:-1]):
                            if user_image_prompt:
                                img_prompts.append(
                                    f"{user_image_prompt}\n\n"
                                    f"규칙: 현실적인 사진. 실제 촬영한 것처럼. AI 느낌 금지. IMPORTANT: Do NOT include ANY text, letters, words, numbers, or typography in the image. The image must contain ZERO text. Only visual elements."
                                )
                            else:
                                sec_text = sec.strip()[-150:] if len(sec.strip()) > 150 else sec.strip()
                                img_prompts.append(
                                    f"다음 내용을 시각적으로 표현하는 고품질 사진:\n{sec_text}\n\n"
                                    f"규칙: 현실적인 사진. 실제 촬영한 것처럼. AI 느낌 금지. 고급 브랜드 부티크 분위기, 따뜻한 조명. IMPORTANT: Do NOT include ANY text, letters, words, numbers, or typography in the image. The image must contain ZERO text. Only visual elements."
                                )
                        if len(img_prompts) < 3:
                            title_for_img = parsed.get("title", "luxury brand")
                            default_prompts = [
                                f"'{title_for_img}' 기사에 어울리는 고급 부티크 매장 내부. 따뜻한 조명, 프리미엄 가방 진열. 현실적 사진. 텍스트 넣지 마세요.",
                                f"명품 가죽 제품 클로즈업. 대리석 위 고급 가방, 골든 라이트. 현실적 제품 사진. 텍스트 넣지 마세요.",
                                f"도쿄 긴자 명품 거리 브랜드 매장. 저녁, 따뜻한 쇼윈도. 현실적 스트리트 사진. 텍스트 넣지 마세요.",
                            ]
                            img_prompts.extend(default_prompts[:3-len(img_prompts)])

                        for idx, ip in enumerate(img_prompts[:5]):
                            try:
                                resp = gclient.models.generate_content(
                                    model="gemini-2.5-flash-image",
                                    contents=ip,
                                    config=_gtypes.GenerateContentConfig(response_modalities=["TEXT", "IMAGE"]),
                                )
                                for part in resp.candidates[0].content.parts:
                                    if part.inline_data:
                                        img_data = part.inline_data.data
                                        # PIL로 정확한 한국어 캡션 합성
                                        caption = _gem_captions[idx] if idx < len(_gem_captions) else ""
                                        img_data = _overlay_text_on_image(img_data, caption)
                                        img_filename = f"article_{_dt.now().strftime('%Y%m%d_%H%M%S')}_{idx+1}.png"
                                        img_path = os.path.join(img_dir, img_filename)
                                        with open(img_path, "wb") as out:
                                            out.write(img_data)
                                        image_paths.append(img_path)
                                        logger.info(f"🖼 Gemini 이미지 {idx+1} 생성+텍스트 합성 완료")
                                        break
                            except Exception as ge:
                                logger.warning(f"🖼 Gemini 이미지 {idx+1} 실패: {ge}")
                            import time; time.sleep(2)
                        logger.info(f"🖼 Gemini AI 이미지 {len(image_paths)}장 완료 (PIL 텍스트 합성)")
                    except Exception as e:
                        logger.warning(f"🖼 Gemini 이미지 생성 실패: {e}")

                # Pexels / Unsplash
                if image_source in ("pexels", "unsplash") or (image_source == "gemini" and not image_paths):
                    if image_source == "gemini" and not image_paths:
                        logger.info("🖼 Gemini 실패 → Pexels 폴백")

                for sq in search_queries:
                    if len(image_paths) >= 5:
                        break
                    if image_source == "gemini":
                        break  # Gemini는 위에서 이미 처리
                    img_urls = []
                    try:
                        if image_source == "pexels":
                            r = _req_lib.get("https://api.pexels.com/v1/search",
                                params={"query": sq, "per_page": 1, "orientation": "landscape", "size": "medium"},
                                headers={"Authorization": PEXELS_API_KEY}, timeout=10)
                            if r.status_code == 200:
                                img_urls = [p["src"]["large"] for p in r.json().get("photos", [])]

                        elif image_source == "unsplash":
                            if not UNSPLASH_ACCESS_KEY:
                                # Unsplash 키 없으면 Pexels 폴백
                                r = _req_lib.get("https://api.pexels.com/v1/search",
                                    params={"query": sq, "per_page": 1, "orientation": "landscape"},
                                    headers={"Authorization": PEXELS_API_KEY}, timeout=10)
                                if r.status_code == 200:
                                    img_urls = [p["src"]["large"] for p in r.json().get("photos", [])]
                                logger.info(f"🖼 Unsplash 키 미설정 → Pexels 폴백: {sq}")
                            else:
                                r = _req_lib.get("https://api.unsplash.com/search/photos",
                                    params={"query": sq, "per_page": 1, "orientation": "landscape"},
                                    headers={"Authorization": f"Client-ID {UNSPLASH_ACCESS_KEY}"}, timeout=10)
                                if r.status_code == 200:
                                    img_urls = [p["urls"]["regular"] for p in r.json().get("results", [])]

                        elif image_source == "dalle":
                            oai_key = config.get("openai_key", "")
                            if not oai_key or len(oai_key) <= 20:
                                try:
                                    env_path = os.path.join(os.path.dirname(__file__), ".env")
                                    with open(env_path, encoding="utf-8") as _ef:
                                        for _el in _ef:
                                            if _el.strip().startswith("OPENAI_API_KEY="):
                                                oai_key = _el.strip().split("=", 1)[1].strip()
                                                break
                                except Exception:
                                    pass
                            if oai_key and len(oai_key) > 20:
                                from openai import OpenAI as _OAI
                                oai = _OAI(api_key=oai_key, timeout=30)
                                dalle_prompt = user_image_prompt if user_image_prompt else sq
                                img_resp = oai.images.generate(
                                    model="dall-e-3",
                                    prompt=f"Professional editorial magazine photo: {dalle_prompt}. Elegant luxury fashion photography, warm lighting, photorealistic. Do NOT include any text, letters, words, or typography in the image.",
                                    size="1024x1024", quality="standard", n=1,
                                )
                                img_urls = [img_resp.data[0].url]

                    except Exception as ie:
                        logger.warning(f"🖼 {image_source} 이미지 실패 ({sq}): {ie}")

                    # 다운로드 + PIL 텍스트 합성
                    for img_url in img_urls:
                        try:
                            img_resp = _req_lib.get(img_url, timeout=15)
                            img_data = img_resp.content
                            # PIL로 캡션 합성
                            cap_idx = len(image_paths)
                            caption = _gem_captions[cap_idx] if cap_idx < len(_gem_captions) else ""
                            img_data = _overlay_text_on_image(img_data, caption)
                            img_filename = f"article_{_dt.now().strftime('%Y%m%d_%H%M%S')}_{cap_idx+1}.png"
                            img_path = os.path.join(img_dir, img_filename)
                            with open(img_path, "wb") as out:
                                out.write(img_data)
                            image_paths.append(img_path)
                            logger.info(f"🖼 [{image_source}] 이미지 {len(image_paths)}/5: {sq}")
                        except Exception as de:
                            logger.warning(f"🖼 다운로드 실패: {de}")

                # Gemini 실패 시 Pexels 폴백 (5개 미만이면)
                PEXELS_API_KEY = "ZMFMszrhmZ9oy5UTEC0XKa7h8JGytGpnLWkoFDcE4bdqxLv7r507JHEe"
                if len(image_paths) < 5 and image_source != "pexels":
                    need = 5 - len(image_paths)
                    push_log(f"🖼 이미지 {len(image_paths)}장 → Pexels 폴백으로 {need}장 추가")
                    for sq in search_queries:
                        if len(image_paths) >= 5:
                            break
                        try:
                            pr = _req_lib.get("https://api.pexels.com/v1/search",
                                params={"query": sq, "per_page": 2, "orientation": "landscape"},
                                headers={"Authorization": PEXELS_API_KEY}, timeout=10)
                            if pr.status_code == 200:
                                for pp in pr.json().get("photos", []):
                                    if len(image_paths) >= 5:
                                        break
                                    try:
                                        img_resp = _req_lib.get(pp["src"]["large"], timeout=10)
                                        img_data = img_resp.content
                                        cap_idx = len(image_paths)
                                        caption = _gem_captions[cap_idx] if cap_idx < len(_gem_captions) else ""
                                        img_data = _overlay_text_on_image(img_data, caption)
                                        from datetime import datetime as _dt2
                                        fn = f"article_{_dt2.now().strftime('%Y%m%d_%H%M%S')}_{cap_idx+1}.png"
                                        fp = os.path.join(img_dir, fn)
                                        with open(fp, "wb") as out:
                                            out.write(img_data)
                                        image_paths.append(fp)
                                    except Exception:
                                        pass
                        except Exception:
                            pass
                    logger.info(f"🖼 Pexels 폴백 후 총 {len(image_paths)}장")

                image_path = ",".join(image_paths) if image_paths else ""
                logger.info(f"🖼 [{image_source}] 이미지 {len(image_paths)}장 수집 완료")
            except Exception as e:
                logger.warning(f"🖼 이미지 수집 실패: {e}")
                image_path = ""

        # 이미지 파일명 목록 (미리보기용)
        image_filenames = []
        if image_path:
            for p in image_path.split(","):
                p = p.strip()
                if p:
                    image_filenames.append(os.path.basename(p))

        # Bing 검색 이미지 출처 표기
        if _image_sources:
            source_domains = []
            for src_url in _image_sources:
                try:
                    domain = src_url.split("/")[2].replace("www.", "")
                    if domain not in source_domains:
                        source_domains.append(domain)
                except Exception:
                    pass
            if source_domains:
                content += "\n\n📸 이미지 출처: " + " / ".join(source_domains)

        return jsonify({
            "ok": True,
            "title": parsed.get("title", ""),
            "content": content,
            "keywords": keywords,
            "tags": tags,
            "trending": trending_keywords[:10],
            "luxury_suggestions": luxury_suggestions[:8],
            "image_path": image_path,
            "image_filenames": image_filenames,
        })
    except Exception as e:
        logger.error(f"기사 생성 오류: {e}")
        return jsonify({"ok": False, "message": f"AI 생성 오류: {str(e)}"})


@app.route(f"{URL_PREFIX}/api/free-board/regenerate-images", methods=["POST"])
@admin_required
def regenerate_article_images():
    """미선택 이미지 재생성"""
    data = request.json or {}
    count = data.get("count", 1)
    image_source = data.get("image_source", "pexels")
    article_type = data.get("article_type", "brand")
    title = data.get("title", "luxury brand")

    PEXELS_API_KEY = "ZMFMszrhmZ9oy5UTEC0XKa7h8JGytGpnLWkoFDcE4bdqxLv7r507JHEe"
    import random
    from datetime import datetime as _dt

    if article_type == "economy":
        pool = ["japanese yen currency money", "tokyo luxury shopping street", "stock market finance",
                "japan city skyline", "business newspaper coffee", "luxury watch closeup",
                "gold bars investment", "tokyo shibuya night", "currency exchange office"]
    else:
        pool = ["luxury designer handbag", "luxury boutique interior", "vintage leather bag",
                "luxury fashion accessories", "designer shoes closeup", "premium watch display",
                "luxury brand storefront", "leather goods craftsman", "fashion magazine editorial"]
    random.shuffle(pool)
    queries = pool[:count]

    try:
        import requests as _req_lib
        import json as _json
        img_dir = os.path.join(get_path("db"), "article_images")
        os.makedirs(img_dir, exist_ok=True)
        image_paths = []
        image_filenames = []

        for sq in queries:
            img_urls = []
            try:
                if image_source in ("gemini", "gemini_edit", "google_edit"):
                    try:
                        gemini_key = ""
                        env_path = os.path.join(os.path.dirname(__file__), ".env")
                        with open(env_path, encoding="utf-8") as _ef:
                            for _el in _ef:
                                if _el.strip().startswith("GEMINI_API_KEY="):
                                    gemini_key = _el.strip().split("=", 1)[1].strip()
                                    break
                        if gemini_key:
                            from google import genai
                            from google.genai import types as _gtypes
                            gclient = genai.Client(api_key=gemini_key)

                            if image_source in ("gemini_edit", "google_edit"):
                                photo_data = None
                                if image_source == "google_edit":
                                    # DuckDuckGo 이미지 검색
                                    import re as _re_g2
                                    try:
                                        hdrs = {"User-Agent": "Mozilla/5.0"}
                                        tk_url = f"https://duckduckgo.com/?q={_req_lib.utils.quote(sq)}&iax=images&ia=images"
                                        tr = _req_lib.get(tk_url, headers=hdrs, timeout=10)
                                        vqd = _re_g2.search(r'vqd=\"([^\"]+)\"', tr.text) or _re_g2.search(r'vqd=([^&\"]+)', tr.text)
                                        if vqd:
                                            ij_url = f"https://duckduckgo.com/i.js?l=ko-kr&o=json&q={_req_lib.utils.quote(sq)}&vqd={vqd.group(1)}&f=,,,,,&p=1"
                                            ir = _req_lib.get(ij_url, headers=hdrs, timeout=10)
                                            imgs = ir.json().get("results", [])
                                            random.shuffle(imgs)
                                            for im in imgs[:10]:
                                                try:
                                                    dr = _req_lib.get(im.get("image",""), headers=hdrs, timeout=10)
                                                    if dr.status_code == 200 and len(dr.content) > 5000:
                                                        photo_data = dr.content; break
                                                except: continue
                                    except: pass
                                else:
                                    r = _req_lib.get("https://api.pexels.com/v1/search",
                                        params={"query": sq, "per_page": 3, "orientation": "landscape"},
                                        headers={"Authorization": PEXELS_API_KEY}, timeout=10)
                                    if r.status_code == 200:
                                        photos = r.json().get("photos", [])
                                        if photos:
                                            photo_data = _req_lib.get(random.choice(photos)["src"]["large"], timeout=10).content
                                if photo_data:
                                    img_part = _gtypes.Part.from_bytes(data=photo_data, mime_type="image/jpeg")
                                    resp = gclient.models.generate_content(
                                        model="gemini-2.5-flash-image",
                                        contents=[f"이 사진을 고급 매거진 스타일로 보정. IMPORTANT: Do NOT include ANY text in the image.", img_part],
                                        config=_gtypes.GenerateContentConfig(response_modalities=["TEXT", "IMAGE"]),
                                    )
                                    for part in resp.candidates[0].content.parts:
                                        if part.inline_data:
                                            fn = f"article_{_dt.now().strftime('%Y%m%d_%H%M%S')}_{len(image_paths)+1}.png"
                                            fp = os.path.join(img_dir, fn)
                                            with open(fp, "wb") as out:
                                                out.write(part.inline_data.data)
                                            image_paths.append(fp)
                                            image_filenames.append(fn)
                                            break
                            else:
                                resp = gclient.models.generate_content(
                                    model="gemini-2.5-flash-image",
                                    contents=f"'{sq}' 주제의 고급 매거진 사진. 현실적인 사진처럼. 한국어 핵심 텍스트 포함. 따뜻한 조명, 고급스러운 분위기.",
                                    config=_gtypes.GenerateContentConfig(response_modalities=["TEXT", "IMAGE"]),
                                )
                                for part in resp.candidates[0].content.parts:
                                    if part.inline_data:
                                        fn = f"article_{_dt.now().strftime('%Y%m%d_%H%M%S')}_{len(image_paths)+1}.png"
                                        fp = os.path.join(img_dir, fn)
                                        with open(fp, "wb") as out:
                                            out.write(part.inline_data.data)
                                        image_paths.append(fp)
                                        image_filenames.append(fn)
                                        break
                            import time; time.sleep(2)
                    except Exception as ge:
                        logger.warning(f"🖼 Gemini 재생성 실패: {ge}")
                    continue
                elif image_source == "pexels":
                    r = _req_lib.get("https://api.pexels.com/v1/search",
                        params={"query": sq, "per_page": 3, "orientation": "landscape", "size": "medium"},
                        headers={"Authorization": PEXELS_API_KEY}, timeout=10)
                    if r.status_code == 200:
                        photos = r.json().get("photos", [])
                        if photos:
                            img_urls = [random.choice(photos)["src"]["large"]]
                else:  # pexels 폴백
                    r = _req_lib.get("https://api.pexels.com/v1/search",
                        params={"query": sq, "per_page": 3, "orientation": "landscape"},
                        headers={"Authorization": PEXELS_API_KEY}, timeout=10)
                    if r.status_code == 200:
                        photos = r.json().get("photos", [])
                        if photos:
                            img_urls = [random.choice(photos)["src"]["large"]]
            except Exception as ie:
                logger.warning(f"🖼 재생성 실패 ({sq}): {ie}")

            for img_url in img_urls:
                try:
                    ext = "png" if image_source == "dalle" else "jpg"
                    fn = f"article_{_dt.now().strftime('%Y%m%d_%H%M%S')}_{len(image_paths)+1}.{ext}"
                    fp = os.path.join(img_dir, fn)
                    resp = _req_lib.get(img_url, timeout=15)
                    with open(fp, "wb") as out:
                        out.write(resp.content)
                    image_paths.append(fp)
                    image_filenames.append(fn)
                except Exception:
                    pass

        return jsonify({"ok": True, "image_filenames": image_filenames, "image_paths": image_paths})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)})


@app.route(f"{URL_PREFIX}/api/article-image/<path:filename>")
@admin_required
def serve_article_image(filename):
    """기사 이미지 파일 서빙"""
    img_dir = os.path.join(get_path("db"), "article_images")
    from flask import send_from_directory
    return send_from_directory(img_dir, filename)


@app.route(f"{URL_PREFIX}/api/free-board/<int:post_id>/upload", methods=["POST"])
@admin_required
def upload_free_board_to_cafe(post_id):
    """자유게시판 기사를 네이버 카페에 업로드"""
    _init_free_board_db()
    from user_db import _conn as user_conn
    conn = user_conn()
    try:
        row = conn.execute("SELECT * FROM free_board WHERE id=?", (post_id,)).fetchone()
        if not row:
            return jsonify({"ok": False, "message": "기사를 찾을 수 없습니다"})

        title = row["title"]
        content = row["content"]
        menu_id = row["cafe_menu_id"] or "126"  # 자유게시판 메뉴 ID
        try:
            img_path = row["image_path"] or ""
        except Exception:
            img_path = ""
        try:
            article_tags = row["tags"] or ""
        except Exception:
            article_tags = ""

        def _upload_article():
            import asyncio
            from cafe_uploader import upload_article_to_cafe
            push_log(f"📰 자유게시판 기사 업로드: {title[:30]}...")
            if img_path:
                push_log(f"🖼 이미지 첨부: {os.path.basename(img_path)}")
            try:
                naver_data = _load_naver_accounts()
                active_slot = naver_data.get("active", 1)
                cookie_path = _get_cookie_path(active_slot)
                result = asyncio.run(upload_article_to_cafe(
                    title=title, content=content, menu_id=menu_id,
                    board_name="자유게시판", log=push_log, cookie_path=cookie_path,
                    image_path=img_path, tags=article_tags,
                ))
                if result:
                    conn2 = user_conn()
                    conn2.execute("UPDATE free_board SET status='완료' WHERE id=?", (post_id,))
                    conn2.commit()
                    conn2.close()
                    push_log(f"✅ 기사 업로드 완료: {title[:30]}")
                else:
                    conn2 = user_conn()
                    conn2.execute("UPDATE free_board SET status='실패' WHERE id=?", (post_id,))
                    conn2.commit()
                    conn2.close()
                    push_log(f"❌ 기사 업로드 실패: {title[:30]}")
            except Exception as e:
                push_log(f"❌ 기사 업로드 오류: {e}")

        import threading
        threading.Thread(target=_upload_article, daemon=True).start()
        return jsonify({"ok": True, "message": "업로드 시작"})
    finally:
        conn.close()


# ── NAS 공유 폴더 상품 동기화 ──────────────────
# 윈도우 PC → NAS에 products.db 저장 → 맥미니가 매시 30분 로컬로 가져옴
from data_manager import NAS_SHARED_PATH, get_nas_path
NAS_IMPORT_PATH = os.path.join(NAS_SHARED_PATH, "db")

def _merge_users_db(nas_users_path):
    """NAS users.db → 로컬 users.db 병합 (덮어쓰기 아닌 병합)"""
    import sqlite3 as _sq
    result = {"users_merged": 0, "tasks_merged": 0, "orders_merged": 0}
    try:
        # NAS 파일을 임시로 복사
        tmp_path = "/tmp/users_nas_tmp.db"
        with open(nas_users_path, "rb") as s, open(tmp_path, "wb") as d:
            d.write(s.read())

        nas_conn = _sq.connect(tmp_path, timeout=10)
        nas_conn.row_factory = _sq.Row

        from user_db import _conn as local_conn_fn
        local_conn = local_conn_fn()

        # NAS 테이블 목록 확인
        nas_tables = {r[0] for r in nas_conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        local_tables = {r[0] for r in local_conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}

        # 회원 (users) 병합 — username 기준
        if "users" in nas_tables and "users" in local_tables:
            nas_users = nas_conn.execute("SELECT * FROM users").fetchall()
            for r in nas_users:
                rd = dict(r)
                existing = local_conn.execute("SELECT id FROM users WHERE username=?", (rd.get("username",""),)).fetchone()
                if not existing:
                    cols = [k for k in rd.keys() if k != "id"]
                    vals = [rd[k] for k in cols]
                    try:
                        local_conn.execute(f"INSERT INTO users ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})", vals)
                        result["users_merged"] += 1
                    except Exception:
                        pass

        # 수집 작업 (scrape_tasks) 병합 — site+brand+cat 기준
        if "scrape_tasks" in nas_tables and "scrape_tasks" in local_tables:
            nas_tasks = nas_conn.execute("SELECT * FROM scrape_tasks").fetchall()
            for r in nas_tasks:
                rd = dict(r)
                existing = local_conn.execute(
                    "SELECT id FROM scrape_tasks WHERE site=? AND brand=? AND cat=?",
                    (rd.get("site",""), rd.get("brand",""), rd.get("cat",""))
                ).fetchone()
                if not existing:
                    cols = [k for k in rd.keys() if k != "id"]
                    vals = [rd[k] for k in cols]
                    try:
                        local_conn.execute(f"INSERT INTO scrape_tasks ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})", vals)
                        result["tasks_merged"] += 1
                    except Exception:
                        pass

        # 주문 (orders) 병합 — order_number 기준
        if "orders" in nas_tables and "orders" in local_tables:
            nas_orders = nas_conn.execute("SELECT * FROM orders").fetchall()
            for r in nas_orders:
                rd = dict(r)
                on = rd.get("order_number", "")
                if on:
                    existing = local_conn.execute("SELECT id FROM orders WHERE order_number=?", (on,)).fetchone()
                    if not existing:
                        cols = [k for k in rd.keys() if k != "id"]
                        vals = [rd[k] for k in cols]
                        try:
                            local_conn.execute(f"INSERT INTO orders ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})", vals)
                            result["orders_merged"] += 1
                        except Exception:
                            pass

        # 게시판/리뷰/공지 등 기타 테이블도 병합
        for tbl in ["board", "notices", "reviews", "free_board"]:
            if tbl in nas_tables and tbl in local_tables:
                nas_rows = nas_conn.execute(f"SELECT * FROM {tbl}").fetchall()
                local_ids = {r[0] for r in local_conn.execute(f"SELECT id FROM {tbl}").fetchall()}
                for r in nas_rows:
                    rd = dict(r)
                    if rd.get("id") not in local_ids:
                        cols = [k for k in rd.keys()]
                        vals = [rd[k] for k in cols]
                        try:
                            local_conn.execute(f"INSERT OR IGNORE INTO {tbl} ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})", vals)
                        except Exception:
                            pass

        local_conn.commit()
        local_conn.close()
        nas_conn.close()
        os.remove(tmp_path)

        msg = f"📂 users.db 병합: 회원 {result['users_merged']} / 작업 {result['tasks_merged']} / 주문 {result['orders_merged']}"
        push_log(msg)
        logger.info(msg)
        result["ok"] = True
        return result
    except Exception as e:
        push_log(f"⚠️ users.db 병합 오류: {e}")
        logger.warning(f"users.db 병합 오류: {e}")
        if os.path.exists("/tmp/users_nas_tmp.db"):
            os.remove("/tmp/users_nas_tmp.db")
        return {"ok": False, "message": str(e)}


def sync_all_from_nas(selected_files=None):
    """NAS → 로컬 동기화 (선택된 파일만, products.db는 병합)"""
    nas_db_dir = get_nas_path("db")
    local_db_dir = get_path("db")
    copied = []
    result = {"ok": True}

    # 기본 파일 목록 (JSON 설정 — 덮어쓰기)
    json_sync_files = ["scrape_history.json", "cafe_schedule.json",
                        "vt_cafe_schedule.json", "check_schedule.json", "fb_schedule.json",
                        "uploaded_history.json", "translation_dict.json",
                        "naver_accounts.json", "blog_accounts.json",
                        "price_config.json", "vintage_price.json", "biz_info.json", "admin_config.json"]
    sync_files = [f for f in json_sync_files if not selected_files or f in selected_files]

    try:
        for fn in sync_files:
            nas_file = os.path.join(nas_db_dir, fn)
            local_file = os.path.join(local_db_dir, fn)
            if os.path.exists(nas_file):
                try:
                    with open(nas_file, "rb") as _sf, open(local_file, "wb") as _df:
                        _df.write(_sf.read())
                    copied.append(fn)
                except Exception as e:
                    push_log(f"⚠️ {fn} 복사 실패: {e}")
        if copied:
            push_log(f"📂 NAS → 로컬 복사: {', '.join(copied)}")
    except Exception as e:
        logger.warning(f"NAS 파일 복사 실패: {e}")

    # users.db 병합 (회원/주문/작업 데이터 보존)
    if not selected_files or "users.db" in selected_files:
        try:
            nas_users = os.path.join(nas_db_dir, "users.db")
            if os.path.exists(nas_users):
                result.update(_merge_users_db(nas_users))
        except Exception as e:
            push_log(f"⚠️ users.db 병합 실패: {e}")

    # products.db 병합
    if not selected_files or "products.db" in selected_files:
        result = sync_products_from_nas()
    result["copied_files"] = copied
    return result


def sync_products_from_nas():
    """NAS products.db 파일을 로컬로 복사하여 동기화
    NAS DB를 직접 열지 않음 — 파일 복사 후 로컬에서 병합
    """
    import sqlite3 as _sq
    import shutil
    nas_db_path = os.path.join(NAS_IMPORT_PATH, "products.db")

    if not os.path.exists(nas_db_path):
        logger.debug(f"NAS DB 없음: {nas_db_path}")
        return {"ok": False, "message": "NAS에 products.db 없음"}

    try:
        nas_stat = os.stat(nas_db_path)
        nas_mtime = datetime.fromtimestamp(nas_stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        nas_size = nas_stat.st_size

        # 마지막 동기화 시간 확인
        sync_info_path = os.path.join(get_path("db"), "nas_sync_info.json")
        last_sync_mtime = ""
        if os.path.exists(sync_info_path):
            try:
                import json as _json
                with open(sync_info_path, "r") as f:
                    last_sync_mtime = _json.load(f).get("last_mtime", "")
            except Exception:
                pass

        if nas_mtime == last_sync_mtime:
            logger.debug("NAS DB 변경 없음 — 스킵")
            return {"ok": True, "message": "변경 없음", "skipped": True}

        push_log(f"📂 NAS → 로컬 동기화 시작 ({nas_size/1024/1024:.1f}MB)")

        # 1단계: NAS 파일을 임시 폴더로 복사 (NAS DB 직접 열기 금지)
        tmp_db_path = "/tmp/products_nas_tmp.db"
        # shutil.copy2는 SMB 메타데이터 복사 시 권한 오류 → 직접 읽기/쓰기
        with open(nas_db_path, "rb") as src, open(tmp_db_path, "wb") as dst:
            dst.write(src.read())
        push_log(f"📂 NAS 파일 복사 완료 → 로컬 임시 DB")

        # 2단계: 임시 DB에서 로컬 DB로 병합 (로컬 파일끼리만 작업)
        tmp_conn = _sq.connect(tmp_db_path, timeout=10)
        tmp_conn.row_factory = _sq.Row

        if last_sync_mtime:
            rows = tmp_conn.execute(
                "SELECT * FROM products WHERE created_at > ? OR scraped_at > ? ORDER BY id",
                (last_sync_mtime, last_sync_mtime)
            ).fetchall()
        else:
            rows = tmp_conn.execute("SELECT * FROM products ORDER BY id").fetchall()

        nas_total = tmp_conn.execute("SELECT COUNT(*) c FROM products").fetchone()["c"]
        tmp_conn.close()

        if not rows:
            push_log(f"📂 변경 없음 (전체 {nas_total:,}개)")
            import json as _json
            with open(sync_info_path, "w") as f:
                _json.dump({"last_mtime": nas_mtime, "last_sync": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}, f)
            os.remove(tmp_db_path)
            return {"ok": True, "message": "변경 없음", "skipped": True, "total": nas_total}

        # 로컬 DB에 병합
        from product_db import _conn as local_conn_fn, init_db as init_product_db
        try:
            init_product_db()
        except Exception:
            pass  # 이미 존재하는 테이블/인덱스 에러 무시
        local_conn = local_conn_fn()
        inserted = 0
        updated = 0
        skipped = 0

        for r in rows:
            try:
                rd = dict(r)
                site_id = rd.get("site_id", "")
                product_code = rd.get("product_code", "")
                if not site_id or not product_code:
                    skipped += 1
                    continue
                existing = local_conn.execute(
                    "SELECT id FROM products WHERE site_id=? AND product_code=?",
                    (site_id, product_code)).fetchone()
                if existing:
                    local_conn.execute("""UPDATE products SET price_jpy=?, in_stock=?, original_price=?,
                        discount_rate=?, scraped_at=?, name=?, name_ko=?, brand=?, brand_ko=?,
                        img_url=?, link=?, description=?, description_ko=?,
                        sizes=?, detail_images=?, condition_grade=?, color=?, material=? WHERE id=?""",
                        (rd.get("price_jpy",0), rd.get("in_stock",1), rd.get("original_price",0),
                         rd.get("discount_rate",0), rd.get("scraped_at",""),
                         rd.get("name",""), rd.get("name_ko",""), rd.get("brand",""), rd.get("brand_ko",""),
                         rd.get("img_url",""), rd.get("link",""), rd.get("description",""), rd.get("description_ko",""),
                         rd.get("sizes","[]"), rd.get("detail_images","[]"),
                         rd.get("condition_grade",""), rd.get("color",""), rd.get("material",""), existing["id"]))
                    updated += 1
                else:
                    cols = ["site_id","category_id","product_code","name","name_ko","brand","brand_ko",
                            "price_jpy","link","img_url","description","description_ko","sizes","detail_images",
                            "original_price","discount_rate","in_stock","scraped_at","created_at","source_type",
                            "condition_grade","color","material"]
                    vals = [rd.get(c, "") for c in cols]
                    try:
                        local_conn.execute(f"INSERT OR REPLACE INTO products ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})", vals)
                        inserted += 1
                    except Exception:
                        skipped += 1
            except Exception:
                skipped += 1

        local_conn.commit()
        local_conn.close()

        # 3단계: 임시 파일 삭제 + 동기화 시간 기록
        os.remove(tmp_db_path)
        import json as _json
        with open(sync_info_path, "w") as f:
            _json.dump({"last_mtime": nas_mtime, "last_sync": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}, f)

        msg = f"📂 동기화 완료: 신규 {inserted} / 업데이트 {updated} / 스킵 {skipped} (변경 {len(rows)} / 전체 {nas_total:,})"
        push_log(msg)
        logger.info(msg)
        return {"ok": True, "inserted": inserted, "updated": updated, "skipped": skipped, "total": nas_total}

    except Exception as e:
        # 임시 파일 정리
        import tempfile
        tmp_path = os.path.join(tempfile.gettempdir(), "products_nas_tmp.db")
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        msg = f"❌ NAS 동기화 오류: {e}"
        push_log(msg)
        logger.error(msg)
        return {"ok": False, "message": str(e)}


def export_all_to_nas(selected_files=None):
    """로컬 → NAS 공유 폴더로 복사 (선택 파일)"""
    try:
        local_db_dir = get_path("db")
        nas_db_dir = get_nas_path("db")
        if not os.path.isdir(nas_db_dir):
            return {"ok": False, "message": "NAS 경로 접근 불가"}

        all_files = ["products.db", "users.db", "scrape_history.json",
                     "cafe_schedule.json", "vt_cafe_schedule.json", "check_schedule.json",
                     "fb_schedule.json", "uploaded_history.json", "translation_dict.json",
                     "naver_accounts.json", "blog_accounts.json",
                     "price_config.json", "vintage_price.json", "biz_info.json", "admin_config.json"]
        export_files = [f for f in all_files if not selected_files or f in selected_files]
        copied = []
        for fn in export_files:
            local_file = os.path.join(local_db_dir, fn)
            nas_file = os.path.join(nas_db_dir, fn)
            if os.path.exists(local_file):
                try:
                    with open(local_file, "rb") as _sf, open(nas_file, "wb") as _df:
                        _df.write(_sf.read())
                    copied.append(fn)
                except Exception as e:
                    push_log(f"⚠️ {fn} 내보내기 실패: {e}")

        msg = f"📤 로컬 → NAS ({len(copied)}개: {', '.join(copied)})"
        push_log(msg)
        return {"ok": True, "message": msg, "files": copied}
    except Exception as e:
        logger.error(f"NAS 내보내기 실패: {e}")
        return {"ok": False, "message": str(e)}


@app.route(f"{URL_PREFIX}/api/nas-export", methods=["POST"])
@admin_required
def manual_nas_export():
    """수동: 선택 파일 → NAS 복사"""
    data = request.json or {}
    selected_files = data.get("files", None)
    result = export_all_to_nas(selected_files=selected_files)
    return jsonify(result)


@app.route(f"{URL_PREFIX}/api/nas-sync", methods=["POST"])
@admin_required
def manual_nas_sync():
    """수동 NAS 동기화 (선택 파일)"""
    data = request.json or {}
    selected_files = data.get("files", None)
    result = sync_all_from_nas(selected_files=selected_files)
    return jsonify(result)


@app.route(f"{URL_PREFIX}/api/nas-sync/status", methods=["GET"])
@admin_required
def nas_sync_status():
    """NAS 동기화 상태 확인"""
    import json as _json
    from data_manager import get_local_path
    nas_db_path = os.path.join(NAS_IMPORT_PATH, "products.db")
    local_db_path = os.path.join(get_local_path("db"), "products.db")
    info = {
        "nas_exists": os.path.exists(nas_db_path),
        "nas_path": NAS_IMPORT_PATH,
        "local_path": get_local_path("db"),
        "last_sync": "",
        "nas_mtime": "",
        "nas_size": 0,
        "local_size": 0,
    }
    if info["nas_exists"]:
        st = os.stat(nas_db_path)
        info["nas_mtime"] = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        info["nas_size"] = st.st_size
    if os.path.exists(local_db_path):
        info["local_size"] = os.stat(local_db_path).st_size

    sync_info_path = os.path.join(get_local_path("db"), "nas_sync_info.json")
    if os.path.exists(sync_info_path):
        try:
            with open(sync_info_path, "r") as f:
                d = _json.load(f)
                info["last_sync"] = d.get("last_sync", "")
        except Exception:
            pass
    return jsonify({"ok": True, **info})


# ── 상품DB 자동 업데이트 ─────────────────────
_db_update_status = {
    "running": False,
    "stop_requested": False,
    "log": [],
    "brands": [],          # [{name, code, status, count}]
    "current_brand": "",
}

def _load_db_update_schedule():
    """DB 업데이트 스케줄 설정 로드"""
    path = os.path.join(get_path("db"), "db_update_schedule.json")
    default = {"enabled": False, "interval_days": 7, "run_time": "03:00", "last_run": "", "targets": []}
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                d = json.load(f)
                default.update(d)
        except Exception:
            pass
    if "targets" not in default:
        default["targets"] = []
    return default

def _save_db_update_schedule(data):
    """DB 업데이트 스케줄 설정 저장"""
    path = os.path.join(get_path("db"), "db_update_schedule.json")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _dbu_log(msg):
    """DB 업데이트 로그 추가"""
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    _db_update_status["log"].append(line)
    # 최대 500줄 유지
    if len(_db_update_status["log"]) > 500:
        _db_update_status["log"] = _db_update_status["log"][-300:]
    logger.info(f"[DB업데이트] {msg}")

def _run_db_update_all():
    """저장된 targets 리스트를 순차 실행 (백그라운드 스레드)

    targets가 비어있으면 전체 브랜드 × 전체 카테고리 수집
    """
    import time
    from site_config import get_brands as get_site_brands, get_site
    _db_update_status["running"] = True
    _db_update_status["stop_requested"] = False
    _db_update_status["log"] = []

    site_id = "2ndstreet"
    all_brands = get_site_brands(site_id)

    # 저장된 targets 로드
    sched = _load_db_update_schedule()
    targets = sched.get("targets", [])

    # targets가 없으면 전체 브랜드로 자동 구성
    if not targets:
        targets = [{"brand_code": "", "brand_name": "전체 브랜드", "category_id": "", "category_name": "전체 카테고리"}]

    # 실행할 작업 목록 펼치기: 전체 브랜드인 항목은 개별 브랜드로 확장
    jobs = []
    for t in targets:
        if not t.get("brand_code"):
            # 전체 브랜드 → 개별 브랜드로 확장
            for code, name in all_brands.items():
                jobs.append({"brand_code": code, "brand_name": name, "category_id": t.get("category_id", ""), "category_name": t.get("category_name", "전체 카테고리")})
        else:
            jobs.append(t)

    # 브랜드 상태 초기화
    _db_update_status["brands"] = [
        {"code": j["brand_code"], "name": f'{j["brand_name"]} ({j["category_name"]})', "status": "pending", "count": 0}
        for j in jobs
    ]

    _dbu_log(f"상품DB 업데이트 시작 — {len(jobs)}개 작업 ({len(targets)}개 수집 대상)")

    for idx, job in enumerate(jobs):
        if _db_update_status["stop_requested"]:
            _dbu_log("중지 요청 — 업데이트 중단")
            for b in _db_update_status["brands"]:
                if b["status"] == "pending":
                    b["status"] = "skipped"
            break

        b_code = job["brand_code"]
        b_name = job["brand_name"]
        c_id = job.get("category_id", "")
        c_name = job.get("category_name", "전체")
        label = f"{b_name} / {c_name}"

        _dbu_log(f"[{idx+1}/{len(jobs)}] {label} 수집 시작")
        _db_update_status["current_brand"] = b_name
        _db_update_status["brands"][idx]["status"] = "running"

        try:
            # 스크래핑이 진행 중이면 대기
            wait_count = 0
            while status.get("scraping"):
                if _db_update_status["stop_requested"]:
                    break
                if wait_count == 0:
                    _dbu_log(f"   다른 스크래핑 진행 중 — 대기...")
                time.sleep(10)
                wait_count += 1
                if wait_count > 60:
                    _dbu_log(f"   대기 시간 초과 — {label} 스킵")
                    _db_update_status["brands"][idx]["status"] = "skipped"
                    continue

            if _db_update_status["stop_requested"]:
                continue

            run_scrape(
                site_id=site_id,
                category_id=c_id,
                keyword="",
                pages="",
                brand_code=b_code,
            )

            saved_count = status.get("product_count", 0)
            _db_update_status["brands"][idx]["status"] = "done"
            _db_update_status["brands"][idx]["count"] = saved_count
            _dbu_log(f"   {label} 완료 — {saved_count}개 수집")

            # 작업 간 대기
            if idx < len(jobs) - 1 and not _db_update_status["stop_requested"]:
                wait_sec = 30
                _dbu_log(f"   다음 작업 전 {wait_sec}초 대기...")
                time.sleep(wait_sec)

        except Exception as e:
            _dbu_log(f"   {label} 오류: {e}")
            _db_update_status["brands"][idx]["status"] = "error"

    # 완료 처리
    _db_update_status["running"] = False
    _db_update_status["current_brand"] = ""

    sched = _load_db_update_schedule()
    sched["last_run"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _save_db_update_schedule(sched)

    total = sum(b["count"] for b in _db_update_status["brands"] if b["status"] == "done")
    done_count = sum(1 for b in _db_update_status["brands"] if b["status"] == "done")
    _dbu_log(f"전체 완료 — {done_count}/{len(jobs)} 작업, 총 {total}개 수집")


@app.route(f"{URL_PREFIX}/api/db-update/status", methods=["GET"])
@admin_required
def db_update_status():
    """DB 업데이트 상태 조회"""
    from site_config import get_brands as get_site_brands, get_site
    sched = _load_db_update_schedule()
    # 다음 실행 시간 계산
    next_run = ""
    if sched["enabled"] and sched.get("last_run"):
        try:
            from datetime import timedelta
            last = datetime.strptime(sched["last_run"], "%Y-%m-%d %H:%M:%S")
            nxt = last + timedelta(days=sched.get("interval_days", 7))
            next_run = nxt.strftime("%Y-%m-%d") + " " + sched.get("run_time", "03:00")
        except Exception:
            pass
    elif sched["enabled"]:
        next_run = "오늘 " + sched.get("run_time", "03:00") + " (첫 실행)"

    # 브랜드/카테고리 옵션 목록
    site_info = get_site("2ndstreet") or {}
    brands_map = get_site_brands("2ndstreet")
    brand_options = [{"code": code, "name": name} for code, name in brands_map.items()]
    cat_options = [{"id": cid, "name": cat.get("name", cid)} for cid, cat in site_info.get("categories", {}).items()]

    return jsonify({
        "ok": True,
        "running": _db_update_status["running"],
        "enabled": sched.get("enabled", False),
        "interval_days": sched.get("interval_days", 7),
        "run_time": sched.get("run_time", "03:00"),
        "last_run": sched.get("last_run", ""),
        "next_run": next_run,
        "brands": _db_update_status["brands"],
        "targets": sched.get("targets", []),
        "log": "\n".join(_db_update_status["log"][-100:]),
        "options": {
            "brands": brand_options,
            "categories": cat_options,
        },
    })


@app.route(f"{URL_PREFIX}/api/db-update/schedule", methods=["POST"])
@admin_required
def db_update_schedule_save():
    """DB 업데이트 스케줄 저장"""
    data = request.get_json() or {}
    sched = _load_db_update_schedule()
    sched["enabled"] = data.get("enabled", False)
    sched["interval_days"] = data.get("interval_days", 7)
    sched["run_time"] = data.get("run_time", "03:00")
    _save_db_update_schedule(sched)

    # 스케줄러 재등록
    _register_db_update_job()

    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/api/db-update/run", methods=["POST"])
@admin_required
def db_update_run():
    """DB 업데이트 수동 실행 (저장된 targets 기반)"""
    if _db_update_status["running"]:
        return jsonify({"ok": False, "message": "이미 실행 중입니다"})
    t = threading.Thread(target=_run_db_update_all, daemon=True)
    t.start()
    return jsonify({"ok": True, "message": "업데이트 시작"})


@app.route(f"{URL_PREFIX}/api/db-update/stop", methods=["POST"])
@admin_required
def db_update_stop():
    """DB 업데이트 중지"""
    _db_update_status["stop_requested"] = True
    status["stop_requested"] = True  # 현재 진행 중인 스크래핑도 중지
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/api/db-update/targets", methods=["POST"])
@admin_required
def db_update_targets():
    """수집 대상 리스트 추가/삭제"""
    data = request.get_json() or {}
    sched = _load_db_update_schedule()
    targets = sched.get("targets", [])
    action = data.get("action", "")

    if action == "add":
        brand_code = data.get("brand_code", "")
        category_id = data.get("category_id", "")
        brand_name = data.get("brand_name", "전체 브랜드")
        category_name = data.get("category_name", "전체 카테고리")
        # 중복 체크
        for t in targets:
            if t.get("brand_code") == brand_code and t.get("category_id") == category_id:
                return jsonify({"ok": False, "message": "이미 등록된 수집 대상입니다", "targets": targets})
        targets.append({
            "brand_code": brand_code,
            "brand_name": brand_name,
            "category_id": category_id,
            "category_name": category_name,
        })
    elif action == "remove":
        idx = data.get("index", -1)
        if 0 <= idx < len(targets):
            targets.pop(idx)

    sched["targets"] = targets
    _save_db_update_schedule(sched)
    return jsonify({"ok": True, "targets": targets})


@app.route(f"{URL_PREFIX}/api/price-changes", methods=["GET"])
@admin_required
def api_price_changes():
    """가격 변경 이력 조회 (관리자)"""
    from product_db import get_price_changes
    change_type = request.args.get("type", "")  # 가격인하 / 가격인상 / 빈값=전체
    limit = int(request.args.get("limit", 200))
    changes = get_price_changes(change_type=change_type, limit=limit)
    return jsonify({"ok": True, "changes": changes, "total": len(changes)})


@app.route(f"{URL_PREFIX}/api/price-changes/shop", methods=["GET"])
def api_price_changes_shop():
    """가격 인하 이력 조회 (쇼핑몰 — 로그인 불필요, 원화 변환)"""
    from product_db import get_price_changes
    changes = get_price_changes(change_type="가격인하", limit=100)
    # 고객 레벨 (로그인 시)
    lvl = session.get("level", "b2c") if session.get("logged_in") else "b2c"
    for c in changes:
        c["old_krw"] = _calc_vintage_price(c.get("old_price", 0), lvl)
        c["new_krw"] = _calc_vintage_price(c.get("new_price", 0), lvl)
        # 고객에게 원가(엔화) 노출 금지
        c.pop("old_price", None)
        c.pop("new_price", None)
    return jsonify({"ok": True, "changes": changes})


def _register_db_update_job():
    """DB 업데이트 스케줄러 등록"""
    job_id = "db_update_auto"
    try:
        scheduler.remove_job(job_id)
    except Exception:
        pass

    sched = _load_db_update_schedule()
    if not sched.get("enabled"):
        return

    run_time = sched.get("run_time", "03:00")
    hour, minute = (int(x) for x in run_time.split(":"))
    interval_days = sched.get("interval_days", 7)

    def _scheduled_db_update():
        """스케줄러에서 호출 — 주기 체크 후 실행"""
        s = _load_db_update_schedule()
        last_run = s.get("last_run", "")
        if last_run:
            try:
                from datetime import timedelta
                last = datetime.strptime(last_run, "%Y-%m-%d %H:%M:%S")
                if datetime.now() - last < timedelta(days=interval_days):
                    return  # 아직 주기 안 됨
            except Exception:
                pass
        if not _db_update_status["running"]:
            t = threading.Thread(target=_run_db_update_all, daemon=True)
            t.start()

    scheduler.add_job(
        _scheduled_db_update,
        "cron",
        hour=hour,
        minute=minute,
        id=job_id,
        replace_existing=True,
    )
    logger.info(f"[DB업데이트] 스케줄 등록: 매일 {hour:02d}:{minute:02d} 체크 (주기 {interval_days}일)")


# ── 빈티지 가격 설정 ─────────────────────
_vintage_price_config = {
    "jp_fee_pct": 3.0,
    "buy_markup_pct": 2.0,
    "margin_b2c_pct": 15.0,
    "margin_b2b_pct": 8.0,
    "jp_domestic_shipping": 800,
    "intl_shipping_krw": 15000,
}

# 파일에서 로드
def _load_vintage_price():
    import json as _json
    path = os.path.join(get_path("db"), "vintage_price.json")
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                _vintage_price_config.update(_json.load(f))
        except Exception:
            pass

def _save_vintage_price():
    import json as _json
    path = os.path.join(get_path("db"), "vintage_price.json")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        _json.dump(_vintage_price_config, f)

_load_vintage_price()

# 서버 시작 시 '수집중' 멈춤 작업 자동 복구
try:
    import sqlite3 as _sq
    _db = os.path.join(get_path("db"), "users.db")
    if os.path.exists(_db):
        _c = _sq.connect(_db)
        _stuck = _c.execute("SELECT count(*) FROM scrape_tasks WHERE status='수집중'").fetchone()[0]
        if _stuck > 0:
            _c.execute("UPDATE scrape_tasks SET status='대기' WHERE status='수집중'")
            _c.commit()
            logger.info(f"🔧 서버 시작: 수집중 멈춤 작업 {_stuck}건 → 대기로 복구")
        _c.close()
except Exception:
    pass


@app.route(f"{URL_PREFIX}/settings/vintage-price", methods=["GET"])
@admin_required
def get_vintage_price():
    return jsonify({"ok": True, **_vintage_price_config})


@app.route(f"{URL_PREFIX}/settings/vintage-price", methods=["POST"])
@admin_required
def update_vintage_price():
    data = request.json or {}
    if data.get("jp_fee_pct") is not None:
        _vintage_price_config["jp_fee_pct"] = float(data["jp_fee_pct"])
    if data.get("buy_markup_pct") is not None:
        _vintage_price_config["buy_markup_pct"] = float(data["buy_markup_pct"])
    if data.get("margin_b2c_pct") is not None:
        _vintage_price_config["margin_b2c_pct"] = float(data["margin_b2c_pct"])
    if data.get("margin_b2b_pct") is not None:
        _vintage_price_config["margin_b2b_pct"] = float(data["margin_b2b_pct"])
    if data.get("jp_domestic_shipping") is not None:
        _vintage_price_config["jp_domestic_shipping"] = int(data["jp_domestic_shipping"])
    if data.get("intl_shipping_krw") is not None:
        _vintage_price_config["intl_shipping_krw"] = int(data["intl_shipping_krw"])
    _save_vintage_price()
    msg = (f"빈티지 가격설정: 수수료={_vintage_price_config['jp_fee_pct']}% "
           f"환율추가={_vintage_price_config['buy_markup_pct']}% "
           f"B2C={_vintage_price_config['margin_b2c_pct']}% "
           f"B2B={_vintage_price_config['margin_b2b_pct']}% "
           f"일본택배=¥{_vintage_price_config.get('jp_domestic_shipping',800):,} "
           f"국제배송={_vintage_price_config['intl_shipping_krw']:,}원")
    push_log("🎺 " + msg)
    return jsonify({"ok": True, **_vintage_price_config, "message": msg})


# ── 데이터 경로 설정 ─────────────────────

@app.route(f"{URL_PREFIX}/settings/data-path", methods=["GET"])
@admin_required
def get_data_path():
    """데이터 저장 경로 상태 조회"""
    return jsonify({"ok": True, **get_data_status()})


@app.route(f"{URL_PREFIX}/settings/data-path", methods=["POST"])
@admin_required
def update_data_path():
    """데이터 저장 경로 변경"""
    data = request.json or {}
    new_path = data.get("path", "").strip()
    if not new_path:
        return jsonify({"ok": False, "message": "경로를 입력해주세요"})

    ok = set_data_root(new_path)
    if ok:
        push_log(f"📁 데이터 경로 변경: {new_path}")
        return jsonify({"ok": True, "message": f"경로 변경 완료: {new_path}", **get_data_status()})
    else:
        return jsonify({"ok": False, "message": "경로 생성 실패 — 경로를 확인해주세요"})


@app.route(f"{URL_PREFIX}/settings/data-path/reset", methods=["POST"])
@admin_required
def reset_data_path():
    """데이터 저장 경로 초기화 (OS 기본값)"""
    from data_manager import _default_path
    default = _default_path()
    ok = set_data_root(default)
    if ok:
        push_log(f"📁 데이터 경로 초기화: {default}")
        return jsonify({"ok": True, "message": f"기본 경로로 초기화: {default}", **get_data_status()})
    return jsonify({"ok": False, "message": "초기화 실패"})


# ── AI 설정 ─────────────────────────────

@app.route(f"{URL_PREFIX}/settings/ai", methods=["GET"])
@admin_required
def get_ai_settings():
    """AI 설정 조회"""
    return jsonify(get_ai_config())


@app.route(f"{URL_PREFIX}/settings/ai", methods=["POST"])
@admin_required
def update_ai_settings():
    """AI 설정 변경 (provider, gemini_key, claude_key, openai_key)"""
    data = request.json or {}
    set_ai_config(
        provider=data.get("provider"),
        gemini_key=data.get("gemini_key"),
        claude_key=data.get("claude_key"),
        openai_key=data.get("openai_key"),
    )
    push_log(f"🤖 AI 설정 변경: {data.get('provider', '변경없음')}")
    return jsonify({"ok": True, **get_ai_config()})


@app.route(f"{URL_PREFIX}/settings/ai/test", methods=["POST"])
@admin_required
def test_ai():
    """AI 연결 테스트"""
    try:
        from post_generator import verify_ai_key, _ai_config
        provider = _ai_config["provider"]
        has_key = bool(_ai_config.get("openai_key")) if provider == "openai" else \
                  bool(_ai_config.get("gemini_key")) if provider == "gemini" else \
                  bool(_ai_config.get("claude_key"))
        logger.info(f"🧪 AI 테스트 — provider: {provider}, key_set: {has_key}")
        result = verify_ai_key()
        logger.info(f"🧪 AI 테스트 결과 — ok: {result['ok']}, msg: {result['message']}")
        if result["ok"]:
            return jsonify({"ok": True, "provider": result["provider"], "response": result["message"]})
        else:
            return jsonify({"ok": False, "message": f"[{result['provider']}] {result['message']}"})
    except Exception as e:
        import traceback
        logger.error(f"🧪 AI 테스트 예외: {traceback.format_exc()}")
        return jsonify({"ok": False, "message": str(e)})


# ── AI 채팅 위젯 API ───────────────────────

@app.route(f"{URL_PREFIX}/chat", methods=["POST"])
@admin_required
def api_chat():
    """AI 채팅 위젯 — 선택된 AI 모델과 대화"""
    data = request.json or {}
    message = data.get("message", "").strip()
    history = data.get("history", [])
    if not message:
        return jsonify({"ok": False, "reply": "메시지를 입력해주세요."})
    from post_generator import chat_with_ai
    result = chat_with_ai(message, history)
    return jsonify(result)


# ── 텔레그램 알림 설정 ────────────────────

@app.route(f"{URL_PREFIX}/settings/telegram", methods=["GET"])
@admin_required
def get_telegram_settings():
    """텔레그램 설정 조회"""
    from notifier import get_telegram_config
    return jsonify({"ok": True, **get_telegram_config()})


@app.route(f"{URL_PREFIX}/settings/telegram", methods=["POST"])
@admin_required
def update_telegram_settings():
    """텔레그램 설정 변경"""
    from notifier import set_telegram_config, get_telegram_config
    data = request.json or {}
    set_telegram_config(
        bot_token=data.get("bot_token"),
        chat_id=data.get("chat_id"),
    )
    push_log("📬 텔레그램 설정 변경")
    return jsonify({"ok": True, **get_telegram_config()})


@app.route(f"{URL_PREFIX}/settings/telegram/test", methods=["POST"])
@admin_required
def test_telegram():
    """텔레그램 연결 테스트"""
    from notifier import send_telegram, is_configured
    if not is_configured():
        return jsonify({"ok": False, "message": "텔레그램 설정이 필요합니다 (Bot Token + Chat ID)"})
    ok = send_telegram("🔔 JP Sourcing 텔레그램 알림 테스트!")
    if ok:
        return jsonify({"ok": True, "message": "테스트 메시지 전송 성공!"})
    return jsonify({"ok": False, "message": "전송 실패 — Token/Chat ID를 확인해주세요"})


# ── 네이버 로그인 (쿠키 저장) ────────────

@app.route(f"{URL_PREFIX}/naver/status")
@admin_required
def naver_status():
    """네이버 로그인 상태 확인"""
    return jsonify({"logged_in": has_saved_cookies()})


# ── 네이버 계정 관리 (최대 3개) ─────────────────

_NAVER_ACCOUNTS_DB = os.path.join(get_path("db"), "naver_accounts.json")


def _load_naver_accounts() -> dict:
    """네이버 계정 목록 로드"""
    if os.path.exists(_NAVER_ACCOUNTS_DB):
        try:
            with open(_NAVER_ACCOUNTS_DB, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"active": 1, "accounts": {}}


def _save_naver_accounts(data: dict):
    """네이버 계정 목록 저장"""
    os.makedirs(os.path.dirname(_NAVER_ACCOUNTS_DB), exist_ok=True)
    with open(_NAVER_ACCOUNTS_DB, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _get_cookie_path(slot: int) -> str:
    """슬롯별 쿠키 파일 경로"""
    if slot == 1:
        return "naver_cookies.json"  # 기존 호환
    return f"naver_cookies_{slot}.json"


@app.route(f"{URL_PREFIX}/naver/accounts", methods=["GET"])
@admin_required
def get_naver_accounts():
    """네이버 계정 목록 조회 (비밀번호 마스킹)"""
    data = _load_naver_accounts()
    result = {"active": data.get("active", 1), "accounts": {}}
    for slot, acc in data.get("accounts", {}).items():
        has_cookie = os.path.exists(_get_cookie_path(int(slot)))
        result["accounts"][slot] = {
            "naver_id": acc.get("naver_id", ""),
            "has_password": bool(acc.get("password")),
            "has_cookie": has_cookie,
        }
    return jsonify(result)


@app.route(f"{URL_PREFIX}/naver/accounts", methods=["POST"])
@admin_required
def save_naver_account():
    """네이버 계정 저장"""
    d = request.json or {}
    slot = str(d.get("slot", 1))
    naver_id = d.get("naver_id", "").strip()
    password = d.get("password", "").strip()
    if not naver_id:
        return jsonify({"ok": False, "message": "아이디를 입력해주세요"})

    data = _load_naver_accounts()
    if "accounts" not in data:
        data["accounts"] = {}
    existing = data["accounts"].get(slot, {})
    # 비밀번호가 비어있으면 기존 비밀번호 유지
    if not password and existing.get("password"):
        password = existing["password"]
    data["accounts"][slot] = {"naver_id": naver_id, "password": password}
    _save_naver_accounts(data)
    pw_msg = "비밀번호 저장됨" if password else "비밀번호 미설정"
    push_log(f"💾 네이버 계정 {slot} 저장: {naver_id} ({pw_msg})")
    return jsonify({"ok": True, "message": f"저장 완료 ({pw_msg})"})


@app.route(f"{URL_PREFIX}/naver/accounts/delete", methods=["POST"])
@admin_required
def delete_naver_account():
    """네이버 계정 삭제"""
    d = request.json or {}
    slot = str(d.get("slot", 1))
    data = _load_naver_accounts()
    if slot in data.get("accounts", {}):
        del data["accounts"][slot]
        _save_naver_accounts(data)
    # 쿠키 파일도 삭제
    cookie_path = _get_cookie_path(int(slot))
    if os.path.exists(cookie_path):
        os.remove(cookie_path)
    push_log(f"🗑️ 네이버 계정 {slot} 삭제")
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/naver/accounts/active", methods=["POST"])
@admin_required
def set_active_naver_account():
    """활성 계정 변경"""
    d = request.json or {}
    slot = int(d.get("slot", 1))
    data = _load_naver_accounts()
    data["active"] = slot
    _save_naver_accounts(data)
    # 활성 계정의 쿠키를 기본 쿠키 경로에 복사
    src = _get_cookie_path(slot)
    if os.path.exists(src) and slot != 1:
        import shutil
        shutil.copy2(src, "naver_cookies.json")
    push_log(f"✅ 활성 네이버 계정 변경: 계정 {slot}")
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/naver/login", methods=["POST"])
@admin_required
def naver_cafe_login():
    """네이버 카페 로그인 시작 (저장된 계정 자동 입력)"""
    d = request.json or {}
    slot = int(d.get("slot", 1))
    cookie_path = _get_cookie_path(slot)

    # 저장된 계정 정보 가져오기
    acc_data = _load_naver_accounts()
    acc = acc_data.get("accounts", {}).get(str(slot), {})
    naver_id = acc.get("naver_id", "")
    password = acc.get("password", "")

    def run_login():
        from cafe_uploader import naver_manual_login_with_cookie_path
        result = asyncio.run(naver_manual_login_with_cookie_path(
            cookie_path=cookie_path, status_callback=push_log,
            naver_id=naver_id, password=password,
        ))
        if result:
            push_log(f"✅ 네이버 계정 {slot} 로그인 & 쿠키 저장 완료!")
            # 활성 계정이면 기본 쿠키에도 복사
            data = _load_naver_accounts()
            if data.get("active") == slot and slot != 1:
                import shutil
                shutil.copy2(cookie_path, "naver_cookies.json")
        else:
            push_log(f"❌ 네이버 계정 {slot} 로그인 실패 또는 시간 초과")

    thread = threading.Thread(target=run_login, daemon=True)
    thread.start()
    return jsonify({"ok": True, "message": f"계정 {slot} 로그인 브라우저가 열립니다."})


@app.route(f"{URL_PREFIX}/naver/logout", methods=["POST"])
@admin_required
def naver_logout():
    """네이버 쿠키 삭제"""
    delete_cookies()
    push_log("🗑️ 네이버 쿠키 삭제 완료")
    return jsonify({"ok": True, "message": "네이버 로그인 정보가 삭제되었습니다"})


# ── 블로그 계정 관리 ────────────────────────
_BLOG_ACCOUNTS_DB = os.path.join(get_path("db"), "blog_accounts.json")


def _load_blog_accounts():
    if os.path.exists(_BLOG_ACCOUNTS_DB):
        try:
            with open(_BLOG_ACCOUNTS_DB, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"active": 1, "accounts": {}}


def _save_blog_accounts(data: dict):
    os.makedirs(os.path.dirname(_BLOG_ACCOUNTS_DB), exist_ok=True)
    with open(_BLOG_ACCOUNTS_DB, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _get_blog_cookie_path(slot: int) -> str:
    return f"blog_cookies_{slot}.json"


@app.route(f"{URL_PREFIX}/blog/accounts", methods=["GET"])
@admin_required
def get_blog_accounts():
    data = _load_blog_accounts()
    result = {"active": data.get("active", 1), "accounts": {}}
    for slot, acc in data.get("accounts", {}).items():
        has_cookie = os.path.exists(_get_blog_cookie_path(int(slot)))
        result["accounts"][slot] = {
            "naver_id": acc.get("naver_id", ""),
            "blog_id": acc.get("blog_id", ""),
            "has_password": bool(acc.get("password")),
            "has_cookie": has_cookie,
        }
    return jsonify(result)


@app.route(f"{URL_PREFIX}/blog/accounts", methods=["POST"])
@admin_required
def save_blog_account():
    d = request.json or {}
    slot = str(d.get("slot", 1))
    naver_id = d.get("naver_id", "").strip()
    blog_id = d.get("blog_id", "").strip()
    password = d.get("password", "").strip()
    if not naver_id:
        return jsonify({"ok": False, "message": "아이디를 입력해주세요"})
    data = _load_blog_accounts()
    if "accounts" not in data:
        data["accounts"] = {}
    data["accounts"][slot] = {"naver_id": naver_id, "blog_id": blog_id, "password": password}
    _save_blog_accounts(data)
    push_log(f"💾 블로그 계정 {slot} 저장: {naver_id} (블로그: {blog_id})")
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/blog/accounts/delete", methods=["POST"])
@admin_required
def delete_blog_account():
    d = request.json or {}
    slot = str(d.get("slot", 1))
    data = _load_blog_accounts()
    if slot in data.get("accounts", {}):
        del data["accounts"][slot]
        _save_blog_accounts(data)
    cookie_path = _get_blog_cookie_path(int(slot))
    if os.path.exists(cookie_path):
        os.remove(cookie_path)
    push_log(f"🗑️ 블로그 계정 {slot} 삭제")
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/blog/accounts/active", methods=["POST"])
@admin_required
def set_active_blog_account():
    d = request.json or {}
    slot = int(d.get("slot", 1))
    data = _load_blog_accounts()
    data["active"] = slot
    _save_blog_accounts(data)
    push_log(f"✅ 활성 블로그 계정 변경: 계정 {slot}")
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/blog/login", methods=["POST"])
@admin_required
def blog_login():
    d = request.json or {}
    slot = int(d.get("slot", 1))
    cookie_path = _get_blog_cookie_path(slot)

    # 저장된 계정 정보 가져오기 (카페 계정과 동일)
    acc_data = _load_naver_accounts()
    acc = acc_data.get("accounts", {}).get(str(slot), {})
    naver_id = acc.get("naver_id", "")
    password = acc.get("password", "")

    def run_login():
        from cafe_uploader import naver_manual_login_with_cookie_path
        result = asyncio.run(naver_manual_login_with_cookie_path(
            cookie_path=cookie_path, status_callback=push_log,
            naver_id=naver_id, password=password,
        ))
        if result:
            push_log(f"✅ 블로그 계정 {slot} 로그인 & 쿠키 저장 완료!")
        else:
            push_log(f"❌ 블로그 계정 {slot} 로그인 실패 또는 시간 초과")

    thread = threading.Thread(target=run_login, daemon=True)
    thread.start()
    return jsonify({"ok": True, "message": f"블로그 계정 {slot} 로그인 브라우저가 열립니다."})


@app.route(f"{URL_PREFIX}/blog/fetch-url", methods=["POST"])
@admin_required
def blog_fetch_url():
    """URL에서 제목, 본문, 이미지 추출 (JS 렌더링 사이트는 Playwright 사용)"""
    d = request.json or {}
    url = d.get("url", "").strip()
    if not url:
        return jsonify({"error": "URL이 비어있습니다"})

    # JS 렌더링이 필요한 사이트 목록
    js_sites = ["smartstore.naver.com", "shopping.naver.com", "brand.naver.com"]
    needs_playwright = any(site in url for site in js_sites)

    if needs_playwright:
        try:
            result = asyncio.run(_fetch_url_playwright(url))
            push_log(f"🌐 URL 추출 완료 (Playwright): {result.get('title', '')[:40]}... (이미지 {len(result.get('images', []))}개)")
            return jsonify(result)
        except Exception as e:
            return jsonify({"error": f"Playwright 추출 실패: {e}"})

    try:
        import requests as _req
        from bs4 import BeautifulSoup
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        resp = _req.get(url, headers=headers, timeout=15)
        resp.encoding = resp.apparent_encoding or "utf-8"
        soup = BeautifulSoup(resp.text, "html.parser")

        # 제목 추출
        title = ""
        for sel in [soup.find("meta", property="og:title"),
                    soup.find("meta", attrs={"name": "title"}),
                    soup.find("title")]:
            if sel:
                title = sel.get("content", "") if sel.name == "meta" else sel.get_text()
                if title:
                    break

        # 본문 추출
        body = ""
        for tag in ["article", "main", "[class*='content']", "[class*='detail']", "[class*='post']", "body"]:
            el = soup.select_one(tag)
            if el:
                for s in el.find_all(["script", "style", "nav", "header", "footer"]):
                    s.decompose()
                text = el.get_text(separator="\n", strip=True)
                if len(text) > 100:
                    body = text
                    break

        # 이미지 추출
        images = _extract_images_from_soup(soup, url)

        # 본문이 너무 짧으면 Playwright로 재시도
        if len(body) < 100:
            try:
                result = asyncio.run(_fetch_url_playwright(url))
                push_log(f"🌐 URL 추출 완료 (Playwright 폴백): {result.get('title', '')[:40]}...")
                return jsonify(result)
            except Exception:
                pass

        push_log(f"🌐 URL 추출 완료: {title[:40]}... (이미지 {len(images)}개)")
        return jsonify({"title": title.strip(), "body": body[:5000], "images": images})
    except Exception as e:
        return jsonify({"error": str(e)})


def _extract_images_from_soup(soup, base_url):
    """BeautifulSoup에서 이미지 URL 추출"""
    from urllib.parse import urlparse
    images = []
    seen = set()
    for img in soup.find_all("img"):
        src = img.get("src") or img.get("data-src") or ""
        if not src or src in seen:
            continue
        if src.startswith("//"):
            src = "https:" + src
        elif src.startswith("/"):
            parsed = urlparse(base_url)
            src = f"{parsed.scheme}://{parsed.netloc}{src}"
        w = img.get("width", "")
        h = img.get("height", "")
        if w and w.isdigit() and int(w) < 50:
            continue
        if h and h.isdigit() and int(h) < 50:
            continue
        if any(x in src.lower() for x in ["logo", "icon", "banner", "ad", "pixel", "tracking", "1x1"]):
            continue
        seen.add(src)
        images.append(src)
        if len(images) >= 30:
            break
    return images


async def _fetch_url_playwright(url: str) -> dict:
    """Playwright로 JS 렌더링 후 콘텐츠 추출 (스마트스토어 등)"""
    from playwright.async_api import async_playwright
    from cafe_uploader import load_cookies

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, args=["--no-sandbox"])
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )

        # 네이버 계정 쿠키 로드 (로그인 상태로 접근)
        try:
            naver_data = _load_naver_accounts()
            active_slot = naver_data.get("active", 1)
            cookie_path = _get_cookie_path(active_slot)
            cookies = load_cookies(cookie_path)
            if cookies:
                await context.add_cookies(cookies)
                logger.info(f"🍪 네이버 쿠키 로드 (계정 {active_slot})")
        except Exception:
            pass

        page = await context.new_page()
        try:
            await page.goto(url, wait_until="networkidle", timeout=30000)
            await asyncio.sleep(3)

            # 스마트스토어 상세 영역 스크롤 (이미지 lazy load 대응)
            for _ in range(5):
                await page.keyboard.press("PageDown")
                await asyncio.sleep(0.5)

            # 더보기 버튼 클릭 (상세 정보 펼치기)
            try:
                more_btn = page.locator("a:has-text('상품정보 더보기'), button:has-text('더보기'), [class*='more']").first
                if await more_btn.count() > 0:
                    await more_btn.click()
                    await asyncio.sleep(2)
            except Exception:
                pass

            # 상세 페이지 끝까지 스크롤 (lazy load 이미지 전부 로드)
            for _ in range(20):
                await page.keyboard.press("PageDown")
                await asyncio.sleep(0.3)
            await asyncio.sleep(2)

            # 제목 추출
            title = ""
            for sel in ["h3._22kNQuEXmb", "h3[class*='title']", "._3oDjSvLDjZ", "h2", "h3", "title"]:
                try:
                    el = page.locator(sel).first
                    if await el.count() > 0:
                        title = (await el.inner_text()).strip()
                        if title:
                            break
                except Exception:
                    continue
            if not title:
                title = await page.title()

            # 본문 텍스트 추출 — se-main-container 내 se-text, se-sectionTitle만
            # se-quotation(인삿말/공지) 완전 제외
            body = ""
            try:
                body = await page.evaluate("""() => {
                    const container = document.querySelector('div.se-main-container');
                    if (!container) return '';
                    const lines = [];
                    // se-main-container 안의 모든 텍스트 paragraph 추출 (이미지/구분선 제외, 나머지 전부)
                    const paragraphs = container.querySelectorAll('p.se-text-paragraph');
                    for (const p of paragraphs) {
                        const text = p.innerText.replace(/\\u200B/g, '').trim();
                        if (text) lines.push(text);
                    }
                    return lines.join('\\n');
                }""")
            except Exception:
                pass

            # se-main-container 못 찾으면 폴백
            if not body:
                for sel in ["div._1Hj-MkenCi", "div._3e8dOKsKKM", "div[class*='detail']", "div[class*='content']"]:
                    try:
                        el = page.locator(sel).first
                        if await el.count() > 0:
                            text = await el.inner_text()
                            if len(text) > len(body):
                                body = text
                    except Exception:
                        continue

            # 이미지 추출 — se-main-container 내 shop-phinf 이미지만
            images = []
            seen = set()

            # 스마트스토어 상세 본문: se-main-container > se-image-resource
            detail_img_selectors = [
                "div.se-main-container img.se-image-resource",
                "div.se-viewer img.se-image-resource",
                "div._1Hj-MkenCi img",
                "div._3e8dOKsKKM img",
            ]
            detail_imgs = []
            for sel in detail_img_selectors:
                try:
                    imgs = await page.query_selector_all(sel)
                    if imgs:
                        detail_imgs = imgs
                        break
                except Exception:
                    continue

            if not detail_imgs:
                detail_imgs = await page.query_selector_all("img")

            for img in detail_imgs:
                for attr in ["src", "data-src", "data-lazy-src", "data-original"]:
                    src = await img.get_attribute(attr) or ""
                    if src and "shop-phinf.pstatic.net" in src:
                        if src.startswith("//"):
                            src = "https:" + src
                        if src not in seen:
                            seen.add(src)
                            images.append(src)
                        break
                if len(images) >= 30:
                    break

            body = body.strip()[:8000]
            return {"title": title.strip(), "body": body, "images": images}

        finally:
            await browser.close()


@app.route(f"{URL_PREFIX}/blog/post-url-content", methods=["POST"])
@admin_required
def blog_post_url_content():
    """URL에서 추출한 콘텐츠를 블로그에 발행"""
    d = request.json or {}
    title = d.get("title", "").strip()
    body = d.get("body", "").strip()
    images = d.get("images", [])
    category = d.get("category", "").strip()
    if not title or not body:
        return jsonify({"ok": False, "error": "제목과 본문이 필요합니다"})

    def run_post():
        try:
            from blog_uploader import blog_post_custom_content
            result = asyncio.run(blog_post_custom_content(
                title=title, body=body, images=images, log=push_log,
                category=category
            ))
            if result:
                push_log(f"✅ 블로그 URL 콘텐츠 발행 성공!")
            else:
                push_log(f"❌ 블로그 URL 콘텐츠 발행 실패")
        except Exception as e:
            push_log(f"❌ 블로그 발행 오류: {e}")

    thread = threading.Thread(target=run_post, daemon=True)
    thread.start()
    return jsonify({"ok": True, "message": "블로그 발행 시작"})


@app.route(f"{URL_PREFIX}/run/stop", methods=["POST"])
def stop_all():
    """가벼운 중단: 크롤링/업로드 중지 + 브라우저 정리 (데이터 삭제 없음)"""
    status["stop_requested"] = True
    status["paused"] = False

    def _cleanup():
        try:
            asyncio.run(force_close_browser())
        except Exception:
            pass
        # 2ndstreet 브라우저도 정리
        try:
            from secondst_crawler import force_close_browser as close_2nd
            asyncio.run(close_2nd())
        except Exception:
            pass
    threading.Thread(target=_cleanup, daemon=True).start()

    status["scraping"] = False
    status["uploading"] = False
    push_log("⛔ 페이지 이탈 감지 — 작업 중단 + 브라우저 정리")
    return jsonify({"ok": True})


@app.route(f"{URL_PREFIX}/run/pause", methods=["POST"])
@admin_required
def pause_scrape():
    """일시정지: 현재 상품 완료 후 멈춤"""
    if not status["scraping"]:
        return jsonify({"ok": False, "message": "실행 중인 작업이 없습니다"})
    status["paused"] = True
    push_log("⏸️ 일시정지 요청 — 현재 상품 수집 완료 후 멈춥니다...")
    return jsonify({"ok": True, "message": "일시정지 요청됨"})


@app.route(f"{URL_PREFIX}/run/resume", methods=["POST"])
@admin_required
def resume_scrape():
    """일시정지 해제"""
    status["paused"] = False
    push_log("▶️ 재개 — 수집을 계속합니다!")
    return jsonify({"ok": True, "message": "재개됨"})


@app.route(f"{URL_PREFIX}/run/unlock", methods=["POST"])
@admin_required
def unlock_status():
    """상태 잠금 해제 (데이터 삭제 없이 stuck 상태만 리셋)"""
    was_scraping = status["scraping"]
    was_uploading = status["uploading"]
    status["scraping"] = False
    status["uploading"] = False
    status["paused"] = False
    status["stop_requested"] = False
    msg = "🔓 상태 잠금 해제 완료"
    if was_scraping or was_uploading:
        msg += f" (scraping={was_scraping}, uploading={was_uploading} → False)"
    push_log(msg)
    return jsonify({"ok": True, "message": msg})


@app.route(f"{URL_PREFIX}/run/force-stop", methods=["POST"])
@admin_required
def force_stop_scrape():
    """스크래핑 강제 중지 (상태 리셋 + 브라우저 종료)"""
    import asyncio
    status["scraping"] = False
    status["stop_requested"] = True
    status["paused"] = False
    try:
        asyncio.run(force_close_browser())
    except Exception:
        pass
    try:
        from secondst_crawler import force_close_browser as fc2
        asyncio.run(fc2())
    except Exception:
        pass
    push_log("⛔ 스크래핑 강제 중지 완료")
    return jsonify({"ok": True, "message": "강제 중지 완료"})


@app.route(f"{URL_PREFIX}/run/reset", methods=["POST"])
@admin_required
def reset_all():
    """리셋: 수집 중단 + 브라우저 강제 종료 + 데이터 삭제 + 상태 초기화"""
    import glob, shutil

    # 중단 요청 (진행 중인 작업만)
    if status["scraping"] or status.get("uploading"):
        status["stop_requested"] = True
    status["paused"] = False

    # 브라우저 강제 종료 (백그라운드 스레드에서 실행)
    def close_browser():
        try:
            asyncio.run(force_close_browser())
            push_log("🔄 브라우저 종료 완료")
        except Exception as e:
            logger.debug(f"브라우저 종료 오류: {e}")
    threading.Thread(target=close_browser, daemon=True).start()

    # output 폴더 데이터 삭제
    for f in glob.glob("output/*.json"):
        try: os.remove(f)
        except: pass
    img_dir = "output/images"
    if os.path.exists(img_dir):
        shutil.rmtree(img_dir)
        os.makedirs(img_dir, exist_ok=True)

    # scraping/uploading 즉시 False로 → 백그라운드 스레드가 루프 탈출
    status["scraping"]  = False
    status["uploading"] = False
    status["stop_requested"] = False  # 즉시 초기화
    status["paused"] = False

    push_log("✅ 리셋 완료 — 초기 상태로 돌아갔습니다")

    push_log("🔄 리셋 완료 — 모든 데이터가 삭제되고 초기화되었습니다")
    return jsonify({"ok": True, "message": "리셋 완료"})


# ── 실시간 로그 스트리밍 (SSE) ─────────────

@app.route(f"{URL_PREFIX}/logs/stream")
@admin_required
def log_stream():
    """
    Server-Sent Events로 실시간 로그 전송
    멀티 클라이언트 지원 — 데스크탑/태블릿/모바일 모두 동시 수신
    """
    client_queue = _subscribe_logs()

    def generate():
        try:
            while True:
                try:
                    msg = client_queue.get(timeout=30)
                    yield f"data: {json.dumps({'msg': msg})}\n\n"
                except queue.Empty:
                    yield f"data: {json.dumps({'msg': '.'})}\n\n"  # heartbeat
        except GeneratorExit:
            # 클라이언트 연결 끊김 (F5/탭닫기/강제종료)
            logger.info("🔌 SSE 클라이언트 연결 끊김 감지")
            # 다른 SSE 클라이언트가 남아있으면 크롤링 중단하지 않음
            remaining = len(_log_subscribers) - 1  # 현재 끊기는 클라이언트 제외
            if status.get("scraping") and remaining <= 0:
                logger.info("⛔ 마지막 SSE 클라이언트 끊김 — 작업 중단 + 브라우저 정리")
                status["stop_requested"] = True
                status["paused"] = False
                try:
                    asyncio.run(force_close_browser())
                except Exception:
                    pass
                try:
                    from secondst_crawler import force_close_browser as close_2nd
                    asyncio.run(close_2nd())
                except Exception:
                    pass
                status["scraping"] = False
            elif status.get("scraping"):
                logger.info(f"   ℹ️ SSE 클라이언트 끊김이지만 {remaining}개 클라이언트 남아있음 — 크롤링 계속")
        finally:
            _unsubscribe_logs(client_queue)

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# =============================================
# 서버 실행
# =============================================

if __name__ == "__main__":
    # 데이터 폴더 생성 (외부 저장소)
    try:
        ensure_dirs()
    except Exception:
        # 외부 저장소 미연결 시 로컬 fallback
        os.makedirs("output", exist_ok=True)
        os.makedirs("logs", exist_ok=True)

    # 빅데이터 DB 초기화
    try:
        init_product_db()
    except Exception as e:
        print(f"⚠️ 빅데이터 DB 초기화 실패: {e}")

    # 회원 DB 초기화
    try:
        init_user_db()
    except Exception as e:
        print(f"⚠️ 회원 DB 초기화 실패: {e}")

    # 스케줄러 시작 — 1번만 실행 보장
    _start_scheduler_once()
    try:
        _register_fb_schedule_jobs()
    except Exception:
        pass

    # 큐 워커 자동 시작 (예약 작업 복구 포함)
    _start_queue_worker()

    print(f"\n  Xebio Dashboard: http://{SERVER_HOST}:{SERVER_PORT}{URL_PREFIX}\n")

    app.run(
        host=SERVER_HOST,
        port=SERVER_PORT,
        debug=False,
        threaded=True,
        use_reloader=False,      # 스케줄러 중복 방지 (파일 수정 시 수동 재기동)
    )